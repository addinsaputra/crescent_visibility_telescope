#!/usr/bin/env python3
"""
══════════════════════════════════════════════════════════════════════
VALIDASI MODEL CRUMEY (2014) vs DATA OBSERVASI HILAL INDONESIA
══════════════════════════════════════════════════════════════════════

Menjalankan model visibilitas hilal pada seluruh data observasi rukyatul
hilal dari lokasi-lokasi BMKG Indonesia, lalu membandingkan hasil prediksi
model (naked eye & teleskop) dengan data observasi aktual.

Fitur:
  1. Batch processing observasi di banyak lokasi sekaligus
  2. Perhitungan visibilitas naked eye dan teleskop
  3. Perbandingan prediksi model vs observasi (Y/N)
  4. Output Excel dengan format rapi

CARA PAKAI:
  Letakkan script ini di direktori Core/ bersama modul-modul lainnya,
  lalu jalankan:
      python batch_validation_crumey.py

PRASYARAT:
  - Koneksi internet (untuk ERA5 API)
  - Semua modul Core/ sudah terinstall dan berfungsi
  - File de440s.bsp ada di direktori data-hisab/

Author: Pipeline validasi untuk skripsi hilal
Referensi: Crumey, A. (2014), MNRAS 442, 2600-2619
══════════════════════════════════════════════════════════════════════
"""

import os
import sys
import time
import traceback
from typing import List

# Pastikan Core/ di PATH
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from core_crescent_visibility import (
    HilalVisibilityCalculator,
    tentukan_timezone_indonesia,
)


# ═══════════════════════════════════════════════════════════════════
# KONFIGURASI GLOBAL
# ═══════════════════════════════════════════════════════════════════

# --- Field factor referensi ---
F_NAKED_REF = 1.5       # naked-eye reference
FIELD_FACTOR_REF = 1.5   # telescope reference

# --- Parameter teleskop default BMKG ---
TEL_PARAMS = dict(
    aperture=100.0,             # mm (refraktor BMKG tipikal)
    magnification=50.0,         # pembesaran
    transmission=0.95,          # transmisi per permukaan
    n_surfaces=6,               # jumlah permukaan optik
    central_obstruction=0.0,    # refraktor → 0
    observer_age=22.0,          # usia pengamat tipikal
    field_factor=FIELD_FACTOR_REF,
)

# --- Mode perhitungan ---
CALC_MODE = "optimal"       # "sunset" atau "optimal"
SUMBER_ATMOSFER = "era5"    # "era5", "merra2", "manual"

# --- Interval loop optimal ---
INTERVAL_MENIT = 1
MIN_MOON_ALT = 2.0
START_DELAY_MENIT = 1


# ═══════════════════════════════════════════════════════════════════
# DATA OBSERVASI
# ═══════════════════════════════════════════════════════════════════
# Format per entry:
#   no, tanggal (YYYY-MM-DD), nama lokasi,
#   lat, lon, elv, adm4_code,
#   bulan_hijri, tahun_hijri,
#   bias_t, bias_rh,
#   observed (True=Y, False=N)

OBSERVATIONS = [
    # ── Muharram 1444 (29 Juli 2022) ─────────────────────────────
    ( 1, "2022-07-29", "Rooftop Observatorium UIN WS",
      -6.99167, 110.34806, 89.0, "33.74.15.1010",
       1, 1444,  -1, 2,  False),
    ( 2, "2022-07-29", "Tower Hilal Sulamu_BMKG Kupang",
      -10.14333, 123.62667, 10.42, "53.01.07.2001",
       1, 1444,  -1, -2,  True),
    ( 3, "2022-07-29", "Hotel mina tanjung",
      -8.346986, 116.149, 6.0, "52.08.01.2001",
       1, 1444,  -1, 2,  False),
    ( 4, "2022-07-29", "Pantai Lhoknga_chiek kuta_BMKG Aceh Besar",
       5.46667, 95.24233, 11.65, "11.06.02.2001",
       1, 1444,  -1, 4,  False),
    ( 5, "2022-07-29", "POB Condrodipo",
      -7.16967, 112.61733, 61.0, "35.25.14.2002",
       1, 1444,   0, -4,  False),
    ( 6, "2022-07-29", "Tower Hilal Marana_BMKG Palu",
      -0.57861, 119.79070, 17.49, "72.03.10.2007",
       1, 1444,  -4, 10,  False),
    ( 7, "2022-07-29", "Tower Hilal Cikelet",
      -7.59367, 107.62350, 7.0, "32.05.02.2005",
       1, 1444,  -1, 2,  False),
    ( 8, "2022-07-29", "Lapangan Tembak Desa Kebutuhjurang_BMKG Banjarnegara",
      -7.480125, 109.677931, 496.0, "33.04.20.2005",
       1, 1444,   0, 0,  False),
    ( 9, "2022-07-29", "POB Syekh Bela-Belu",
      -7.73983, 110.35017, 45.0, "34.02.04.2005",
       1, 1444,   0, 0,  False),
    (10, "2022-07-29", "Tower Hilal Ternate",
      -0.79983, 127.29483, 33.61, "82.71.01.1005",
       1, 1444,  -1, 2,  False),
 
    # ── Ramadhan 1444 (22 Maret 2023) ────────────────────────────
    (11, "2023-03-22", "Rooftop Observatorium UIN WS",
      -6.99167, 110.34806, 89.0, "33.74.15.1010",
       9, 1444,  -1, 2,  False),
    (12, "2023-03-22", "POB Pedalen Kebumen_BMKG Banjarnegara",
      -7.731322, 109.390878, 9.0, "33.05.01.2001",
       9, 1444,   0, 0,  False),
    (13, "2023-03-22", "Pantai Loang Baloq-MATARAM",
      -8.60408, 116.07440, 5.0, "52.71.04.1002",
       9, 1444,  -1, 2,  True),
    (14, "2023-03-22", "Tower Hilal Marana_BMKG Palu",
      -0.57861, 119.79070, 17.49, "72.03.10.2007",
       9, 1444,  -4, 10,  True),
    (15, "2023-03-22", "Pantai Lhoknga_chiek kuta_BMKG Aceh Besar",
       5.46667, 95.24233, 11.65, "11.06.02.2001",
       9, 1444,  -1, 4,  True),
    (16, "2023-03-22", "Tower Hilal Meras_MTC Manado_BMKG Manado",
       1.48033, 124.83367, 15.14, "71.71.06.1005",
       9, 1444,  -1, 4,  False),
    (17, "2023-03-22", "POB Cibeas Pel. Ratu",
      -7.07400, 106.53133, 114.0, "32.02.02.2003",
       9, 1444,   0, -2,  False),
    (18, "2023-03-22", "POB Syekh Bela-Belu",
      -7.73983, 110.35017, 45.0, "34.02.04.2005",
       9, 1444,   0, 0,  False),
    (19, "2023-03-22", "Tower Hilal Cikelet",
      -7.59367, 107.62350, 7.0, "32.05.02.2005",
       9, 1444,  -1, 2,  False),
    (20, "2023-03-22", "Gedung BMKG NTT-Kupang",
      -10.15278, 123.60833, 40.0, "53.71.06.1007",
       9, 1444,  -1, -2,  False),
 
    # ── Syawal 1445 (9 April 2024) ───────────────────────────────
    (21, "2024-04-09", "Rooftop Observatorium UIN WS",
      -6.99167, 110.34806, 89.0, "33.74.15.1010",
      10, 1445,  -1, 2,  False),
    (22, "2024-04-09", "POB Pedalen Kebumen_BMKG Banjarnegara",
      -7.731322, 109.390878, 9.0, "33.05.01.2001",
      10, 1445,   0, 0,  False),
    (23, "2024-04-09", "Tower Hilal Meras_MTC Manado_BMKG Manado",
       1.48033, 124.83367, 15.14, "71.71.06.1005",
      10, 1445,  -1, 4,  True),
    (24, "2024-04-09", "Pantai Loang Baloq-MATARAM",
      -8.60408, 116.07440, 5.0, "52.71.04.1002",
      10, 1445,  -1, 2,  False),
    (25, "2024-04-09", "Pantai Lhoknga_chiek kuta_BMKG Aceh Besar",
       5.46667, 95.24233, 11.65, "11.06.02.2001",
      10, 1445,  -1, 4,  False),
    (26, "2024-04-09", "Gedung BMKG NTT-Kupang",
      -10.15278, 123.60833, 40.0, "53.71.06.1007",
      10, 1445,  -1, -2,  False),
    (27, "2024-04-09", "Tower Hilal Ternate",
      -0.79983, 127.29483, 33.61, "82.71.01.1005",
      10, 1445,  -1, 2,  False),
    (28, "2024-04-09", "Pantai Ngliyep Malang_BMKG Malang",
      -8.35000, 112.43333, 10.0, "35.07.15.2001",
      10, 1445,   0, 0,  False),
    (29, "2024-04-09", "Tower Hilal Marana_BMKG Palu",
      -0.57861, 119.79070, 17.49, "72.03.10.2007",
      10, 1445,  -4, 10,  False),
    (30, "2024-04-09", "kantor stageof lampung utara_BMKG Lampung Utara",
       4.836136, 104.87005, 33.0, "18.03.07.1007",
      10, 1445,  -1, 6,  False),
    (31, "2024-04-09", "POB Syekh Bela-Belu",
      -7.73983, 110.35017, 45.0, "34.02.04.2005",
      10, 1445,   0, 0,  False),
]


# ═══════════════════════════════════════════════════════════════════
# HELPER
# ═══════════════════════════════════════════════════════════════════

N_OBS = len(OBSERVATIONS)


def parse_obs(entry: tuple) -> dict:
    """Parse satu entry tuple menjadi dictionary."""
    return {
        'no': entry[0],
        'tanggal': entry[1],
        'nama': entry[2],
        'lat': entry[3],
        'lon': entry[4],
        'elv': entry[5],
        'adm4': entry[6],
        'bulan_hijri': entry[7],
        'tahun_hijri': entry[8],
        'bias_t': entry[9],
        'bias_rh': entry[10],
        'observed': entry[11],  # True = Y, False = N
    }


# ═══════════════════════════════════════════════════════════════════
# BATCH PROCESSING
# ═══════════════════════════════════════════════════════════════════

def run_single_observation(obs: dict, verbose: bool = True) -> dict:
    """Jalankan model untuk satu observasi.

    Returns
    -------
    dict : berisi semua hasil + metadata observasi
    """
    no = obs['no']
    nama = obs['nama']
    tanggal = obs['tanggal']
    tz = tentukan_timezone_indonesia(obs['lon'])

    if verbose:
        print(f"\n{'═' * 70}")
        print(f"[{no:2d}/{N_OBS}] {nama}")
        print(f"        Tanggal: {tanggal}  |  Hijri: {obs['bulan_hijri']}/{obs['tahun_hijri']}")
        print(f"        Lat={obs['lat']:.4f}  Lon={obs['lon']:.4f}  Elv={obs['elv']:.0f}m")
        print(f"        Observasi: {'TERLIHAT (Y)' if obs['observed'] else 'TIDAK TERLIHAT (N)'}")
        print(f"{'═' * 70}")

    try:
        calc = HilalVisibilityCalculator(
            nama_tempat=nama,
            lintang=obs['lat'],
            bujur=obs['lon'],
            elevasi=obs['elv'],
            timezone_str=tz,
            bulan_hijri=obs['bulan_hijri'],
            tahun_hijri=obs['tahun_hijri'],
            delta_day_offset=0,
            bias_t=obs['bias_t'],
            bias_rh=obs['bias_rh'],
            sumber_atmosfer=SUMBER_ATMOSFER,
            adm4_code=obs['adm4'],
        )

        hasil = calc.jalankan_perhitungan_lengkap(
            use_telescope=True,
            mode=CALC_MODE,
            F_naked=F_NAKED_REF,
            interval_menit=INTERVAL_MENIT,
            min_moon_alt=MIN_MOON_ALT,
            start_delay_menit=START_DELAY_MENIT,
            **TEL_PARAMS,
        )

        # Tanggal pengamatan yang dihitung model
        tgl_model = hasil.get('tanggal_pengamatan')
        if tgl_model:
            tgl_model_str = tgl_model.strftime('%Y-%m-%d')
        else:
            tgl_model_str = "?"

        # Verifikasi tanggal
        tgl_cocok = (tgl_model_str == tanggal)
        if not tgl_cocok and verbose:
            print(f"\n  ⚠ TANGGAL TIDAK COCOK: model={tgl_model_str}, "
                  f"observasi={tanggal}")

        # Kumpulkan hasil
        result = {
            # Metadata observasi
            'no': no,
            'nama': nama,
            'tanggal_obs': tanggal,
            'tanggal_model': tgl_model_str,
            'tanggal_cocok': tgl_cocok,
            'bulan_hijri': obs['bulan_hijri'],
            'tahun_hijri': obs['tahun_hijri'],
            'lat': obs['lat'],
            'lon': obs['lon'],
            'elv': obs['elv'],
            'observed': obs['observed'],
            'success': True,

            # Hasil saat sunset
            'sunset_local': str(hasil.get('sunset_local', '')),
            'moon_alt_sunset': hasil.get('moon_alt', 0),
            'sun_alt_sunset': hasil.get('sun_alt', 0),
            'elongation': hasil.get('elongation', 0),
            'moon_width': hasil.get('moon_width', 0),
            'phase_angle': hasil.get('phase_angle', 0),
            'sky_brightness_nl': hasil.get('sky_brightness_nl', 0),
            'luminansi_hilal_nl': hasil.get('luminansi_hilal_nl', 0),
            'k_v': hasil.get('k_v', 0),
            'rh': hasil.get('rh', 0),
            'temperature': hasil.get('temperature', 0),
            'delta_m_ne_sunset': hasil.get('delta_m_ne', -99.0),
            'delta_m_tel_sunset': hasil.get('delta_m_tel', -99.0),
            'telescope_gain_sunset': hasil.get('telescope_gain', 0),
        }

        # Hasil optimal (jika mode optimal)
        if CALC_MODE == "optimal":
            # Ambil data astronomis dari result dict optimal
            opt_ne = hasil.get('optimal_result_ne') or {}
            opt_tel = hasil.get('optimal_result_tel') or {}

            result.update({
                'delta_m_ne_opt': hasil.get('optimal_delta_m_ne', -99.0),
                'delta_m_tel_opt': hasil.get('optimal_delta_m_tel', -99.0),
                'optimal_time_ne': str(hasil.get('optimal_time_ne', '')),
                'optimal_time_tel': str(hasil.get('optimal_time_tel', '')),
                'optimal_moon_alt_ne': hasil.get('optimal_moon_alt_ne', 0),
                'optimal_moon_alt_tel': hasil.get('optimal_moon_alt_tel', 0),
                'telescope_gain_opt': hasil.get('optimal_telescope_gain', 0),
                'vis_duration_ne': hasil.get('visibility_duration_ne', 0),
                'vis_duration_tel': hasil.get('visibility_duration_tel', 0),
                # Data astronomis optimal NE
                'opt_ne_elongation': opt_ne.get('elongation', 0),
                'opt_ne_sky_brightness_nl': opt_ne.get('sky_brightness_nl', 0),
                'opt_ne_luminansi_hilal_nl': opt_ne.get('luminansi_hilal_nl', 0),
                'opt_ne_k_v': opt_ne.get('k_v', 0),
                'opt_ne_rh': opt_ne.get('rh', 0),
                'opt_ne_temperature': opt_ne.get('temperature', 0),
                # Data astronomis optimal Teleskop
                'opt_tel_elongation': opt_tel.get('elongation', 0),
                'opt_tel_sky_brightness_nl': opt_tel.get('sky_brightness_nl', 0),
                'opt_tel_luminansi_hilal_nl': opt_tel.get('luminansi_hilal_nl', 0),
                'opt_tel_k_v': opt_tel.get('k_v', 0),
                'opt_tel_rh': opt_tel.get('rh', 0),
                'opt_tel_temperature': opt_tel.get('temperature', 0),
            })
        else:
            result.update({
                'delta_m_ne_opt': hasil.get('delta_m_ne', -99.0),
                'delta_m_tel_opt': hasil.get('delta_m_tel', -99.0),
                'optimal_time_ne': '',
                'optimal_time_tel': '',
                'optimal_moon_alt_ne': 0,
                'optimal_moon_alt_tel': 0,
                'telescope_gain_opt': 0,
                'vis_duration_ne': 0,
                'vis_duration_tel': 0,
                'opt_ne_elongation': 0, 'opt_ne_sky_brightness_nl': 0,
                'opt_ne_luminansi_hilal_nl': 0, 'opt_ne_k_v': 0,
                'opt_ne_rh': 0, 'opt_ne_temperature': 0,
                'opt_tel_elongation': 0, 'opt_tel_sky_brightness_nl': 0,
                'opt_tel_luminansi_hilal_nl': 0, 'opt_tel_k_v': 0,
                'opt_tel_rh': 0, 'opt_tel_temperature': 0,
            })

        # Ringkasan cepat
        dm_ne = result['delta_m_ne_opt']
        dm_tel = result['delta_m_tel_opt']
        pred_ne = "Y" if dm_ne > 0 else "N"
        pred_tel = "Y" if dm_tel > 0 else "N"
        obs_str = "Y" if obs['observed'] else "N"
        match_tel = "✓" if (dm_tel > 0) == obs['observed'] else "✗"

        if verbose:
            print(f"\n  RINGKASAN:")
            print(f"    Naked Eye : Δm={dm_ne:+.3f}  Pred={pred_ne}")
            print(f"    Teleskop  : Δm={dm_tel:+.3f}  Pred={pred_tel}  Obs={obs_str}  {match_tel}")

        return result

    except Exception as e:
        print(f"\n  ✗ ERROR pada obs #{no}: {e}")
        traceback.print_exc()
        return {
            'no': no,
            'nama': nama,
            'tanggal_obs': tanggal,
            'tanggal_model': '?',
            'tanggal_cocok': False,
            'bulan_hijri': obs['bulan_hijri'],
            'tahun_hijri': obs['tahun_hijri'],
            'lat': obs['lat'],
            'lon': obs['lon'],
            'elv': obs['elv'],
            'observed': obs['observed'],
            'success': False,
            'delta_m_ne_sunset': -99.0,
            'delta_m_tel_sunset': -99.0,
            'delta_m_ne_opt': -99.0,
            'delta_m_tel_opt': -99.0,
            'error': str(e),
        }


def run_batch(bias_mode: str = '1', manual_bias_t: float = 0.0, manual_bias_rh: float = 0.0) -> List[dict]:
    """Jalankan model untuk semua observasi."""
    print("\n" + "█" * 70)
    print("  BATCH VALIDATION: Model Crumey (2014) vs Observasi Hilal")
    print(f"  {N_OBS} data observasi × 4 event rukyat (2022–2024)")
    print(f"  Mode: {CALC_MODE}  |  Atmosfer: {SUMBER_ATMOSFER}")
    print(f"  F_naked_ref={F_NAKED_REF}  |  F_tel_ref={FIELD_FACTOR_REF}")
    if bias_mode == '2':
        print("  Koreksi Bias: Tanpa koreksi (bias = 0)")
    elif bias_mode == '3':
        print(f"  Koreksi Bias: Manual Seragam (T={manual_bias_t:+.1f}°C, RH={manual_bias_rh:+.1f}%)")
    else:
        print("  Koreksi Bias: Data Bawaan Lokasi")
    print("█" * 70)

    results = []
    t_start = time.time()

    for seq_no, entry in enumerate(OBSERVATIONS, start=1):
        obs = parse_obs(entry)
        obs['no'] = seq_no
        
        # Terapkan opsi bias
        if bias_mode == '2':
            obs['bias_t'] = 0.0
            obs['bias_rh'] = 0.0
        elif bias_mode == '3':
            obs['bias_t'] = manual_bias_t
            obs['bias_rh'] = manual_bias_rh

        t_obs_start = time.time()
        result = run_single_observation(obs)
        elapsed = time.time() - t_obs_start
        result['elapsed_sec'] = elapsed
        results.append(result)
        print(f"  ⏱ Waktu: {elapsed:.1f} detik")

    total_time = time.time() - t_start
    n_success = sum(1 for r in results if r.get('success', False))
    print(f"\n{'═' * 70}")
    print(f"BATCH SELESAI: {n_success}/{N_OBS} berhasil dalam {total_time:.0f} detik")
    print(f"{'═' * 70}")

    return results


# ═══════════════════════════════════════════════════════════════════
# TABEL RINGKASAN
# ═══════════════════════════════════════════════════════════════════

def print_results_table(results: List[dict]):
    """Cetak tabel ringkasan hasil."""
    print(f"\n{'═' * 120}")
    print(f"TABEL HASIL OBSERVASI (F_naked={F_NAKED_REF:.1f}, F_tel={FIELD_FACTOR_REF:.1f})")
    print(f"{'═' * 120}")
    hdr = (f"{'No':>3} {'Tanggal':>10} {'Lokasi':<35} "
           f"{'Obs':>3} {'Moon°':>6} {'Elong°':>6} {'Lebar':>6} "
           f"{'Δm_NE':>8} {'Pr_NE':>5} "
           f"{'Δm_Tel':>8} {'Pr_Tel':>6} {'Match':>5}")
    print(hdr)
    print("─" * 120)

    for r in results:
        if not r.get('success'):
            print(f"{r['no']:>3} {r['tanggal_obs']:>10} {r['nama']:<35} "
                  f"{'?' :>3} {'?':>6} {'?':>6} {'?':>6} {'ERROR':>8} {'?':>5} "
                  f"{'ERROR':>8} {'?':>6} {'?':>5}")
            continue

        obs_str = "Y" if r['observed'] else "N"
        dm_ne = r.get('delta_m_ne_opt', -99)
        dm_tel = r.get('delta_m_tel_opt', -99)
        pred_ne = "Y" if dm_ne > 0 else "N"
        pred_tel = "Y" if dm_tel > 0 else "N"
        # Cocok jika prediksi teleskop sesuai observasi
        match = "✓" if pred_tel == obs_str else "✗"
        moon_alt = r.get('moon_alt_sunset', 0)
        elong = r.get('elongation', 0)
        moon_width_arcmin = r.get('moon_width', 0) * 60.0

        print(f"{r['no']:>3} {r['tanggal_obs']:>10} {r['nama']:<35} "
              f"{obs_str:>3} {moon_alt:>6.2f} {elong:>6.2f} {moon_width_arcmin:>6.2f} "
              f"{dm_ne:>+8.3f} {pred_ne:>5} "
              f"{dm_tel:>+8.3f} {pred_tel:>6} {match:>5}")

    print("─" * 120)

    # Ringkasan kecocokan (observasi = data teleskop)
    valid = [r for r in results if r.get('success', False)]
    if valid:
        n_match_tel = sum(1 for r in valid if (r.get('delta_m_tel_opt', -99) > 0) == r['observed'])
        print(f"\n  Kecocokan Teleskop vs Observasi : {n_match_tel}/{len(valid)} ({n_match_tel/len(valid):.1%})")


# ═══════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════════

def save_to_excel(results: List[dict], filepath: str, bias_mode_str: str = "Data Bawaan Lokasi"):
    """Simpan semua hasil ke file Excel dengan format rapi 2-baris header."""
    from datetime import time as dt_time
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.styles.colors import Color
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Styles ──
    thin_border = Border(
        left=Side('thin'), right=Side('thin'),
        top=Side('thin'), bottom=Side('thin'))
    thin_border_no_top = Border(
        left=Side('thin'), right=Side('thin'),
        bottom=Side('thin'))

    hdr_font = Font(name='Times New Roman', bold=True, size=11,
                    color=Color(theme=1))
    data_font = Font(name='Times New Roman', size=10)
    center = Alignment(horizontal='center', vertical='center')
    center_no_v = Alignment(horizontal='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # Header fills (warna grup)
    meta_fill = PatternFill('solid', fgColor='0070C0')
    sunset_fill = PatternFill('solid',
                              fgColor=Color(theme=9, tint=-0.249977111117893))
    opt_ne_fill = PatternFill('solid',
                              fgColor=Color(theme=7, tint=-0.249977111117893))
    tel_fill = PatternFill('solid',
                           fgColor=Color(theme=6, tint=-0.249977111117893))

    # Data fills (warna muda untuk kolom Δm & prediksi)
    sunset_data_fill = PatternFill('solid',
                                   fgColor=Color(theme=9, tint=0.3999755851924192))
    opt_ne_data_fill = PatternFill('solid',
                                   fgColor=Color(theme=7, tint=0.3999755851924192))
    tel_data_fill = PatternFill('solid',
                                fgColor=Color(theme=6, tint=0.3999755851924192))

    green_fill = PatternFill('solid', fgColor='C6EFCE')
    red_fill = PatternFill('solid', fgColor='FFC7CE')

    # Ringkasan styles (tetap Arial)
    ringkasan_hdr_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    ringkasan_hdr_fill = PatternFill('solid', fgColor='1F4E79')
    ringkasan_data_font = Font(name='Arial', size=10)
    ringkasan_border = Border(
        left=Side('thin', 'B0B0B0'), right=Side('thin', 'B0B0B0'),
        top=Side('thin', 'B0B0B0'), bottom=Side('thin', 'B0B0B0'))

    # ═══ Sheet 1: Hasil Observasi ═══
    ws = wb.active
    ws.title = "Hasil Observasi"

    # ── Row 1: Header grup (merged) + metadata kolom (merged 2 baris) ──
    meta_headers = ['No', 'Tanggal', 'Lokasi', 'Lat', 'Lon', 'Elv',
                    'Bulan Hijri', 'Time Zone']
    for ci, h in enumerate(meta_headers, 1):
        col_letter = get_column_letter(ci)
        ws.merge_cells(f'{col_letter}1:{col_letter}2')
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font
        c.fill = meta_fill
        c.alignment = center
        c.border = thin_border

    # Grup: Sunset (I1:V1)
    ws.merge_cells('I1:V1')
    c = ws.cell(row=1, column=9,
                value='DATA VISIBILITAS HILAL NAKED EYE DAN TELESKOP SAAT SUNSET')
    c.font = hdr_font; c.fill = sunset_fill
    c.alignment = center_no_v; c.border = thin_border

    # Grup: Optimal NE (W1:AF1)
    ws.merge_cells('W1:AF1')
    c = ws.cell(row=1, column=23,
                value='DATA VISIBILITAS HILAL NAKED EYE WAKTU OPTIMAL ATAU BEST TIME')
    c.font = hdr_font; c.fill = opt_ne_fill
    c.alignment = center_no_v; c.border = thin_border

    # Grup: Teleskop (AG1:AT1)
    ws.merge_cells('AG1:AT1')
    c = ws.cell(row=1, column=33,
                value='DATA VISIBILITAS HILAL BERBANTUAN TELESKOP BEST TIME')
    c.font = hdr_font; c.fill = tel_fill
    c.alignment = center_no_v; c.border = thin_border

    # ── Row 2: Sub-header kolom per grup ──
    # Sunset (I-V)
    sunset_hdrs = {
        9: 'Sunset Lokal', 10: 'Moon Alt (\u00b0)', 11: 'Elongasi (\u00b0)',
        12: 'Lebar Sabit (arcmin)', 13: 'Phase Angle (\u00b0)',
        14: 'Sky Bright (nL)', 15: 'Lum Hilal (nL)',
        16: 'k_V', 17: 'RH (%)', 18: 'T (\u00b0C)',
    }
    for ci, h in sunset_hdrs.items():
        c = ws.cell(row=2, column=ci, value=h)
        c.font = hdr_font; c.fill = sunset_fill
        c.alignment = center; c.border = thin_border

    # Δm NE sunset (S2:T2)
    ws.merge_cells('S2:T2')
    c = ws.cell(row=2, column=19, value='\u0394m NE')
    c.font = hdr_font; c.fill = sunset_fill
    c.alignment = center; c.border = thin_border

    # Δm Tel sunset (U2:V2)
    ws.merge_cells('U2:V2')
    c = ws.cell(row=2, column=21, value='\u0394m Tel ')
    c.font = hdr_font; c.fill = sunset_fill
    c.alignment = center; c.border = thin_border

    # Optimal NE (W-AF)
    opt_ne_hdrs = {
        23: 'Best Time NE', 24: 'Moon Alt (\u00b0)', 25: 'Elongasi (\u00b0)',
        26: 'Sky Bright (nL)', 27: 'Lum Hilal (nL)',
        28: 'k_V', 29: 'RH (%)', 30: 'T (\u00b0C)',
    }
    for ci, h in opt_ne_hdrs.items():
        c = ws.cell(row=2, column=ci, value=h)
        c.font = hdr_font; c.fill = opt_ne_fill
        c.alignment = center; c.border = thin_border

    # Δm NE optimal (AE2:AF2)
    ws.merge_cells('AE2:AF2')
    c = ws.cell(row=2, column=31, value='\u0394m NE')
    c.font = hdr_font; c.fill = opt_ne_fill
    c.alignment = center; c.border = thin_border

    # Teleskop (AG-AT)
    tel_hdrs = {
        33: 'Best Time Tel', 34: 'Moon Alt (\u00b0)', 35: 'Elongasi (\u00b0)',
        36: 'Sky Bright (nL)', 37: 'Lum Hilal (nL)',
        38: 'k_V', 39: 'RH (%)', 40: 'T (\u00b0C)',
        41: 'Tel Gain Opt', 42: 'Leg time (mnt)',
        45: 'Observasi Tel', 46: 'Correct',
    }
    for ci, h in tel_hdrs.items():
        c = ws.cell(row=2, column=ci, value=h)
        c.font = hdr_font; c.fill = tel_fill
        c.alignment = center; c.border = thin_border

    # Δm Tel teleskop (AQ2:AR2)
    ws.merge_cells('AQ2:AR2')
    c = ws.cell(row=2, column=43, value='\u0394m Tel')
    c.font = hdr_font; c.fill = tel_fill
    c.alignment = center; c.border = thin_border

    # ── Helper functions ──
    def _parse_time(time_str):
        """Parse string waktu ke datetime.time."""
        if not time_str:
            return None
        try:
            s = str(time_str).strip()
            if ' ' in s:
                time_part = s.split(' ')[1].split('+')[0].split('-')[0]
            else:
                time_part = s
            parts = time_part.split(':')
            return dt_time(int(parts[0]), int(parts[1]),
                           int(parts[2]) if len(parts) > 2 else 0)
        except Exception:
            return None

    def _tz_offset(lon):
        """UTC offset zona waktu Indonesia dari bujur."""
        if lon < 115:
            return 7   # WIB
        elif lon < 135:
            return 8   # WITA
        return 9       # WIT

    # ── Data rows (mulai baris 3) ──
    # Kolom yang diberi warna Δm data
    SUNSET_DM_COLS = {19, 20, 21, 22}
    OPT_NE_DM_COLS = {31, 32}
    TEL_DM_COLS = {43, 44, 45}

    for i, r in enumerate(results, 3):
        obs_str = "Y" if r['observed'] else "N"
        dm_ne_sun = r.get('delta_m_ne_sunset', -99)
        dm_tel_sun = r.get('delta_m_tel_sunset', -99)
        dm_ne_opt = r.get('delta_m_ne_opt', -99)
        dm_tel_opt = r.get('delta_m_tel_opt', -99)
        pred_ne_sun = "Y" if dm_ne_sun > 0 else "N"
        pred_tel_sun = "Y" if dm_tel_sun > 0 else "N"
        pred_ne_opt = "Y" if dm_ne_opt > 0 else "N"
        pred_tel_opt = "Y" if dm_tel_opt > 0 else "N"
        cocok = "\u2713" if pred_tel_opt == obs_str else "\u2717"

        row_data = {
            # Metadata (A-H)
            1: r['no'],
            2: r['tanggal_obs'],
            3: r['nama'],
            4: r.get('lat', 0),
            5: r.get('lon', 0),
            6: r.get('elv', 0),
            7: f"{r.get('bulan_hijri', 0)}/{r.get('tahun_hijri', 0)}",
            8: _tz_offset(r.get('lon', 0)),
            # Sunset (I-V)
            9: _parse_time(r.get('sunset_local', '')),
            10: round(r.get('moon_alt_sunset', 0), 4),
            11: round(r.get('elongation', 0), 4),
            12: round(r.get('moon_width', 0) * 60.0, 4),
            13: round(r.get('phase_angle', 0), 4),
            14: f"{r.get('sky_brightness_nl', 0):.4e}",
            15: f"{r.get('luminansi_hilal_nl', 0):.4e}",
            16: round(r.get('k_v', 0), 4),
            17: round(r.get('rh', 0), 2),
            18: round(r.get('temperature', 0), 2),
            19: round(dm_ne_sun, 4),
            20: pred_ne_sun,
            21: round(dm_tel_sun, 4),
            22: pred_tel_sun,
            # Optimal NE (W-AF)
            23: _parse_time(r.get('optimal_time_ne', '')),
            24: round(r.get('optimal_moon_alt_ne', 0), 4),
            25: round(r.get('opt_ne_elongation', 0), 4),
            26: f"{r.get('opt_ne_sky_brightness_nl', 0):.4e}",
            27: f"{r.get('opt_ne_luminansi_hilal_nl', 0):.4e}",
            28: round(r.get('opt_ne_k_v', 0), 4),
            29: round(r.get('opt_ne_rh', 0), 2),
            30: round(r.get('opt_ne_temperature', 0), 2),
            31: round(dm_ne_opt, 4),
            32: pred_ne_opt,
            # Teleskop (AG-AT)
            33: _parse_time(r.get('optimal_time_tel', '')),
            34: round(r.get('optimal_moon_alt_tel', 0), 4),
            35: round(r.get('opt_tel_elongation', 0), 4),
            36: f"{r.get('opt_tel_sky_brightness_nl', 0):.4e}",
            37: f"{r.get('opt_tel_luminansi_hilal_nl', 0):.4e}",
            38: round(r.get('opt_tel_k_v', 0), 4),
            39: round(r.get('opt_tel_rh', 0), 2),
            40: round(r.get('opt_tel_temperature', 0), 2),
            41: round(r.get('telescope_gain_opt', 0), 4),
            42: int(round(r.get('vis_duration_tel', 0))),
            43: round(dm_tel_opt, 4),
            44: pred_tel_opt,
            45: obs_str,
            46: cocok,
        }

        for ci, val in row_data.items():
            c = ws.cell(row=i, column=ci, value=val)
            c.font = data_font
            c.alignment = left_align if ci == 3 else center
            c.border = thin_border_no_top

            # Format waktu
            if ci in (9, 23, 33) and isinstance(val, dt_time):
                c.number_format = 'h:mm:ss'

            # Warna Δm data
            if ci in SUNSET_DM_COLS:
                c.fill = sunset_data_fill
            elif ci in OPT_NE_DM_COLS:
                c.fill = opt_ne_data_fill
            elif ci in TEL_DM_COLS:
                c.fill = tel_data_fill

            # Kolom Correct
            if ci == 46:
                c.fill = green_fill if cocok == "\u2713" else red_fill

    # ── Column widths ──
    col_widths = {
        'A': 4, 'B': 12, 'C': 30, 'D': 11, 'E': 12, 'F': 7,
        'G': 13, 'H': 7.57,
        'I': 14.29, 'J': 12.71, 'K': 14, 'L': 17, 'M': 17, 'N': 13,
        'O': 16, 'P': 8, 'Q': 13, 'R': 13, 'S': 9.29, 'T': 4,
        'U': 9.29, 'V': 4.43,
        'W': 14.57, 'X': 12.86, 'Y': 11.86, 'Z': 15.71, 'AA': 15,
        'AB': 8.29, 'AC': 9.14, 'AD': 7.57, 'AE': 9, 'AF': 3.86,
        'AG': 14.71, 'AH': 12.86, 'AI': 11.86, 'AJ': 15.71,
        'AK': 15, 'AL': 7.29, 'AM': 9, 'AN': 7.43, 'AO': 13.29,
        'AP': 15.14, 'AQ': 8.43, 'AR': 4.29, 'AS': 14.43, 'AT': 8.29,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = 'A3'

    # ═══ Sheet 2: Ringkasan ═══
    ws2 = wb.create_sheet("Ringkasan")
    ws2.sheet_properties.tabColor = 'BF8F00'
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 25

    valid = [r for r in results if r.get('success', False)]
    n_match_tel = sum(1 for r in valid
                      if (r.get('delta_m_tel_opt', -99) > 0) == r['observed'])

    summary_data = [
        ("KONFIGURASI", ""),
        ("Mode Perhitungan", CALC_MODE),
        ("Sumber Atmosfer", SUMBER_ATMOSFER),
        ("Koreksi Bias", bias_mode_str),
        ("F Naked Eye", F_NAKED_REF),
        ("F Teleskop", FIELD_FACTOR_REF),
        ("Aperture Teleskop (mm)", TEL_PARAMS['aperture']),
        ("Magnifikasi", TEL_PARAMS['magnification']),
        ("Transmisi/permukaan", TEL_PARAMS['transmission']),
        ("Jumlah permukaan", TEL_PARAMS['n_surfaces']),
        ("", ""),
        ("HASIL VALIDASI", ""),
        ("Total Observasi", len(results)),
        ("Observasi Berhasil", len(valid)),
        ("Observasi Terlihat (Y)", sum(1 for r in results if r['observed'])),
        ("Observasi Tidak Terlihat (N)",
         sum(1 for r in results if not r['observed'])),
        ("", ""),
        ("KECOCOKAN PREDIKSI TELESKOP vs OBSERVASI", ""),
        ("Kecocokan Teleskop",
         f"{n_match_tel}/{len(valid)} ({n_match_tel/len(valid):.1%})"
         if valid else "N/A"),
    ]

    for i, (label, val) in enumerate(summary_data, 1):
        cl = ws2.cell(row=i, column=1, value=label)
        cv = ws2.cell(row=i, column=2, value=val)
        if label and not val and val != 0:
            cl.font = ringkasan_hdr_font
            cl.fill = ringkasan_hdr_fill
            cv.fill = ringkasan_hdr_fill
        else:
            cl.font = ringkasan_data_font
            cv.font = ringkasan_data_font
        cl.border = ringkasan_border
        cv.border = ringkasan_border

    # Save
    out_dir = os.path.dirname(filepath)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(filepath)
    print(f"\n  \u2713 Excel disimpan: {filepath}")


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════

def _input_koreksi_bias_batch() -> tuple:
    print("\n--- KOREKSI BIAS ---")
    print("  Pilih opsi koreksi bias untuk semua lokasi:")
    print("  1. Gunakan data bawaan lokasi (jika ada)")
    print("  2. Tanpa koreksi (bias_t = 0, bias_rh = 0)")
    print("  3. Input manual nilai bias seragam untuk semua lokasi")
    try:
        pilihan = input("\n  Pilih opsi (1/2/3) [enter=1]: ").strip() or "1"
        if pilihan == "2":
            print("  ✓ Menggunakan data tanpa koreksi")
            return "2", 0.0, 0.0, "Tanpa koreksi (bias = 0)"
        elif pilihan == "3":
            t_str = input("  Masukkan bias suhu (°C) [enter=0]: ").strip()
            bias_t = float(t_str) if t_str else 0.0
            rh_str = input("  Masukkan bias RH (%) [enter=0]: ").strip()
            bias_rh = float(rh_str) if rh_str else 0.0
            print(f"  ✓ Koreksi bias seragam: T={bias_t:+.1f}°C, RH={bias_rh:+.1f}%")
            return "3", bias_t, bias_rh, f"Manual Seragam (T={bias_t:+.1f}°C, RH={bias_rh:+.1f}%)"
        else:
            print("  ✓ Menggunakan data bawaan lokasi")
            return "1", 0.0, 0.0, "Data Bawaan Lokasi"
    except EOFError:
        return "1", 0.0, 0.0, "Data Bawaan Lokasi"
    except ValueError:
        print("  [!] Input tidak valid, menggunakan bawaan lokasi.")
        return "1", 0.0, 0.0, "Data Bawaan Lokasi"

def main():
    """Entry point utama."""
    # 0. Prompt Bias
    bias_mode, manual_bias_t, manual_bias_rh, bias_mode_str = _input_koreksi_bias_batch()

    # 1. Batch run
    results = run_batch(bias_mode, manual_bias_t, manual_bias_rh)

    # 2. Tampilkan tabel
    print_results_table(results)

    # 3. Simpan ke Excel
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(script_dir, 'output')

    bias_tag = "NoBias" if bias_mode == '2' else ("ManualBias" if bias_mode == '3' else "BiasBawaan")
    excel_path = os.path.join(output_dir,
        f"Validasi_Crumey_{SUMBER_ATMOSFER}_{CALC_MODE}_{bias_tag}.xlsx")
    save_to_excel(results, excel_path, bias_mode_str)

    # 4. Ringkasan akhir
    print(f"\n{'█' * 70}")
    print("  VALIDASI SELESAI")
    print(f"{'█' * 70}")
    print(f"  Excel : {excel_path}")
    print(f"{'█' * 70}\n")

    return results


if __name__ == "__main__":
    results = main()