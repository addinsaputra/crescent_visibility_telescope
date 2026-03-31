#!/usr/bin/env python3
"""
Script test untuk memverifikasi perbaikan kolom waktu pada output Excel.
"""

import os
import sys

# Pastikan Core/ di PATH
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from batch_validation_crumey import (
    OBSERVATIONS, parse_obs, run_single_observation,
    save_to_excel, CALC_MODE, SUMBER_ATMOSFER,
    F_NAKED_REF, FIELD_FACTOR_REF, TEL_PARAMS
)

def main_test():
    """Jalankan test sederhana dengan beberapa observasi."""
    print("=" * 70)
    print("TEST PERBAIKAN KOLOM WAKTU EXCEL")
    print("=" * 70)
    
    # Gunakan mode optimal untuk melihat kolom Best Time
    CALC_MODE = "optimal"
    print(f"Mode: {CALC_MODE}")
    
    # Ambil 3 observasi pertama untuk test
    test_observations = OBSERVATIONS[:3]
    results = []
    
    for i, entry in enumerate(test_observations, 1):
        obs = parse_obs(entry)
        obs['no'] = i
        print(f"\n[{i}/{len(test_observations)}] Memproses: {obs['nama']}")
        
        # Jalankan perhitungan
        result = run_single_observation(obs, verbose=True)
        results.append(result)
        
        # Cek hasil waktu
        print(f"  Sunset Lokal    : {result.get('sunset_local')}")
        print(f"  Best Time NE    : {result.get('optimal_time_ne')}")
        print(f"  Best Time Tel   : {result.get('optimal_time_tel')}")
    
    # Simpan ke Excel
    output_dir = os.path.join(os.path.dirname(__file__), 'output')
    os.makedirs(output_dir, exist_ok=True)
    excel_path = os.path.join(output_dir, 'test_perbaikan_waktu.xlsx')
    
    save_to_excel(results, excel_path, "Test Mode")
    
    print(f"\n✓ Test selesai! Output: {excel_path}")
    
    # Verifikasi dengan openpyxl
    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    ws = wb.active
    
    print("\n" + "=" * 70)
    print("VERIFIKASI EXCEL:")
    print("=" * 70)
    
    for i in range(3, 3 + len(results) + 1):
        sunset_val = ws.cell(row=i, column=9).value
        opt_ne_val = ws.cell(row=i, column=23).value
        opt_tel_val = ws.cell(row=i, column=33).value
        
        print(f"\nRow {i}:")
        print(f"  Col I (Sunset Lokal)  : {sunset_val} (type: {type(sunset_val).__name__})")
        print(f"  Col W (Best Time NE)  : {opt_ne_val} (type: {type(opt_ne_val).__name__})")
        print(f"  Col AG (Best Time Tel): {opt_tel_val} (type: {type(opt_tel_val).__name__})")
    
    print("\n" + "=" * 70)
    print("Jika kolom waktu masih kosong, periksa:")
    print("1. Apakah data dari HilalVisibilityCalculator mengembalikan objek datetime?")
    print("2. Apakah fungsi _parse_time sudah diperbaiki?")
    print("=" * 70)

if __name__ == "__main__":
    main_test()
