import math
from datetime import datetime, timezone, timedelta
from typing import Dict, Any, Optional, Tuple
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

# Import modul yang diperlukan
from visual_limit_schaefer import hitung_sky_brightness
from visual_limit_kastner import hitung_luminansi_kastner
# Import langsung dari modul Crumey (tanpa intermediary crumey_telescope_correction.py)
from full_rumus_crumey import (
    hilal_naked_eye_visibility,
    nL_to_cd_m2,
    cd_m2_to_nL,
    arcmin2_to_sr,
    contrast_threshold,
    crescent_area_arcmin2,
)
from telescope_limit import TelescopeVisibilityModel

# Import modul cuaca ERA5 (Open-Meteo)
from atmosfer_era5 import (
    ObservingLocation, get_rh_t_at_time as era5_get_rh_t,
    ERA5APIError, apply_bias_correction
)

# Import modul cuaca MERRA-2 (NASA POWER)
from atmosfer_merra2 import (
    ObservingLocation as MERRA2Location,
    get_rh_t_at_time as merra2_get_rh_t,
    PowerAPIError
)

# Import modul cuaca BMKG (Prakiraan Cuaca)
from atmosfer_bmkg import (
    ObservingLocation as BMKGLocation,
    get_rh_t_at_time_local as bmkg_get_rh_t_local,
    BMKGAPIError
)

# Import modul data_hisab menggantikan sunmoon
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'data-hisab'))
from data_hisab import (
    newmoon_hijri_month_utc,
    convert_utc_to_localtime,
    convert_localtime_to_utc,
    sunrise_sunset_utc,
    sunrise_sunset_local,
    sun_position_time_utc,
    sun_position_time_local,
    moon_position_time_utc,
    moon_position_time_local,
    moon_elongation_time_utc,
    moon_elongation_time_local,
    moon_phase_angle_time_utc,
    moon_phase_angle_time_local,
    moon_illumination_width_utc,
    moon_illumination_width_local,
    set_location,
    refraction_horizon_degree
)



def deg_to_dms(deg: float) -> str:
    """
    Konversi derajat desimal ke format derajat-menit-detik (DMS).
    
    Parameters:
    -----------
    deg : float
        Nilai dalam derajat desimal
        
    Returns:
    --------
    str
        String dalam format "DD° MM' SS.SS\""
    """
    sign = "-" if deg < 0 else ""
    deg = abs(deg)
    d = int(deg)
    m = int((deg - d) * 60)
    s = (deg - d - m / 60) * 3600
    return f"{sign}{d}° {m}' {s:.2f}\""


class HilalVisibilityCalculator:
    """Kelas utama untuk kalkulasi visibilitas hilal"""
    
    # Mapping nama sumber atmosfer untuk label
    SUMBER_ATMOSFER_LABEL = {
        'era5': 'ERA5 Reanalysis (Open-Meteo API)',
        'merra2': 'MERRA-2 Reanalysis (NASA POWER API)',
        'bmkg': 'BMKG Prakiraan Cuaca',
        'manual': 'Input Manual',
    }

    def __init__(self,
                 nama_tempat: str,
                 lintang: float,
                 bujur: float,
                 elevasi: float,
                 timezone_str: str,
                 bulan_hijri: int,
                 tahun_hijri: int,
                 delta_day_offset: int = 0,
                 bias_t: float = 0.0,
                 bias_rh: float = 0.0,
                 sumber_atmosfer: str = 'era5',
                 adm4_code: str = '',
                 manual_rh: float = 80.0,
                 manual_t: float = 25.0,
                 manual_p: float = 1013.25):
        """
        Inisialisasi kalkulator visibilitas hilal.

        Parameters:
        -----------
        nama_tempat : str
            Nama lokasi pengamatan
        lintang : float
            Lintang dalam derajat (positif untuk utara, negatif untuk selatan)
        bujur : float
            Bujur dalam derajat (positif untuk timur, negatif untuk barat)
        elevasi : float
            Ketinggian dalam meter di atas permukaan laut
        timezone_str : str
            Zona waktu (contoh: "Asia/Jakarta" atau "+7")
        bulan_hijri : int
            Bulan hijriah (1-12)
        tahun_hijri : int
            Tahun hijriah
        delta_day_offset : int
            Offset hari untuk pengamatan (default: 0)
        bias_t : float
            Bias suhu reanalisis (°C), definisi: bias = Reanalisis - Obs (default: 0.0)
        bias_rh : float
            Bias RH reanalisis (%), definisi: bias = Reanalisis - Obs (default: 0.0)
        sumber_atmosfer : str
            Sumber data atmosfer: 'era5', 'merra2', 'bmkg', atau 'manual'
        adm4_code : str
            Kode wilayah BMKG (hanya untuk sumber 'bmkg')
        manual_rh : float
            RH manual (%) jika sumber='manual'
        manual_t : float
            Suhu manual (°C) jika sumber='manual'
        manual_p : float
            Tekanan manual (mbar) jika sumber='manual'
        """
        self.nama_tempat = nama_tempat
        self.lintang = lintang
        self.bujur = bujur
        self.elevasi = elevasi
        self.timezone_str = timezone_str
        self.bulan_hijri = bulan_hijri
        self.tahun_hijri = tahun_hijri
        self.delta_day_offset = delta_day_offset
        self.bias_t = bias_t
        self.bias_rh = bias_rh
        self.sumber_atmosfer = sumber_atmosfer.lower()
        self.adm4_code = adm4_code
        self.manual_rh = manual_rh
        self.manual_t = manual_t
        self.manual_p = manual_p
        
        # Setup lokasi untuk perhitungan astronomis
        self.location = set_location(lintang, bujur, elevasi)
        
        # Catatan: parameter moonlight (ALTMOON, AZIMOON, PHASE_MOON, SNELLEN)
        # telah dihapus karena modul Schaefer sudah diadaptasi khusus untuk
        # visibilitas hilal (Bulan = objek pengamatan, bukan sumber background)
        
        # Hasil perhitungan akan disimpan di sini
        self.hasil: Dict[str, Any] = {}

    def _estimasi_tekanan(self) -> float:
        """Estimasi tekanan udara dari elevasi menggunakan rumus barometrik standar."""
        return 1013.25 * (1 - 2.25577e-5 * self.elevasi) ** 5.25588

    def _fetch_atmosfer(self,
                        observing_location: ObservingLocation,
                        waktu_utc: datetime,
                        verbose: bool = True) -> Tuple[float, float, float, float, float]:
        """
        Mengambil data atmosfer dan menerapkan koreksi bias.
        Dispatch berdasarkan self.sumber_atmosfer: 'era5', 'merra2', 'bmkg', atau 'manual'.

        Returns:
        --------
        rh_raw, temperature_raw, rh, temperature, pressure
        """
        indent = "  " if verbose else "    "
        sumber = self.sumber_atmosfer
        label = self.SUMBER_ATMOSFER_LABEL.get(sumber, sumber.upper())

        if sumber == 'manual':
            rh_raw = self.manual_rh
            temperature_raw = self.manual_t
            pressure = self.manual_p
            if verbose:
                print(f"{indent}[✓] Data atmosfer MANUAL:")
                print(f"{indent}     RH={rh_raw:.2f}%, T={temperature_raw:.2f}°C, P={pressure:.2f} mbar")
            # Manual: tidak ada koreksi bias
            return rh_raw, temperature_raw, rh_raw, temperature_raw, pressure

        if sumber == 'era5':
            try:
                rh_raw, temperature_raw, pressure = era5_get_rh_t(observing_location, waktu_utc)
                if verbose:
                    print(f"{indent}[✓] Data atmosfer ERA5 berhasil diambil:")
                    print(f"{indent}     RH={rh_raw:.2f}%, T={temperature_raw:.2f}°C, P={pressure:.2f} mbar")
            except ERA5APIError as e:
                if verbose:
                    print(f"{indent}[!] ERA5 API Error: {e}")
                    print(f"{indent}[!] Data atmosfer TIDAK tersedia - Menggunakan nilai DEFAULT:")
                    print(f"{indent}     RH=80.00%, T=25.00°C, P=1013.25 mbar")
                rh_raw, temperature_raw, pressure = 80.0, 25.0, 1013.25

        elif sumber == 'merra2':
            try:
                loc_merra = MERRA2Location(
                    name=self.nama_tempat, latitude=self.lintang,
                    longitude=self.bujur, altitude=self.elevasi,
                    timezone=self.timezone_str
                )
                rh_raw, temperature_raw, pressure = merra2_get_rh_t(loc_merra, waktu_utc)
                if verbose:
                    print(f"{indent}[✓] Data atmosfer MERRA-2 berhasil diambil:")
                    print(f"{indent}     RH={rh_raw:.2f}%, T={temperature_raw:.2f}°C, P={pressure:.2f} kPa")
                # MERRA-2 mengembalikan tekanan dalam kPa, konversi ke mbar
                pressure = pressure * 10.0
            except PowerAPIError as e:
                if verbose:
                    print(f"{indent}[!] MERRA-2 API Error: {e}")
                    print(f"{indent}[!] Data atmosfer TIDAK tersedia - Menggunakan nilai DEFAULT:")
                    print(f"{indent}     RH=80.00%, T=25.00°C, P=1013.25 mbar")
                rh_raw, temperature_raw, pressure = 80.0, 25.0, 1013.25

        elif sumber == 'bmkg':
            try:
                loc_bmkg = BMKGLocation(
                    name=self.nama_tempat, latitude=self.lintang,
                    longitude=self.bujur, altitude=self.elevasi,
                    timezone=self.timezone_str, adm4_code=self.adm4_code
                )
                # BMKG menggunakan waktu lokal, konversi dari UTC
                waktu_local = convert_utc_to_localtime(self.timezone_str, utc_datetime=waktu_utc)
                # Hapus tzinfo karena BMKG menerima naive datetime
                if hasattr(waktu_local, 'tzinfo') and waktu_local.tzinfo is not None:
                    waktu_local = waktu_local.replace(tzinfo=None)
                rh_raw, temperature_raw = bmkg_get_rh_t_local(loc_bmkg, waktu_local)
                # BMKG tidak menyediakan tekanan, estimasi dari elevasi
                pressure = self._estimasi_tekanan()
                if verbose:
                    print(f"{indent}[✓] Data atmosfer BMKG berhasil diambil:")
                    print(f"{indent}     RH={rh_raw:.2f}%, T={temperature_raw:.2f}°C, P={pressure:.2f} mbar (estimasi)")
            except BMKGAPIError as e:
                if verbose:
                    print(f"{indent}[!] BMKG API Error: {e}")
                    print(f"{indent}[!] Data atmosfer TIDAK tersedia - Menggunakan nilai DEFAULT:")
                    print(f"{indent}     RH=80.00%, T=25.00°C, P=1013.25 mbar")
                rh_raw, temperature_raw, pressure = 80.0, 25.0, 1013.25
        else:
            raise ValueError(f"Sumber atmosfer tidak dikenal: '{sumber}'")

        # Terapkan koreksi bias (untuk semua sumber API)
        rh, temperature, pressure = apply_bias_correction(
            rh_raw, temperature_raw, pressure,
            bias_t=self.bias_t,
            bias_rh=self.bias_rh
        )

        if verbose and (self.bias_t != 0.0 or self.bias_rh != 0.0):
            print(f"{indent}[Bias Correction] T: {temperature_raw:.2f} → {temperature:.2f}°C (bias={self.bias_t:+.1f})")
            print(f"{indent}[Bias Correction] RH: {rh_raw:.2f} → {rh:.2f}% (bias={self.bias_rh:+.1f})")

        return rh_raw, temperature_raw, rh, temperature, pressure

    def _fetch_atmosfer_dua_titik(self,
                                   observing_location: ObservingLocation,
                                   waktu_utc_0: datetime,
                                   waktu_utc_1: datetime,
                                   verbose: bool = True
                                   ) -> Tuple[Tuple[float, float, float], Tuple[float, float, float]]:
        """
        Mengambil data atmosfer pada dua titik waktu untuk interpolasi linear.

        Parameters
        ----------
        observing_location : ObservingLocation
        waktu_utc_0 : datetime
            Titik waktu pertama (sunset)
        waktu_utc_1 : datetime
            Titik waktu kedua (sunset + 1 jam)

        Returns
        -------
        atm_0 : (rh, temperature, pressure) pada waktu_utc_0
        atm_1 : (rh, temperature, pressure) pada waktu_utc_1
        """
        if verbose:
            print(f"  Mengambil data atmosfer pada 2 titik untuk interpolasi...")

        _, _, rh_0, t_0, p_0 = self._fetch_atmosfer(
            observing_location, waktu_utc_0, verbose=verbose
        )
        _, _, rh_1, t_1, p_1 = self._fetch_atmosfer(
            observing_location, waktu_utc_1, verbose=False
        )

        if verbose:
            w0_str = waktu_utc_0.strftime('%H:%M')
            w1_str = waktu_utc_1.strftime('%H:%M')
            print(f"    Titik 1 ({w0_str} UTC): RH={rh_0:.1f}%, T={t_0:.1f}°C, P={p_0:.1f} mbar")
            print(f"    Titik 2 ({w1_str} UTC): RH={rh_1:.1f}%, T={t_1:.1f}°C, P={p_1:.1f} mbar")

        return (rh_0, t_0, p_0), (rh_1, t_1, p_1)

    @staticmethod
    def _interpolasi_atmosfer(atm_0: Tuple[float, float, float],
                               atm_1: Tuple[float, float, float],
                               t0: datetime, t1: datetime,
                               t: datetime) -> Tuple[float, float, float]:
        """
        Interpolasi linear data atmosfer antara dua titik waktu.

        Parameters
        ----------
        atm_0 : (rh, temperature, pressure) pada t0
        atm_1 : (rh, temperature, pressure) pada t1
        t0, t1 : datetime — waktu referensi
        t : datetime — waktu target

        Returns
        -------
        (rh, temperature, pressure) hasil interpolasi
        """
        total = (t1 - t0).total_seconds()
        if total <= 0:
            return atm_0

        frac = max(0.0, min(1.0, (t - t0).total_seconds() / total))

        rh = atm_0[0] + frac * (atm_1[0] - atm_0[0])
        temperature = atm_0[1] + frac * (atm_1[1] - atm_0[1])
        pressure = atm_0[2] + frac * (atm_1[2] - atm_0[2])

        return rh, temperature, pressure

    def hitung_ijtima(self) -> Tuple[datetime, datetime]:
        """
        Menghitung waktu ijtima (konjungsi) untuk bulan hijriah yang ditentukan.
        
        Returns:
        --------
        ijtima_utc : datetime
            Waktu ijtima dalam UTC
        ijtima_local : datetime
            Waktu ijtima dalam waktu lokal
        """
        ijtima_utc = newmoon_hijri_month_utc(self.tahun_hijri, self.bulan_hijri)
        ijtima_local = convert_utc_to_localtime(self.timezone_str, utc_datetime=ijtima_utc)
        
        self.hasil['ijtima_utc'] = ijtima_utc
        self.hasil['ijtima_local'] = ijtima_local
        
        return ijtima_utc, ijtima_local
    
    def tentukan_tanggal_pengamatan(self, ijtima_utc: datetime) -> Tuple[datetime, datetime, float, float]:
        """
        Menentukan tanggal pengamatan berdasarkan waktu ijtima dan sunset.

        ALUR KOREKSI SUNSET:
        1. Hitung sunset GEOMETRIS (tanpa koreksi refraksi) menggunakan skyfield
        2. Gunakan waktu sunset geometris untuk mengambil data atmosfer (RH, T) dari sumber yang dipilih
        3. Hitung sunset APPARENT dengan koreksi refraksi menggunakan T dari API
        4. Data RH dan T digunakan untuk perhitungan sky brightness

        ATURAN HISAB:
        1. Konversi ijtima UTC ke waktu lokal
        2. Jika ijtima lokal terjadi sebelum jam 12:00 (tengah malam - siang):
           - Gunakan tanggal ijtima lokal untuk pengamatan (sore hari itu)
        3. Jika ijtima lokal terjadi setelah jam 12:00 (siang - tengah malam):
           - Bandingkan dengan sunset lokal sore hari itu
           - Jika ijtima < sunset: amati hari berikutnya
           - Jika ijtima >= sunset: amati hari yang sama

        Contoh untuk kasus Muharram 1444 (29 Juli 2022):
        - Ijtima UTC: 2022-07-28 17:55:02
        - Ijtima Lokal (WIB): 2022-07-29 00:55:02
        - Karena ijtima lokal (00:55) < 12:00, maka pengamatan dilakukan pada 29 Juli 2022 sore
        - Sunset 29 Juli sore: ~17:30 WIB -> Bulan sudah cukup tinggi untuk diamati

        Parameters:
        -----------
        ijtima_utc : datetime
            Waktu ijtima dalam UTC

        Returns:
        --------
        sunset_utc : datetime
            Waktu sunset dalam UTC pada hari pengamatan
        sunset_local : datetime
            Waktu sunset dalam waktu lokal (dengan koreksi refraksi)
        rh : float
            Relative humidity dalam persen (dari API, terkoreksi bias)
        temperature : float
            Suhu dalam derajat Celsius (dari API, terkoreksi bias)
        pressure : float
            Tekanan udara dalam mbar (dari API)
        """
        # Konversi ijtima ke waktu lokal
        ijtima_local = convert_utc_to_localtime(self.timezone_str, utc_datetime=ijtima_utc)

        # Aturan sederhana berdasarkan jam ijtima lokal
        if ijtima_local.hour < 12:
            # Ijtima terjadi sebelum jam 12:00 (tengah malam sampai sebelum siang)
            tanggal_pengamatan = ijtima_local.date()
        else:
            # Ijtima terjadi setelah jam 12:00 (siang sampai tengah malam)
            # Gunakan sunrise_sunset_local untuk estimasi sunset
            _, sunset_local_ijtima = sunrise_sunset_local(
                self.location,
                self.timezone_str,
                year=ijtima_utc.year,
                month=ijtima_utc.month,
                day=ijtima_utc.day
            )

            if ijtima_local < sunset_local_ijtima:
                tanggal_pengamatan = ijtima_local.date() + timedelta(days=1)
            else:
                tanggal_pengamatan = ijtima_local.date()

        # Terapkan delta day offset
        tanggal_pengamatan += timedelta(days=self.delta_day_offset)

        # LANGKAH 1: Hitung ESTIMASI sunset untuk fetch weather
        # Gunakan default T dan P untuk estimasi awal (standard refraction)
        _, sunset_est_local = sunrise_sunset_local(
            self.location,
            self.timezone_str,
            year=tanggal_pengamatan.year,
            month=tanggal_pengamatan.month,
            day=tanggal_pengamatan.day
        )
        sunset_est_utc = convert_localtime_to_utc(self.timezone_str, local_datetime=sunset_est_local)

        # LANGKAH 2: Ambil data atmosfer (RH dan T) pada waktu sunset geometris
        loc = ObservingLocation(
            name=self.nama_tempat,
            latitude=self.lintang,
            longitude=self.bujur,
            altitude=self.elevasi,
            timezone=self.timezone_str
        )

        # Pastikan sunset_est_utc adalah timezone-aware UTC datetime
        if sunset_est_utc.tzinfo is None or sunset_est_utc.tzinfo.utcoffset(sunset_est_utc) is None:
            sunset_est_utc = sunset_est_utc.replace(tzinfo=timezone.utc)

        # Ambil data atmosfer dan terapkan koreksi bias
        rh_raw, temperature_raw, rh, temperature, pressure = self._fetch_atmosfer(
            loc, sunset_est_utc, verbose=True
        )

        # LANGKAH 3: Hitung sunset APPARENT dengan koreksi refraksi menggunakan T terkoreksi
        _, sunset_local = sunrise_sunset_local(
            self.location,
            self.timezone_str,
            year=tanggal_pengamatan.year,
            month=tanggal_pengamatan.month,
            day=tanggal_pengamatan.day,
            temperature_C=temperature,  # Gunakan temperature terkoreksi
            pressure_mbar=pressure  # Gunakan pressure dari API
        )

        # Konversi sunset lokal ke UTC
        sunset_utc = convert_localtime_to_utc(self.timezone_str, local_datetime=sunset_local)

        # Simpan hasil (raw + corrected)
        self.hasil['sunset_utc'] = sunset_utc
        self.hasil['sunset_local'] = sunset_local
        self.hasil['sunset_est_utc'] = sunset_est_utc
        self.hasil['sunset_est_local'] = sunset_est_local
        self.hasil['tanggal_pengamatan'] = datetime.combine(tanggal_pengamatan, datetime.min.time())
        self.hasil['rh_raw'] = rh_raw
        self.hasil['temperature_raw'] = temperature_raw
        self.hasil['rh'] = rh
        self.hasil['temperature'] = temperature
        self.hasil['pressure'] = pressure
        self.hasil['bias_t'] = self.bias_t
        self.hasil['bias_rh'] = self.bias_rh
        self.hasil['sumber_atmosfer'] = self.sumber_atmosfer
        self.hasil['observing_location'] = loc  # Simpan untuk digunakan di loop optimal

        return sunset_utc, sunset_local, rh, temperature, pressure
    
    def hitung_posisi_matahari_bulan(self, sunset_local: datetime,
                                     temperature_C: float = 10.0,
                                     pressure_mbar: float = 1030.0) -> Dict[str, float]:
        """
        Menghitung posisi matahari dan bulan saat sunset.
        
        Parameters:
        -----------
        sunset_local : datetime
            Waktu sunset dalam waktu lokal
        temperature_C : float
            Suhu udara dalam derajat Celsius (untuk koreksi refraksi)
        pressure_mbar : float
            Tekanan udara dalam mbar (untuk koreksi refraksi)
            
        Returns:
        --------
        Dict[str, float]
            Dictionary berisi:
            - sun_alt: Altitude matahari (derajat)
            - sun_az: Azimuth matahari (derajat)
            - moon_alt: Altitude bulan (derajat)
            - moon_az: Azimuth bulan (derajat)
            - elongation: Elongasi toposentrik (derajat)
            - phase_angle: Sudut fase bulan (derajat)
            - moon_semidiameter: Semidiameter bulan (derajat)
        """
        # Posisi matahari (menggunakan local time + koreksi refraksi dinamis)
        sun_alt, sun_az, _ = sun_position_time_local(
            self.location,
            self.timezone_str,
            local_datetime=sunset_local,
            temperature_C=temperature_C,
            pressure_mbar=pressure_mbar
        )
        
        # Posisi bulan (menggunakan local time + koreksi refraksi dinamis)
        moon_alt, moon_az, _ = moon_position_time_local(
            self.location,
            self.timezone_str,
            local_datetime=sunset_local,
            temperature_C=temperature_C,
            pressure_mbar=pressure_mbar
        )
        
        # Elongasi toposentrik
        elongation = moon_elongation_time_local(
            self.timezone_str,
            location=self.location,
            local_datetime=sunset_local
        )
        
        # Sudut fase bulan
        phase_angle = moon_phase_angle_time_local(
            self.timezone_str,
            location=self.location,
            local_datetime=sunset_local
        )
        
        # Semidiameter bulan
        _, _, parallax, SD = moon_illumination_width_local(
            self.timezone_str,
            location=self.location,
            local_datetime=sunset_local
        )
        moon_semidiameter = SD  # dalam derajat
        
        return {
            'sun_alt': sun_alt,
            'sun_az': sun_az,
            'moon_alt': moon_alt,
            'moon_az': moon_az,
            'elongation': elongation,
            'phase_angle': phase_angle,
            'moon_semidiameter': moon_semidiameter
        }
    
    def hitung_sky_brightness_schaefer(self,
                                        rh: float,
                                        temperature: float,
                                        posisi: Dict[str, float]) -> Tuple[float, float]:
        """
        Menghitung sky brightness menggunakan model Schaefer.
        
        Parameters:
        -----------
        rh : float
            Relative humidity dalam persen
        temperature : float
            Suhu dalam derajat Celsius
        posisi : Dict[str, float]
            Dictionary posisi matahari dan bulan
            
        Returns:
        --------
        sky_brightness_nl : float
            Sky brightness dalam nanoLambert
        k_v : float
            Koefisien ekstingsi band V
        """
        # Hitung selisih azimuth sun-moon
        azisun = abs(posisi['sun_az'] - posisi['moon_az'])

        result = {}
        try:
            result = hitung_sky_brightness(
                month=self.hasil['tanggal_pengamatan'].month,
                year=self.hasil['tanggal_pengamatan'].year,
                altsun=posisi['sun_alt'],
                azisun=azisun,
                humidity=rh,
                temperature=temperature,
                latitude=self.lintang,
                elevation=self.elevasi,
                alt_objek=max(posisi['moon_alt'], 0.0),
            )

            sky_brightness_nl = float(result.get("sky_brightness", 0.0))

            if sky_brightness_nl <= 0:
                sky_brightness_nl = 1.0

        except (ValueError, ZeroDivisionError) as e:
            print(f"  [!] Error sky brightness: {e} - Menggunakan default 1000 nL")
            sky_brightness_nl = 1000.0

        K_list = result.get("K", [])
        k_v = float(K_list[2]) if len(K_list) > 2 else 0.3

        return sky_brightness_nl, k_v
    
    def hitung_luminansi_hilal_kastner(self,
                                       posisi: Dict[str, float],
                                       k_v: float) -> float:
        """
        Menghitung luminansi hilal menggunakan model Kastner.
        
        Parameters:
        -----------
        posisi : Dict[str, float]
            Dictionary posisi matahari dan bulan
        k_v : float
            Koefisien ekstingsi band V
            
        Returns:
        --------
        luminansi_hilal_nl : float
            Luminansi hilal dalam nanoLambert
        """
        zenith_distance = 90.0 - posisi['moon_alt']
        
        luminansi_hilal_nl = hitung_luminansi_kastner(
            alpha=posisi['phase_angle'],
            r=float(posisi['moon_semidiameter']),
            z=zenith_distance,
            k=k_v
        )
        
        return luminansi_hilal_nl
    
    def hitung_visibilitas_naked_eye(self,
                                      luminansi_hilal_nl: float,
                                      sky_brightness_nl: float,
                                      posisi: dict,
                                      F_naked: float = 2.5) -> tuple:
        """
        Menghitung visibilitas hilal mata telanjang menggunakan model Crumey (2014).

        Perubahan dari versi lama:
          LAMA : hanya menghitung rasio L/B → tidak ada threshold persepsi
          BARU : menghitung Weber contrast → bandingkan dengan Crumey threshold
                 menggunakan luas sabit dan kecerahan langit yang benar

        Parameters
        ----------
        luminansi_hilal_nl : float
            Luminansi permukaan hilal [nanoLambert] dari model Kastner
        sky_brightness_nl : float
            Kecerahan langit [nanoLambert] dari model Schaefer
        posisi : dict
            Dictionary posisi matahari & bulan (harus berisi:
            'elongation', 'moon_semidiameter')
        F_naked : float
            Field factor untuk naked eye (default 2.5)

        Returns
        -------
        rasio_kontras : float
            Weber contrast C = (L - B) / B
        delta_m : float
            Margin visibilitas [mag]: 2.5 × log₁₀(|C_obj| / C_th)
            Positif = terlihat, negatif = tidak terlihat
        crumey_result : dict
            Hasil lengkap dari hilal_naked_eye_visibility()
        """
        if sky_brightness_nl <= 0.0:
            raise ValueError("Sky brightness harus positif")

        # Panggil model Crumey untuk naked eye
        result = hilal_naked_eye_visibility(
            L_hilal_nL=luminansi_hilal_nl,
            B_sky_nL=sky_brightness_nl,
            phase_angle_deg=posisi['phase_angle'],
            moon_sd_deg=float(posisi['moon_semidiameter']),
            F=F_naked,
            mode='auto',  # otomatis pilih scotopic/combined berdasarkan B
        )

        rasio_kontras = result['C_obj']
        delta_m = result['delta_m']

        # Hindari -inf: hitung margin saat C_obj ≤ 0
        # Formula: 2.5 × (log10(L/B) − log10(C_th))
        # Ini mempertahankan telescope gain:
        #   delta_m_tel − delta_m_ne = 2.5 × log10(C_th_ne / C_th_tel)
        if math.isinf(delta_m) and delta_m < 0:
            C_obj = result['C_obj']
            C_th = result['C_th']
            L_over_B = max(1.0 + C_obj, 1e-15)
            if C_th > 0:
                delta_m = 2.5 * (math.log10(L_over_B) - math.log10(C_th))
            else:
                delta_m = -99.0

        return rasio_kontras, delta_m, result
    
    def hitung_visibilitas_teleskop(self,
                                     luminansi_hilal_nl: float,
                                     sky_brightness_nl: float,
                                     posisi: Dict[str, float],
                                     aperture: float = 66.0,
                                     magnification: float = 50.0,
                                     transmission: float = 0.95,
                                     n_surfaces: int = 6,
                                     central_obstruction: float = 0.0,
                                     observer_age: float = 22.0,
                                     seeing: float = 3.0,
                                     field_factor: float = 2.4) -> Tuple[float, float, float, float, float]:
        """
        Menghitung visibilitas hilal melalui teleskop.
        Pipeline langsung: Schaefer (telescope_limit) + Crumey (full_rumus_crumey).

        Pipeline:
          1. Hitung luas sabit dari elongasi & semidiameter
          2. Hitung faktor teleskop Schaefer: Fb, Ft, Fp, Fa, Fm
          3. Weber contrast: C_obj = (Lt − B) / B  (dipertahankan, Crumey Eq. 76)
          4. Background apparent: Ba = (δmin/p)² × B / Ft  (Crumey Eq. 66)
          5. Area efektif retina: A_eff = A_sr × M²
          6. Threshold Crumey: φ = FT × FM × F, C_th = contrast_threshold(A_eff, Ba, φ)
          7. Keputusan: visible = (C_obj > C_th)

        Parameters:
        -----------
        luminansi_hilal_nl : float
            Luminansi hilal dalam nL (I_0)
        sky_brightness_nl : float
            Sky brightness dalam nL (B_0)
        posisi : Dict[str, float]
            Dictionary posisi matahari dan bulan
        aperture : float
            Diameter aperture teleskop (mm)
        magnification : float
            Pembesaran teleskop
        transmission : float
            Transmisi per permukaan optik (default: 0.95)
        n_surfaces : int
            Jumlah permukaan optik (default: 6)
        central_obstruction : float
            Diameter obstruksi pusat (mm, default: 0.0 untuk refraktor)
        observer_age : float
            Usia pengamat (tahun, default: 22.0)
        seeing : float
            Ukuran seeing disk (arcseconds, default: 3.0)
        field_factor : float
            Personal/kondisi lapangan field factor F (default: 2.4)

        Returns:
        --------
        luminansi_hilal_tel_nl : float
            Luminansi hilal efektif teleskop (I_eff) dalam nL
        sky_brightness_tel_nl : float
            Sky brightness efektif teleskop (B_eff) dalam nL
        rasio_kontras_tel : float
            Weber contrast objek terhadap latar belakang (C_obj)
        c_th_tel : float
            Contrast Threshold berdasarkan Crumey
        delta_m_tel : float
            Margin teleskop [mag]: 2.5 × log₁₀(C_obj / C_th), >0 = detectable
        """
        Lt_nL = luminansi_hilal_nl
        B_nL = sky_brightness_nl

        # ── Langkah 1: Luas sabit ─────────────────────────────────────────
        A_arcmin2 = crescent_area_arcmin2(
            posisi['phase_angle'], float(posisi['moon_semidiameter'])
        )
        A_sr = arcmin2_to_sr(A_arcmin2)

        # ── Langkah 2: Faktor koreksi teleskop (Schaefer) ─────────────────
        model = TelescopeVisibilityModel()
        factors = model.calculate_factors(
            D=aperture,
            Ds=central_obstruction,
            M=magnification,
            age=observer_age,
            t1=transmission,
            n=n_surfaces,
            theta=seeing
        )

        # ── Langkah 3: Weber contrast (dipertahankan, Crumey Eq. 76) ─────
        # Hilal = extended source → kontras Weber dipertahankan melalui teleskop
        if B_nL <= 0:
            return 0.0, 0.0, float('nan'), float('nan'), -99.0

        C_obj = (Lt_nL - B_nL) / B_nL

        # ── Langkah 4: Background apparent melalui teleskop ───────────────
        # Crumey (2014) Eq. 66: Ba = (δmin/p)² × B / Ft
        d_exit = factors["exit_pupil"]   # exit pupil teleskop [mm]
        De = factors["De"]               # diameter pupil mata [mm]
        Ft = factors["Ft"]               # 1/transmittance
        delta_min = min(d_exit, De)
        Ba_factor = (delta_min / De) ** 2 / Ft   # Crumey Eq. 66
        B_corr_nL  = B_nL  * Ba_factor
        Lt_corr_nL = Lt_nL * Ba_factor   # extended source: faktor sama

        B_cd  = nL_to_cd_m2(B_corr_nL)
        Lt_cd = nL_to_cd_m2(Lt_corr_nL)

        # ── Langkah 5: Area efektif retina ────────────────────────────────
        # Magnifikasi memperbesar area angular di retina: A_eff = A × M²
        A_eff = A_sr * (magnification ** 2)

        # ── Langkah 6: Threshold Crumey ───────────────────────────────────
        # Crumey (2014) Sec. 3.2: φ = FT × FM × F
        #   FT = √2 (koreksi monocular, Sec. 1.6.4)
        #   FM = 1.0 (faktor magnifikasi, Eq. 83 — netral untuk extended source)
        #   F  = field_factor (personal/kondisi lapangan)
        if B_cd <= 0:
            return cd_m2_to_nL(Lt_cd), 0.0, C_obj, float('nan'), -99.0

        FT = math.sqrt(2)   # monocular viewing correction (Sec. 1.6.4)
        FM = 1.0             # magnification factor (Eq. 83)
        phi = FT * FM * field_factor  # Crumey Sec. 3.2
        c_th_tel = contrast_threshold(A_eff, B_cd, F=phi, mode='auto')

        # ── Langkah 7: Margin visibilitas ─────────────────────────────────
        if C_obj > 0 and c_th_tel > 0:
            delta_m_tel = 2.5 * math.log10(C_obj / c_th_tel)
        else:
            delta_m_tel = float('-inf')

        # Hindari -inf: hitung margin saat C_obj ≤ 0
        # Formula: 2.5 × (log10(L/B) − log10(C_th))
        # Ini mempertahankan telescope gain:
        #   delta_m_tel − delta_m_ne = 2.5 × log10(C_th_ne / C_th_tel)
        if math.isinf(delta_m_tel) and delta_m_tel < 0:
            L_over_B = max(1.0 + C_obj, 1e-15)
            if c_th_tel > 0:
                delta_m_tel = 2.5 * (math.log10(L_over_B) - math.log10(c_th_tel))
            else:
                delta_m_tel = -99.0

        # Konversi balik cd/m² → nL untuk output
        I_eff = cd_m2_to_nL(Lt_cd)
        B_eff = cd_m2_to_nL(B_cd)

        return I_eff, B_eff, C_obj, c_th_tel, delta_m_tel

    def hitung_visibilitas_pada_waktu(self,
                                       waktu_local: datetime,
                                       observing_location: ObservingLocation,
                                       aperture: float = 66.0,
                                       magnification: float = 50.0,
                                       F_naked: float = 2.5,
                                       field_factor: float = 2.4,
                                       cached_atm: Optional[Tuple[float, float, float]] = None
                                       ) -> Dict[str, Any]:
        """
        Menghitung visibilitas hilal pada waktu tertentu.

        Parameters
        ----------
        cached_atm : (rh, temperature, pressure) atau None
            Jika diberikan, gunakan data atmosfer ini (dari interpolasi)
            tanpa melakukan API call. Jika None, fetch via API.
        """
        if cached_atm is not None:
            rh, temperature, pressure = cached_atm
        else:
            # Konversi waktu lokal ke UTC untuk API call
            waktu_utc = waktu_local.astimezone(timezone.utc)
            _, _, rh, temperature, pressure = self._fetch_atmosfer(
                observing_location, waktu_utc, verbose=False
            )
        
        # LANGKAH 2: Hitung posisi matahari dan bulan dengan T/P dinamis
        posisi = self.hitung_posisi_matahari_bulan(
            waktu_local,
            temperature_C=temperature,
            pressure_mbar=pressure
        )
        
        # Jika bulan sudah di bawah horizon, kembalikan hasil kosong
        if posisi['moon_alt'] <= 0:
            return {
                'waktu_local': waktu_local,
                'moon_alt': posisi['moon_alt'],
                'sun_alt': posisi['sun_alt'],
                'valid': False,
                'delta_m_ne': -99.0,
                'delta_m_tel': -99.0,
                'telescope_gain': 0.0,
                'rh': rh,
                'temperature': temperature
            }
        
        # LANGKAH 3: Hitung sky brightness (Schaefer)
        sky_brightness_nl, k_v = self.hitung_sky_brightness_schaefer(rh, temperature, posisi)
        
        # Hitung luminansi hilal (Kastner)
        luminansi_hilal_nl = self.hitung_luminansi_hilal_kastner(posisi, k_v)
        
        # Hitung visibilitas naked eye
        rasio_kontras_ne, delta_m_ne, crumey_ne = self.hitung_visibilitas_naked_eye(
            luminansi_hilal_nl, sky_brightness_nl, posisi, F_naked=F_naked
        )

        # Hitung visibilitas teleskop
        (luminansi_hilal_tel_nl, sky_brightness_tel_nl,
         rasio_kontras_tel, c_th_tel, delta_m_tel) = self.hitung_visibilitas_teleskop(
            luminansi_hilal_nl, sky_brightness_nl, posisi,
            aperture=aperture,
            magnification=magnification,
            field_factor=field_factor
        )

        # Telescope gain: keuntungan threshold teleskop vs naked eye [mag]
        c_th_ne = crumey_ne['C_th']
        if c_th_tel > 0 and c_th_ne > 0:
            telescope_gain = 2.5 * math.log10(c_th_ne / c_th_tel)
        else:
            telescope_gain = 0.0

        return {
            'waktu_local': waktu_local,
            'moon_alt': posisi['moon_alt'],
            'sun_alt': posisi['sun_alt'],
            'elongation': posisi['elongation'],
            'sky_brightness_nl': sky_brightness_nl,
            'luminansi_hilal_nl': luminansi_hilal_nl,
            'luminansi_hilal_tel_nl': luminansi_hilal_tel_nl,
            'sky_brightness_tel_nl': sky_brightness_tel_nl,
            'k_v': k_v,
            'delta_m_ne': delta_m_ne,
            'delta_m_tel': delta_m_tel,
            'rasio_kontras_ne': rasio_kontras_ne,
            'rasio_kontras_tel': rasio_kontras_tel,
            'c_th_tel': c_th_tel,
            'telescope_gain': telescope_gain,
            'rh': rh,
            'temperature': temperature,
            'crumey_ne_C_th': crumey_ne['C_th'],
            'valid': True
        }
    
    def cari_visibilitas_optimal(self,
                                  sunset_local: datetime,
                                  observing_location: ObservingLocation,
                                  aperture: float = 66.0,
                                  magnification: float = 50.0,
                                  F_naked: float = 2.5,
                                  field_factor: float = 2.4,
                                  interval_menit: int = 1,
                                  min_moon_alt: float = 2.0,
                                  start_delay_menit: int = 1) -> Dict[str, Any]:
        """
        Loop dari sunset hingga bulan mendekati horizon untuk mencari
        waktu optimal (delta_m maksimum).

        Perbaikan v2:
        - Atmosfer di-fetch 2 titik (sunset & +1 jam), lalu diinterpolasi linear
        - Default interval 1 menit (sebelumnya 2 menit)
        - Refinement ±2 menit di sekitar puncak dengan step 15 detik
        - Track window visibilitas kontinu (start, end, durasi terpanjang)
        - Start delay 1 menit agar teleskop bisa mendeteksi lebih awal

        Parameters
        ----------
        interval_menit : int
            Interval antar timestep dalam menit (default 1)
        min_moon_alt : float
            Altitude bulan minimum (derajat) sebelum loop berhenti (default 2.0)
        start_delay_menit : int
            Delay setelah sunset sebelum loop dimulai (default 1)
        """
        print(f"\n  Mencari visibilitas optimal (interval: {interval_menit} menit, "
              f"start delay: {start_delay_menit} menit)...")

        # ── Fetch atmosfer 2 titik untuk interpolasi ──────────────────────
        sunset_utc = convert_localtime_to_utc(self.timezone_str, local_datetime=sunset_local)
        if sunset_utc.tzinfo is None:
            sunset_utc = sunset_utc.replace(tzinfo=timezone.utc)
        sunset_utc_plus1h = sunset_utc + timedelta(hours=1)

        atm_0, atm_1 = self._fetch_atmosfer_dua_titik(
            observing_location, sunset_utc, sunset_utc_plus1h, verbose=True
        )
        t0_utc = sunset_utc
        t1_utc = sunset_utc_plus1h
        print(f"  Atmosfer diinterpolasi linear antara sunset dan +1 jam")

        # ── Inisialisasi tracking ─────────────────────────────────────────
        best_result_ne = None
        best_result_tel = None
        best_delta_m_ne = -99.0
        best_delta_m_tel = -99.0

        # Tracking visibility window: first & last visible
        visibility_start_ne = None
        visibility_end_ne = None
        visibility_start_tel = None
        visibility_end_tel = None

        # Tracking window kontinu terpanjang (NE)
        current_streak_ne = 0
        longest_streak_ne = 0
        streak_start_ne = None
        last_visible_in_streak_ne = None
        best_window_start_ne = None
        best_window_end_ne = None

        # Tracking window kontinu terpanjang (Teleskop)
        current_streak_tel = 0
        longest_streak_tel = 0
        streak_start_tel = None
        last_visible_in_streak_tel = None
        best_window_start_tel = None
        best_window_end_tel = None

        all_results = []
        current_time = sunset_local + timedelta(minutes=start_delay_menit)
        step_count = 0
        max_steps = 120

        # ── Loop utama (kasar) ────────────────────────────────────────────
        while step_count < max_steps:
            # Interpolasi atmosfer untuk waktu ini
            waktu_utc = convert_localtime_to_utc(self.timezone_str, local_datetime=current_time)
            if waktu_utc.tzinfo is None:
                waktu_utc = waktu_utc.replace(tzinfo=timezone.utc)
            rh, temperature, pressure = self._interpolasi_atmosfer(
                atm_0, atm_1, t0_utc, t1_utc, waktu_utc
            )

            result = self.hitung_visibilitas_pada_waktu(
                current_time, observing_location, aperture, magnification,
                F_naked=F_naked, field_factor=field_factor,
                cached_atm=(rh, temperature, pressure)
            )

            all_results.append(result)

            if not result['valid'] or result['moon_alt'] < min_moon_alt:
                print(f"    Berhenti: moon_alt = {result['moon_alt']:.2f}° (< {min_moon_alt}°)")
                break

            # Track best NE
            if result['delta_m_ne'] > best_delta_m_ne:
                best_delta_m_ne = result['delta_m_ne']
                best_result_ne = result

            # Track best telescope
            if result['delta_m_tel'] > best_delta_m_tel:
                best_delta_m_tel = result['delta_m_tel']
                best_result_tel = result

            # ── Track NE visibility window ────────────────────────────────
            if result['delta_m_ne'] > 0:
                if visibility_start_ne is None:
                    visibility_start_ne = current_time
                visibility_end_ne = current_time
                # Track streak kontinu
                if current_streak_ne == 0:
                    streak_start_ne = current_time
                current_streak_ne += 1
                last_visible_in_streak_ne = current_time
            else:
                if current_streak_ne > longest_streak_ne:
                    longest_streak_ne = current_streak_ne
                    best_window_start_ne = streak_start_ne
                    best_window_end_ne = last_visible_in_streak_ne
                current_streak_ne = 0

            # ── Track telescope visibility window ─────────────────────────
            if result['delta_m_tel'] > 0:
                if visibility_start_tel is None:
                    visibility_start_tel = current_time
                visibility_end_tel = current_time
                if current_streak_tel == 0:
                    streak_start_tel = current_time
                current_streak_tel += 1
                last_visible_in_streak_tel = current_time
            else:
                if current_streak_tel > longest_streak_tel:
                    longest_streak_tel = current_streak_tel
                    best_window_start_tel = streak_start_tel
                    best_window_end_tel = last_visible_in_streak_tel
                current_streak_tel = 0

            # Tampilkan progress
            waktu_str = result['waktu_local'].strftime('%H:%M:%S')
            rh_val = result.get('rh', 0) or 0
            t_val = result.get('temperature', 0) or 0
            print(f"    {waktu_str} | alt={result['moon_alt']:5.2f} | "
                  f"RH={rh_val:5.1f}% | T={t_val:5.1f}C | "
                  f"B={result['sky_brightness_nl']:8.1e} | "
                  f"L={result['luminansi_hilal_nl']:8.1e} | "
                  f"dm_ne={result['delta_m_ne']:+5.2f} | "
                  f"dm_tel={result['delta_m_tel']:+5.2f} | "
                  f"gain={result['telescope_gain']:+.2f}")

            current_time += timedelta(minutes=interval_menit)
            step_count += 1

        # Finalisasi streak terakhir (jika loop berakhir saat masih visible)
        if current_streak_ne > longest_streak_ne:
            longest_streak_ne = current_streak_ne
            best_window_start_ne = streak_start_ne
            best_window_end_ne = last_visible_in_streak_ne
        if current_streak_tel > longest_streak_tel:
            longest_streak_tel = current_streak_tel
            best_window_start_tel = streak_start_tel
            best_window_end_tel = last_visible_in_streak_tel

        print(f"  Selesai loop kasar: {len(all_results)} timestep dihitung")

        # ── Refinement: presisi tinggi di sekitar puncak ──────────────────
        if best_result_ne or best_result_tel:
            peak_times = []
            if best_result_ne:
                peak_times.append(best_result_ne['waktu_local'])
            if best_result_tel:
                peak_times.append(best_result_tel['waktu_local'])

            refine_start = min(peak_times) - timedelta(minutes=2)
            refine_end = max(peak_times) + timedelta(minutes=2)
            refine_step = timedelta(seconds=15)

            print(f"\n  Refinement: {refine_start.strftime('%H:%M:%S')} - "
                  f"{refine_end.strftime('%H:%M:%S')} (per 15 detik)")

            t_refine = refine_start
            while t_refine <= refine_end:
                waktu_utc_r = convert_localtime_to_utc(
                    self.timezone_str, local_datetime=t_refine
                )
                if waktu_utc_r.tzinfo is None:
                    waktu_utc_r = waktu_utc_r.replace(tzinfo=timezone.utc)
                rh_r, temp_r, pres_r = self._interpolasi_atmosfer(
                    atm_0, atm_1, t0_utc, t1_utc, waktu_utc_r
                )
                result_r = self.hitung_visibilitas_pada_waktu(
                    t_refine, observing_location, aperture, magnification,
                    F_naked=F_naked, field_factor=field_factor,
                    cached_atm=(rh_r, temp_r, pres_r)
                )
                if result_r['valid']:
                    if result_r['delta_m_ne'] > best_delta_m_ne:
                        best_delta_m_ne = result_r['delta_m_ne']
                        best_result_ne = result_r
                        print(f"    ^ NE peak refined: "
                              f"{t_refine.strftime('%H:%M:%S')} "
                              f"dm={result_r['delta_m_ne']:+.4f}")
                    if result_r['delta_m_tel'] > best_delta_m_tel:
                        best_delta_m_tel = result_r['delta_m_tel']
                        best_result_tel = result_r
                        print(f"    ^ Tel peak refined: "
                              f"{t_refine.strftime('%H:%M:%S')} "
                              f"dm={result_r['delta_m_tel']:+.4f}")
                t_refine += refine_step

        # ── Durasi window kontinu terpanjang ──────────────────────────────
        visibility_duration_ne = longest_streak_ne * interval_menit
        visibility_duration_tel = longest_streak_tel * interval_menit

        return {
            'optimal_time_ne': best_result_ne['waktu_local'] if best_result_ne else None,
            'optimal_delta_m_ne': best_delta_m_ne,
            'optimal_moon_alt_ne': best_result_ne['moon_alt'] if best_result_ne else None,
            'optimal_time_tel': best_result_tel['waktu_local'] if best_result_tel else None,
            'optimal_delta_m_tel': best_delta_m_tel,
            'optimal_moon_alt_tel': best_result_tel['moon_alt'] if best_result_tel else None,
            'optimal_telescope_gain': best_result_tel.get('telescope_gain', 0.0) if best_result_tel else 0.0,
            'best_result_ne': best_result_ne,
            'best_result_tel': best_result_tel,
            'visibility_duration_ne': visibility_duration_ne,
            'visibility_duration_tel': visibility_duration_tel,
            'visibility_start_ne': visibility_start_ne,
            'visibility_start_tel': visibility_start_tel,
            'visibility_end_ne': visibility_end_ne,
            'visibility_end_tel': visibility_end_tel,
            'best_window_start_ne': best_window_start_ne,
            'best_window_end_ne': best_window_end_ne,
            'best_window_start_tel': best_window_start_tel,
            'best_window_end_tel': best_window_end_tel,
            'total_timesteps': len(all_results),
            'all_results': all_results
        }


    
    def jalankan_perhitungan_lengkap(self,
                                      use_telescope: bool = True,
                                      aperture: float = 100.0,
                                      magnification: float = 50.0,
                                      F_naked: float = 2.5,
                                      mode: str = "sunset",
                                      interval_menit: int = 1,
                                      min_moon_alt: float = 2.0,
                                      start_delay_menit: int = 1,
                                      **telescope_kwargs) -> Dict[str, Any]:
        """
        Menjalankan seluruh algoritma perhitungan visibilitas hilal.

        ALUR PROGRAM:
        1. Hitung ijtima/konjungsi
        2. Tentukan tanggal pengamatan dan hitung sunset dengan alur koreksi:
           a. Hitung sunset geometris (tanpa refraksi)
           b. Ambil data atmosfer (RH, T) dari sumber yang dipilih pada waktu sunset geometris
           c. Hitung sunset apparent dengan koreksi refraksi menggunakan T dari API
        3. Hitung posisi matahari dan bulan saat sunset
        4. Hitung sky brightness menggunakan model Schaefer (dengan RH dan T dari API)
        5. Hitung luminansi hilal menggunakan model Kastner
        6. Hitung visibilitas naked eye dan teleskop

        Parameters:
        -----------
        use_telescope : bool
            Apakah akan menghitung visibilitas teleskop
        aperture : float
            Diameter aperture teleskop (mm)
        magnification : float
            Pembesaran teleskop

        Returns:
        --------
        Dict[str, Any]
            Dictionary berisi seluruh hasil perhitungan
        """
        print("Menjalankan perhitungan visibilitas hilal...")

        # Langkah 1: Hitung ijtima
        ijtima_utc, ijtima_local = self.hitung_ijtima()

        # Langkah 2: Tentukan tanggal pengamatan dan hitung sunset dengan koreksi refraksi
        sunset_utc, sunset_local, rh, temperature, pressure = self.tentukan_tanggal_pengamatan(ijtima_utc)

        # Langkah 3: Hitung posisi matahari dan bulan (menggunakan local time)
        posisi = self.hitung_posisi_matahari_bulan(
            sunset_local,
            temperature_C=temperature,
            pressure_mbar=pressure
        )

        # Langkah 4: Hitung sky brightness (Schaefer)
        sky_brightness_nl, k_v = self.hitung_sky_brightness_schaefer(rh, temperature, posisi)

        # Langkah 5: Hitung luminansi hilal (Kastner)
        luminansi_hilal_nl = self.hitung_luminansi_hilal_kastner(posisi, k_v)

        # Langkah 6: Hitung visibilitas naked eye
        rasio_kontras_ne, delta_m_ne, crumey_ne = self.hitung_visibilitas_naked_eye(
            luminansi_hilal_nl, sky_brightness_nl, posisi, F_naked=F_naked
        )

        # Langkah 7: Hitung visibilitas teleskop (jika diminta)
        if use_telescope:
            (luminansi_hilal_tel_nl, sky_brightness_tel_nl,
             rasio_kontras_tel, c_th_tel, delta_m_tel) = self.hitung_visibilitas_teleskop(
                luminansi_hilal_nl, sky_brightness_nl, posisi,
                aperture=aperture,
                magnification=magnification,
                **telescope_kwargs
            )
            # Telescope gain: keuntungan threshold teleskop vs naked eye [mag]
            c_th_ne = crumey_ne['C_th']
            if c_th_tel > 0 and c_th_ne > 0:
                telescope_gain = 2.5 * math.log10(c_th_ne / c_th_tel)
            else:
                telescope_gain = 0.0
        else:
            luminansi_hilal_tel_nl = sky_brightness_tel_nl = 0.0
            rasio_kontras_tel = c_th_tel = delta_m_tel = 0.0
            telescope_gain = 0.0

        # Simpan semua hasil termasuk data posisi lengkap
        self.hasil.update({
            # Data posisi matahari
            'sun_alt': posisi['sun_alt'],
            'sun_az': posisi['sun_az'],
            # Data posisi bulan
            'moon_alt': posisi['moon_alt'],
            'moon_az': posisi['moon_az'],
            'elongation': posisi['elongation'],
            'phase_angle': posisi['phase_angle'],
            'moon_semidiameter': posisi['moon_semidiameter'],
            # Data perhitungan (saat sunset)
            'luminansi_hilal_nl': luminansi_hilal_nl,
            'sky_brightness_nl': sky_brightness_nl,
            'luminansi_hilal_tel_nl': luminansi_hilal_tel_nl,
            'sky_brightness_tel_nl': sky_brightness_tel_nl,
            'k_v': k_v,
            'rasio_kontras_ne': rasio_kontras_ne,
            'delta_m_ne': delta_m_ne,
            'rasio_kontras_tel': rasio_kontras_tel,
            'c_th_tel': c_th_tel,
            'delta_m_tel': delta_m_tel,
            'telescope_gain': telescope_gain,
            'crumey_ne_C_th': crumey_ne['C_th'],
            'crumey_ne_regime': crumey_ne['regime'],
            'crumey_ne_A_arcmin2': crumey_ne['A_arcmin2'],
        })
        
        # Langkah 8: Cari visibilitas optimal (jika mode = "optimal")
        if mode.lower() == "optimal":
            observing_location = self.hasil.get('observing_location')
            
            hasil_optimal = self.cari_visibilitas_optimal(
                sunset_local=sunset_local,
                observing_location=observing_location,
                aperture=aperture,
                magnification=magnification,
                F_naked=F_naked,
                field_factor=telescope_kwargs.get('field_factor', 2.4),
                interval_menit=interval_menit,
                min_moon_alt=min_moon_alt,
                start_delay_menit=start_delay_menit
            )
            
            self.hasil.update({
                'mode': 'optimal',
                'optimal_time_ne': hasil_optimal['optimal_time_ne'],
                'optimal_delta_m_ne': hasil_optimal['optimal_delta_m_ne'],
                'optimal_moon_alt_ne': hasil_optimal['optimal_moon_alt_ne'],
                'optimal_time_tel': hasil_optimal['optimal_time_tel'],
                'optimal_delta_m_tel': hasil_optimal['optimal_delta_m_tel'],
                'optimal_moon_alt_tel': hasil_optimal['optimal_moon_alt_tel'],
                'optimal_telescope_gain': hasil_optimal['optimal_telescope_gain'],
                'optimal_result_ne': hasil_optimal['best_result_ne'],
                'optimal_result_tel': hasil_optimal['best_result_tel'],
                'visibility_duration_ne': hasil_optimal['visibility_duration_ne'],
                'visibility_duration_tel': hasil_optimal['visibility_duration_tel'],
                'visibility_start_ne': hasil_optimal['visibility_start_ne'],
                'visibility_start_tel': hasil_optimal['visibility_start_tel'],
                'visibility_end_ne': hasil_optimal['visibility_end_ne'],
                'visibility_end_tel': hasil_optimal['visibility_end_tel'],
                'best_window_start_ne': hasil_optimal['best_window_start_ne'],
                'best_window_end_ne': hasil_optimal['best_window_end_ne'],
                'best_window_start_tel': hasil_optimal['best_window_start_tel'],
                'best_window_end_tel': hasil_optimal['best_window_end_tel'],
                'total_timesteps': hasil_optimal['total_timesteps'],
                'all_timestep_results': hasil_optimal['all_results']
            })
        else:
            self.hasil['mode'] = 'sunset'
        
        # Tampilkan hasil akhir
        self.tampilkan_hasil_akhir()
        
        return self.hasil
    
    def tampilkan_hasil_akhir(self):
        """Menampilkan ringkasan hasil perhitungan akhir"""
        print("\n" + "=" * 70)
        print("HASIL PERHITUNGAN VISIBILITAS HILAL")
        print("=" * 70)
        
        # === INFORMASI LOKASI ===
        print(f"\n{'='*30} LOKASI {'='*31}")
        print(f"  Nama Tempat           : {self.nama_tempat}")
        print(f"  Lintang               : {self.lintang}°")
        print(f"  Bujur                 : {self.bujur}°")
        print(f"  Elevasi               : {self.elevasi} m")
        print(f"  Timezone              : {self.timezone_str}")
        print(f"  Bulan/Tahun Hijri     : {self.bulan_hijri}/{self.tahun_hijri}")
        
        # === WAKTU ===
        print(f"\n{'='*30} WAKTU {'='*32}")
        if 'ijtima_utc' in self.hasil:
            print(f"  Ijtima UTC            : {self.hasil['ijtima_utc'].strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"  Ijtima Lokal          : {self.hasil['ijtima_local'].strftime('%Y-%m-%d %H:%M:%S')}")
        if 'tanggal_pengamatan' in self.hasil:
            print(f"  Tanggal Pengamatan    : {self.hasil['tanggal_pengamatan'].strftime('%Y-%m-%d')}")
        if 'sunset_utc' in self.hasil:
            print(f"  Sunset UTC            : {self.hasil['sunset_utc'].strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"  Sunset Lokal          : {self.hasil['sunset_local'].strftime('%Y-%m-%d %H:%M:%S')}")
        
        # === DATA ATMOSFER ===
        print(f"\n{'='*28} DATA ATMOSFER {'='*28}")
        if 'rh' in self.hasil:
            # Tampilkan nilai raw dan corrected hanya jika ada bias
            bias_t = self.hasil.get('bias_t', 0.0)
            bias_rh = self.hasil.get('bias_rh', 0.0)
            src_label = self.SUMBER_ATMOSFER_LABEL.get(self.sumber_atmosfer, self.sumber_atmosfer.upper()).split(' ')[0]
            if bias_t != 0.0 or bias_rh != 0.0:
                print(f"  Suhu {src_label} (raw)           : {self.hasil.get('temperature_raw', 0):.2f}°C")
                print(f"  Suhu Terkoreksi (T)       : {self.hasil['temperature']:.2f}°C  (bias={bias_t:+.1f}°C)")
                print(f"  RH {src_label} (raw)             : {self.hasil.get('rh_raw', 0):.2f}%")
                print(f"  RH Terkoreksi             : {self.hasil['rh']:.2f}%  (bias={bias_rh:+.1f}%)")
            else:
                # Tanpa koreksi bias - tampilkan bias 0.0 untuk jelas
                print(f"  Kelembapan Relatif (RH)   : {self.hasil['rh']:.2f}%  (bias=0.0%)")
                print(f"  Suhu (T)                  : {self.hasil['temperature']:.2f}°C  (bias=0.0°C)")
        if 'pressure' in self.hasil:
            print(f"  Tekanan Udara (P)         : {self.hasil['pressure']:.2f} mbar")
        if 'k_v' in self.hasil:
            print(f"  Koefisien Ekstingsi (k_V) : {self.hasil['k_v']:.4f}")
        
        # === POSISI MATAHARI ===
        print(f"\n{'='*27} POSISI MATAHARI {'='*27}")
        if 'sun_alt' in self.hasil:
            print(f"  Altitude Matahari     : {deg_to_dms(self.hasil['sun_alt'])}")
            print(f"  Azimuth Matahari      : {deg_to_dms(self.hasil['sun_az'])}")
        
        # === POSISI BULAN ===
        print(f"\n{'='*28} POSISI BULAN {'='*29}")
        if 'moon_alt' in self.hasil:
            print(f"  Altitude Bulan        : {deg_to_dms(self.hasil['moon_alt'])}")
            print(f"  Azimuth Bulan         : {deg_to_dms(self.hasil['moon_az'])}")
        if 'elongation' in self.hasil:
            print(f"  Elongasi Toposentrik  : {deg_to_dms(self.hasil['elongation'])}")
        if 'phase_angle' in self.hasil:
            print(f"  Phase Angle           : {deg_to_dms(self.hasil['phase_angle'])}")
        if 'moon_semidiameter' in self.hasil:
            print(f"  Semidiameter Bulan    : {deg_to_dms(float(self.hasil['moon_semidiameter']))}")
        
        # === VISIBILITAS HILAL NAKED EYE (Crumey 2014) ===
        print(f"\n{'='*22} VISIBILITAS HILAL NAKED EYE {'='*21}")
        if 'luminansi_hilal_nl' in self.hasil:
            print(f"  Luminansi Hilal       : {self.hasil['luminansi_hilal_nl']:.4e} nL")
            print(f"  Sky Brightness        : {self.hasil['sky_brightness_nl']:.4e} nL")
        if 'rasio_kontras_ne' in self.hasil:
            print(f"  Weber Contrast (C_obj): {self.hasil['rasio_kontras_ne']:.4e}")
            if 'crumey_ne_C_th' in self.hasil:
                print(f"  Threshold (Crumey)    : {self.hasil['crumey_ne_C_th']:.4e}")
            if 'crumey_ne_regime' in self.hasil:
                print(f"  Regime                : {self.hasil['crumey_ne_regime']}")
            if 'crumey_ne_A_arcmin2' in self.hasil:
                print(f"  Luas Sabit            : {self.hasil['crumey_ne_A_arcmin2']:.4f} arcmin²")
            print(f"  Visib. Margin (D_m)   : {self.hasil['delta_m_ne']:.4f}")
            status_ne = "TERLIHAT" if self.hasil['delta_m_ne'] >= 0 else "TIDAK TERLIHAT"
            print(f"  Status                : {status_ne}")
        
        # Tambahkan hasil optimal naked eye jika mode optimal
        if self.hasil.get('mode') == 'optimal' and self.hasil.get('optimal_time_ne'):
            print(f"\n  --- Optimal ---")
            print(f"  Waktu Optimal         : {self.hasil['optimal_time_ne'].strftime('%H:%M:%S')}")
            print(f"  Moon Alt. Optimal     : {self.hasil['optimal_moon_alt_ne']:.2f}°")
            print(f"  Delta m Optimal       : {self.hasil['optimal_delta_m_ne']:.4f}")
            if self.hasil['visibility_duration_ne'] > 0:
                print(f"  Durasi Visibilitas    : {self.hasil['visibility_duration_ne']} menit (window kontinu terpanjang)")
                ws_ne = self.hasil.get('best_window_start_ne')
                we_ne = self.hasil.get('best_window_end_ne')
                if ws_ne and we_ne:
                    print(f"  Window Visibilitas    : {ws_ne.strftime('%H:%M:%S')} - {we_ne.strftime('%H:%M:%S')}")
            else:
                print(f"  Durasi Visibilitas    : 0 menit (tidak pernah delta_m > 0)")
        
        if self.hasil.get('delta_m_ne', -1) >= 0:
            print(f"\n  >> Hilal BERPOTENSI terlihat dengan mata telanjang")
        else:
            print(f"\n  >> Hilal SULIT terlihat dengan mata telanjang")
        
        # === VISIBILITAS HILAL TELESKOP ===
        print(f"\n{'='*24} VISIBILITAS HILAL TELESKOP {'='*22}")
        if 'rasio_kontras_tel' in self.hasil:
            print(f"  Luminansi (Teleskop)  : {self.hasil['luminansi_hilal_tel_nl']:.4e} nL")
            print(f"  Sky Bright. (Teleskop): {self.hasil['sky_brightness_tel_nl']:.4e} nL")
            print(f"  Weber Contrast (C_obj): {self.hasil['rasio_kontras_tel']:.4e}")
            if 'c_th_tel' in self.hasil:
                print(f"  Threshold (Crumey)    : {self.hasil['c_th_tel']:.4e}")
            print(f"  Visib. Margin (D_m)   : {self.hasil['delta_m_tel']:.4f}")
            if 'telescope_gain' in self.hasil:
                print(f"  Telescope Gain        : {self.hasil['telescope_gain']:+.4f} mag "
                      f"(= 2.5 log C_th_ne/C_th_tel)")
            status_tel = "TERLIHAT" if self.hasil['delta_m_tel'] > 0 else "TIDAK TERLIHAT"
            print(f"  Status                : {status_tel}")
        
        # Tambahkan hasil optimal teleskop jika mode optimal
        if self.hasil.get('mode') == 'optimal' and self.hasil.get('optimal_time_tel'):
            print(f"\n  --- Optimal ---")
            print(f"  Waktu Optimal         : {self.hasil['optimal_time_tel'].strftime('%H:%M:%S')}")
            print(f"  Moon Alt. Optimal     : {self.hasil['optimal_moon_alt_tel']:.2f}°")
            print(f"  Delta m Optimal       : {self.hasil['optimal_delta_m_tel']:.4f}")
            if 'optimal_telescope_gain' in self.hasil:
                print(f"  Telescope Gain        : {self.hasil['optimal_telescope_gain']:+.4f} mag")
            if self.hasil['visibility_duration_tel'] > 0:
                print(f"  Durasi Visibilitas    : {self.hasil['visibility_duration_tel']} menit (window kontinu terpanjang)")
                ws_tel = self.hasil.get('best_window_start_tel')
                we_tel = self.hasil.get('best_window_end_tel')
                if ws_tel and we_tel:
                    print(f"  Window Visibilitas    : {ws_tel.strftime('%H:%M:%S')} - {we_tel.strftime('%H:%M:%S')}")
            else:
                print(f"  Durasi Visibilitas    : 0 menit (tidak pernah delta_m > 0)")
        
        if self.hasil.get('delta_m_tel', -1) > 0:
            print(f"\n  >> Hilal TERDETEKSI dengan teleskop")
        else:
            print(f"\n  >> Hilal SULIT terdeteksi dengan teleskop")
        
        # Total timesteps (jika mode optimal)
        if self.hasil.get('mode') == 'optimal':
            print(f"\n  Total timesteps       : {self.hasil.get('total_timesteps', 0)}")
        
        print("\n" + "=" * 70)

    def plot_visibility_margin(self, save_path: Optional[str] = None) -> bool:
        """
        Plot grafik visibilitas margin (naked eye & teleskop) vs timestep.

        Sumbu X: nomor timestep (interval label 1)
        Sumbu Y: visibility margin / delta_m (interval label 5)

        Parameters
        ----------
        save_path : str or None
            Path file untuk menyimpan gambar. Jika None, hanya tampilkan.

        Returns
        -------
        bool
            True jika plot berhasil dibuat, False jika tidak ada data.
        """
        all_results = self.hasil.get('all_timestep_results', [])
        if not all_results:
            print("  [!] Tidak ada data timestep. Plot hanya tersedia pada mode optimal.")
            return False

        try:
            import matplotlib
            matplotlib.use('Agg')  # non-interactive backend
            import matplotlib.pyplot as plt
            import matplotlib.ticker as ticker
        except ImportError:
            print("  [!] matplotlib belum terinstall. Jalankan: pip install matplotlib")
            return False

        # Kumpulkan data valid
        timesteps = []
        dm_ne_list = []
        dm_tel_list = []
        waktu_labels = []
        idx = 0
        for r in all_results:
            if not r.get('valid'):
                continue
            timesteps.append(idx)
            dm_ne_list.append(r['delta_m_ne'])
            dm_tel_list.append(r['delta_m_tel'])
            waktu_labels.append(r['waktu_local'].strftime('%H:%M'))
            idx += 1

        if not timesteps:
            print("  [!] Tidak ada data valid untuk diplot.")
            return False

        fig, ax = plt.subplots(figsize=(12, 6))

        # Plot kedua garis
        ax.plot(timesteps, dm_ne_list, color='#2196F3', linewidth=2,
                marker='o', markersize=4, label='Margin Naked Eye (Δm NE)')
        ax.plot(timesteps, dm_tel_list, color='#F44336', linewidth=2,
                marker='s', markersize=4, label='Margin Teleskop (Δm Tel)')

        # Garis threshold delta_m = 0
        ax.axhline(y=0, color='#4CAF50', linewidth=1.5, linestyle='--',
                   label='Threshold (Δm = 0)', alpha=0.8)

        # Fill area di atas threshold
        ax.fill_between(timesteps, dm_tel_list, 0,
                        where=[v > 0 for v in dm_tel_list],
                        alpha=0.10, color='#F44336')
        ax.fill_between(timesteps, dm_ne_list, 0,
                        where=[v > 0 for v in dm_ne_list],
                        alpha=0.10, color='#2196F3')

        # Konfigurasi sumbu X — interval label 1, mulai tepat dari 0
        ax.set_xlabel('Timestep (menit setelah sunset)', fontsize=12, fontweight='bold')
        ax.xaxis.set_major_locator(ticker.MultipleLocator(1))
        ax.set_xlim(0, max(timesteps))

        # Konfigurasi sumbu Y — interval label 5, simetris terhadap 0
        ax.set_ylabel('Visibility Margin (Δm) [mag]', fontsize=12, fontweight='bold')
        ax.yaxis.set_major_locator(ticker.MultipleLocator(5))
        ax.yaxis.set_minor_locator(ticker.MultipleLocator(1))

        # Paksa sumbu Y simetris agar Δm = 0 selalu di tengah
        # Default: -20 s/d +20. Jika data melebihi 20, perluas ke kelipatan 5.
        all_dm = dm_ne_list + dm_tel_list
        y_abs_max = max(abs(min(all_dm)), abs(max(all_dm)))
        y_limit = max(20, math.ceil(y_abs_max / 5) * 5)
        ax.set_ylim(-y_limit, y_limit)

        # Format label sumbu Y: ...-10 -5 0 +5 +10...
        def y_formatter(val, pos):
            if val == 0:
                return '0'
            return f'{val:+.0f}'
        ax.yaxis.set_major_formatter(ticker.FuncFormatter(y_formatter))

        # Judul dan informasi
        nama = self.nama_tempat
        bln = self.bulan_hijri
        thn = self.tahun_hijri
        sunset_str = self.hasil.get('sunset_local', '')
        if hasattr(sunset_str, 'strftime'):
            sunset_str = sunset_str.strftime('%Y-%m-%d %H:%M:%S')
        ax.set_title(
            f'Visibilitas Hilal — {nama}\n'
            f'Bulan {bln}/{thn} H | Sunset: {sunset_str}',
            fontsize=14, fontweight='bold', pad=15
        )

        ax.legend(loc='best', fontsize=10, framealpha=0.9)
        ax.grid(True, which='major', linestyle='-', alpha=0.3)
        ax.grid(True, which='minor', linestyle=':', alpha=0.15)

        plt.tight_layout()

        if save_path:
            os.makedirs(os.path.dirname(save_path) or '.', exist_ok=True)
            fig.savefig(save_path, dpi=150, bbox_inches='tight')
            print(f"  [✓] Grafik disimpan: {save_path}")

        plt.close(fig)
        return True

    def _excel_styles(self):
        """Mengembalikan dictionary style untuk Excel."""
        return {
            'header_font': Font(name='Segoe UI', bold=True, size=12, color='FFFFFF'),
            'header_fill': PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid'),
            'section_font': Font(name='Segoe UI', bold=True, size=11, color='FFFFFF'),
            'section_fill': PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid'),
            'data_font': Font(name='Segoe UI', size=11),
            'data_font_bold': Font(name='Segoe UI', size=11, bold=True),
            'thin_border': Border(
                left=Side(style='thin', color='BDD7EE'),
                right=Side(style='thin', color='BDD7EE'),
                top=Side(style='thin', color='BDD7EE'),
                bottom=Side(style='thin', color='BDD7EE')
            ),
            'align_left': Alignment(horizontal='left', vertical='center'),
            'align_center': Alignment(horizontal='center', vertical='center'),
            'green_fill': PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid'),
            'red_fill': PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid'),
            'light_green_fill': PatternFill(start_color='E2F0D9', end_color='E2F0D9', fill_type='solid'),
        }

    def _excel_write_section_header(self, ws, row, title, styles, num_cols=2):
        """Menulis header seksi dengan formatting."""
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = styles['section_font']
        cell.fill = styles['section_fill']
        cell.alignment = styles['align_center']
        cell.border = styles['thin_border']
        for c in range(2, num_cols + 1):
            ws.cell(row=row, column=c).border = styles['thin_border']
            ws.cell(row=row, column=c).fill = styles['section_fill']
        return row + 1

    def _excel_write_data_row(self, ws, row, label, value, styles,
                               bold=False, is_status=False, status_positive=False,
                               merge_to_col=None):
        """Menulis satu baris data label-value, opsional merge B hingga kolom tertentu."""
        cell_label = ws.cell(row=row, column=1, value=label)
        cell_label.font = styles['data_font_bold'] if bold else styles['data_font']
        cell_label.alignment = styles['align_left']
        cell_label.border = styles['thin_border']

        if merge_to_col and merge_to_col > 2:
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=merge_to_col)

        cell_value = ws.cell(row=row, column=2, value=value)
        cell_value.font = styles['data_font_bold'] if bold else styles['data_font']
        if is_status:
            cell_value.alignment = styles['align_center']
            cell_value.fill = styles['green_fill'] if status_positive else styles['red_fill']
        else:
            cell_value.alignment = styles['align_left']
        cell_value.border = styles['thin_border']

        if merge_to_col:
            for c in range(3, merge_to_col + 1):
                ws.cell(row=row, column=c).border = styles['thin_border']

        return row + 1

    def _excel_write_comparison_row(self, ws, row, label, val_sunset, val_optimal, styles,
                                     bold=False, is_status=False,
                                     status_pos_sunset=False, status_pos_opt=False):
        """Menulis baris perbandingan 3 kolom: label | saat sunset | saat optimal."""
        font = styles['data_font_bold'] if bold else styles['data_font']

        cell_label = ws.cell(row=row, column=1, value=label)
        cell_label.font = font
        cell_label.alignment = styles['align_left']
        cell_label.border = styles['thin_border']

        cell_sunset = ws.cell(row=row, column=2, value=val_sunset)
        cell_sunset.font = font
        cell_sunset.alignment = styles['align_center']
        cell_sunset.border = styles['thin_border']

        cell_opt = ws.cell(row=row, column=3, value=val_optimal)
        cell_opt.font = font
        cell_opt.alignment = styles['align_center']
        cell_opt.border = styles['thin_border']

        if is_status:
            cell_sunset.fill = styles['green_fill'] if status_pos_sunset else styles['red_fill']
            if val_optimal != '-':
                cell_opt.fill = styles['green_fill'] if status_pos_opt else styles['red_fill']

        return row + 1

    def _write_ringkasan_sheet(self, ws, styles):
        """Menulis sheet Ringkasan ke worksheet dengan kolom terpisah sunset vs optimal."""
        ws.sheet_properties.tabColor = '2F5496'
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30

        is_optimal = self.hasil.get('mode') == 'optimal'
        nc = 3  # jumlah kolom

        # Style untuk sub-header kolom perbandingan
        sub_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        sub_font = Font(name='Segoe UI', bold=True, size=10, color='1F4E79')

        # Title
        ws.merge_cells('A1:C1')
        title_cell = ws.cell(row=1, column=1, value='HASIL PERHITUNGAN VISIBILITAS HILAL')
        title_cell.font = Font(name='Segoe UI', bold=True, size=16, color='FFFFFF')
        title_cell.fill = styles['header_fill']
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        for c in range(2, nc + 1):
            ws.cell(row=1, column=c).fill = styles['header_fill']
        ws.row_dimensions[1].height = 40

        # Shortcut helpers
        wsh = lambda r, t: self._excel_write_section_header(ws, r, t, styles, num_cols=nc)
        wdr = lambda r, l, v, **kw: self._excel_write_data_row(ws, r, l, v, styles, merge_to_col=nc, **kw)
        wcr = lambda r, l, vs, vo, **kw: self._excel_write_comparison_row(ws, r, l, vs, vo, styles, **kw)

        def write_sub_header(r):
            """Menulis sub-header kolom perbandingan."""
            headers = ['Parameter', 'Saat Sunset', 'Saat Optimal' if is_optimal else '-']
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=r, column=ci, value=h)
                cell.font = sub_font
                cell.alignment = styles['align_center']
                cell.border = styles['thin_border']
                cell.fill = sub_fill
            return r + 1

        row = 3
        na = '-'

        # ===================== INFORMASI LOKASI =====================
        row = wsh(row, '📍 INFORMASI LOKASI')
        row = wdr(row, 'Nama Tempat', self.nama_tempat)
        row = wdr(row, 'Lintang (°)', self.lintang)
        row = wdr(row, 'Bujur (°)', self.bujur)
        row = wdr(row, 'Elevasi (m)', self.elevasi)
        row = wdr(row, 'Timezone', self.timezone_str)
        row = wdr(row, 'Bulan/Tahun Hijri', f"{self.bulan_hijri}/{self.tahun_hijri}")
        row = wdr(row, 'Mode Perhitungan', self.hasil.get('mode', 'sunset').upper())
        row += 1

        # ===================== WAKTU =====================
        row = wsh(row, '🕐 WAKTU')
        if 'ijtima_utc' in self.hasil:
            row = wdr(row, 'Ijtima UTC', self.hasil['ijtima_utc'].strftime('%Y-%m-%d %H:%M:%S'))
            row = wdr(row, 'Ijtima Lokal', self.hasil['ijtima_local'].strftime('%Y-%m-%d %H:%M:%S'))
        if 'tanggal_pengamatan' in self.hasil:
            row = wdr(row, 'Tanggal Pengamatan', self.hasil['tanggal_pengamatan'].strftime('%Y-%m-%d'))
        if 'sunset_utc' in self.hasil:
            row = wdr(row, 'Sunset UTC', self.hasil['sunset_utc'].strftime('%Y-%m-%d %H:%M:%S'))
            row = wdr(row, 'Sunset Lokal', self.hasil['sunset_local'].strftime('%Y-%m-%d %H:%M:%S'))
        row += 1

        # ===================== DATA ATMOSFER (Saat Sunset) =====================
        src_short = self.SUMBER_ATMOSFER_LABEL.get(self.sumber_atmosfer, self.sumber_atmosfer.upper()).split(' ')[0]
        row = wsh(row, '🌡️ DATA ATMOSFER (Saat Sunset)')
        row = wdr(row, 'Sumber Data', self.SUMBER_ATMOSFER_LABEL.get(self.sumber_atmosfer, self.sumber_atmosfer))
        bias_t = self.hasil.get('bias_t', 0.0)
        bias_rh = self.hasil.get('bias_rh', 0.0)
        if 'rh' in self.hasil:
            if bias_t != 0.0 or bias_rh != 0.0:
                row = wdr(row, f'Suhu {src_short} Raw (°C)', f"{self.hasil.get('temperature_raw', 0):.2f}")
                row = wdr(row, f'Suhu Terkoreksi (°C) [bias={bias_t:+.1f}°C]', f"{self.hasil['temperature']:.2f}")
                row = wdr(row, f'RH {src_short} Raw (%)', f"{self.hasil.get('rh_raw', 0):.2f}")
                row = wdr(row, f'RH Terkoreksi (%) [bias={bias_rh:+.1f}%]', f"{self.hasil['rh']:.2f}")
            else:
                row = wdr(row, 'Kelembapan Relatif / RH (%)', f"{self.hasil['rh']:.2f}")
                row = wdr(row, 'Suhu / T (°C)', f"{self.hasil['temperature']:.2f}")
        if 'pressure' in self.hasil:
            row = wdr(row, 'Tekanan Udara / P (mbar)', f"{self.hasil['pressure']:.2f}")
        if 'k_v' in self.hasil:
            row = wdr(row, 'Koefisien Ekstingsi (k_V)', f"{self.hasil['k_v']:.4f}")
        row += 1

        # ===================== POSISI MATAHARI (Saat Sunset) =====================
        row = wsh(row, '☀️ POSISI MATAHARI (Saat Sunset)')
        if 'sun_alt' in self.hasil:
            row = wdr(row, 'Altitude Matahari', deg_to_dms(self.hasil['sun_alt']))
            row = wdr(row, 'Altitude Matahari (°)', f"{self.hasil['sun_alt']:.6f}")
            row = wdr(row, 'Azimuth Matahari', deg_to_dms(self.hasil['sun_az']))
            row = wdr(row, 'Azimuth Matahari (°)', f"{self.hasil['sun_az']:.6f}")
        row += 1

        # ===================== POSISI BULAN (Saat Sunset) =====================
        row = wsh(row, '🌙 POSISI BULAN (Saat Sunset)')
        if 'moon_alt' in self.hasil:
            row = wdr(row, 'Altitude Bulan', deg_to_dms(self.hasil['moon_alt']))
            row = wdr(row, 'Altitude Bulan (°)', f"{self.hasil['moon_alt']:.6f}")
            row = wdr(row, 'Azimuth Bulan', deg_to_dms(self.hasil['moon_az']))
            row = wdr(row, 'Azimuth Bulan (°)', f"{self.hasil['moon_az']:.6f}")
        if 'elongation' in self.hasil:
            row = wdr(row, 'Elongasi Toposentrik', deg_to_dms(self.hasil['elongation']))
            row = wdr(row, 'Elongasi Toposentrik (°)', f"{self.hasil['elongation']:.6f}")
        if 'phase_angle' in self.hasil:
            row = wdr(row, 'Phase Angle', deg_to_dms(self.hasil['phase_angle']))
            row = wdr(row, 'Phase Angle (°)', f"{self.hasil['phase_angle']:.6f}")
        if 'moon_semidiameter' in self.hasil:
            row = wdr(row, 'Semidiameter Bulan', deg_to_dms(float(self.hasil['moon_semidiameter'])))
            row = wdr(row, 'Semidiameter Bulan (°)', f"{float(self.hasil['moon_semidiameter']):.6f}")
        row += 1

        # ===================== VISIBILITAS NAKED EYE =====================
        row = wsh(row, '👁️ VISIBILITAS HILAL NAKED EYE')
        row = write_sub_header(row)

        opt_ne = self.hasil.get('optimal_result_ne') if is_optimal else None
        has_opt_ne = opt_ne is not None and opt_ne.get('valid', False)

        sunset_time = self.hasil['sunset_local'].strftime('%H:%M:%S') if 'sunset_local' in self.hasil else na

        row = wcr(row, 'Waktu', sunset_time,
                  opt_ne['waktu_local'].strftime('%H:%M:%S') if has_opt_ne else na)

        row = wcr(row, 'Altitude Bulan (°)',
                  f"{self.hasil['moon_alt']:.4f}" if 'moon_alt' in self.hasil else na,
                  f"{opt_ne['moon_alt']:.4f}" if has_opt_ne else na)

        row = wcr(row, 'Altitude Matahari (°)',
                  f"{self.hasil['sun_alt']:.4f}" if 'sun_alt' in self.hasil else na,
                  f"{opt_ne['sun_alt']:.4f}" if has_opt_ne else na)

        row = wcr(row, 'Elongasi (°)',
                  f"{self.hasil['elongation']:.4f}" if 'elongation' in self.hasil else na,
                  f"{opt_ne['elongation']:.4f}" if has_opt_ne else na)

        row = wcr(row, 'RH (%)',
                  f"{self.hasil['rh']:.2f}" if 'rh' in self.hasil else na,
                  f"{opt_ne['rh']:.2f}" if has_opt_ne and opt_ne.get('rh') is not None else na)

        row = wcr(row, 'Suhu (°C)',
                  f"{self.hasil['temperature']:.2f}" if 'temperature' in self.hasil else na,
                  f"{opt_ne['temperature']:.2f}" if has_opt_ne and opt_ne.get('temperature') is not None else na)

        row = wcr(row, 'Sky Brightness (nL)',
                  f"{self.hasil['sky_brightness_nl']:.4e}" if 'sky_brightness_nl' in self.hasil else na,
                  f"{opt_ne['sky_brightness_nl']:.4e}" if has_opt_ne else na)

        row = wcr(row, 'Luminansi Hilal (nL)',
                  f"{self.hasil['luminansi_hilal_nl']:.4e}" if 'luminansi_hilal_nl' in self.hasil else na,
                  f"{opt_ne['luminansi_hilal_nl']:.4e}" if has_opt_ne else na)

        row = wcr(row, 'Koefisien Ekstingsi (k_V)',
                  f"{self.hasil['k_v']:.4f}" if 'k_v' in self.hasil else na,
                  f"{opt_ne['k_v']:.4f}" if has_opt_ne else na)

        row = wcr(row, 'Weber Contrast (C_obj)',
                  f"{self.hasil['rasio_kontras_ne']:.4e}" if 'rasio_kontras_ne' in self.hasil else na,
                  f"{opt_ne['rasio_kontras_ne']:.4e}" if has_opt_ne else na)

        row = wcr(row, 'Threshold Crumey (C_th)',
                  f"{self.hasil['crumey_ne_C_th']:.4e}" if 'crumey_ne_C_th' in self.hasil else na,
                  f"{opt_ne.get('crumey_ne_C_th', 0):.4e}" if has_opt_ne and opt_ne.get('crumey_ne_C_th') is not None else na)

        row = wcr(row, 'Regime',
                  self.hasil.get('crumey_ne_regime', na),
                  opt_ne.get('crumey_ne_regime', na) if has_opt_ne else na)

        row = wcr(row, 'Luas Sabit (arcmin²)',
                  f"{self.hasil['crumey_ne_A_arcmin2']:.4f}" if 'crumey_ne_A_arcmin2' in self.hasil else na,
                  f"{opt_ne.get('crumey_ne_A_arcmin2', 0):.4f}" if has_opt_ne and opt_ne.get('crumey_ne_A_arcmin2') is not None else na)

        row = wcr(row, 'Visib. Margin (D_m)',
                  f"{self.hasil['delta_m_ne']:.4f}" if 'delta_m_ne' in self.hasil else na,
                  f"{opt_ne['delta_m_ne']:.4f}" if has_opt_ne else na)

        # Status NE
        is_sunset_ne = self.hasil.get('delta_m_ne', -1) >= 0
        is_opt_ne_vis = has_opt_ne and opt_ne.get('delta_m_ne', -1) >= 0
        row = wcr(row, 'Status',
                  "TERLIHAT" if is_sunset_ne else "TIDAK TERLIHAT",
                  "TERLIHAT" if is_opt_ne_vis else ("TIDAK TERLIHAT" if has_opt_ne else na),
                  bold=True, is_status=True,
                  status_pos_sunset=is_sunset_ne, status_pos_opt=is_opt_ne_vis)

        # Durasi visibilitas (hanya bermakna pada mode optimal)
        if is_optimal:
            dur_ne = self.hasil.get('visibility_duration_ne', 0)
            row = wcr(row, 'Durasi Visibilitas (menit)', na, str(dur_ne))
            # Window visibilitas kontinu terpanjang
            ws_ne = self.hasil.get('best_window_start_ne')
            we_ne = self.hasil.get('best_window_end_ne')
            if ws_ne and we_ne:
                window_str = f"{ws_ne.strftime('%H:%M:%S')} - {we_ne.strftime('%H:%M:%S')}"
            else:
                window_str = '-'
            row = wcr(row, 'Window Visibilitas', na, window_str)

        # Interpretasi keseluruhan
        overall_ne = is_sunset_ne or is_opt_ne_vis
        if overall_ne:
            row = wdr(row, 'Interpretasi', '✓ Hilal BERPOTENSI terlihat dengan mata telanjang', bold=True)
        else:
            row = wdr(row, 'Interpretasi', '✗ Hilal SULIT terlihat dengan mata telanjang', bold=True)
        row += 1

        # ===================== VISIBILITAS TELESKOP =====================
        row = wsh(row, '🔭 VISIBILITAS HILAL TELESKOP')
        row = write_sub_header(row)

        opt_tel = self.hasil.get('optimal_result_tel') if is_optimal else None
        has_opt_tel = opt_tel is not None and opt_tel.get('valid', False)

        row = wcr(row, 'Waktu', sunset_time,
                  opt_tel['waktu_local'].strftime('%H:%M:%S') if has_opt_tel else na)

        row = wcr(row, 'Altitude Bulan (°)',
                  f"{self.hasil['moon_alt']:.4f}" if 'moon_alt' in self.hasil else na,
                  f"{opt_tel['moon_alt']:.4f}" if has_opt_tel else na)

        row = wcr(row, 'Altitude Matahari (°)',
                  f"{self.hasil['sun_alt']:.4f}" if 'sun_alt' in self.hasil else na,
                  f"{opt_tel['sun_alt']:.4f}" if has_opt_tel else na)

        row = wcr(row, 'Elongasi (°)',
                  f"{self.hasil['elongation']:.4f}" if 'elongation' in self.hasil else na,
                  f"{opt_tel['elongation']:.4f}" if has_opt_tel else na)

        row = wcr(row, 'RH (%)',
                  f"{self.hasil['rh']:.2f}" if 'rh' in self.hasil else na,
                  f"{opt_tel['rh']:.2f}" if has_opt_tel and opt_tel.get('rh') is not None else na)

        row = wcr(row, 'Suhu (°C)',
                  f"{self.hasil['temperature']:.2f}" if 'temperature' in self.hasil else na,
                  f"{opt_tel['temperature']:.2f}" if has_opt_tel and opt_tel.get('temperature') is not None else na)

        row = wcr(row, 'Luminansi Hilal Teleskop (nL)',
                  f"{self.hasil['luminansi_hilal_tel_nl']:.4e}" if 'luminansi_hilal_tel_nl' in self.hasil else na,
                  f"{opt_tel.get('luminansi_hilal_tel_nl', 0):.4e}" if has_opt_tel and 'luminansi_hilal_tel_nl' in opt_tel else na)

        row = wcr(row, 'Sky Brightness Teleskop (nL)',
                  f"{self.hasil['sky_brightness_tel_nl']:.4e}" if 'sky_brightness_tel_nl' in self.hasil else na,
                  f"{opt_tel.get('sky_brightness_tel_nl', 0):.4e}" if has_opt_tel and 'sky_brightness_tel_nl' in opt_tel else na)

        row = wcr(row, 'Weber Contrast (C_obj)',
                  f"{self.hasil['rasio_kontras_tel']:.4e}" if 'rasio_kontras_tel' in self.hasil else na,
                  f"{opt_tel['rasio_kontras_tel']:.4e}" if has_opt_tel else na)

        if 'c_th_tel' in self.hasil:
            row = wcr(row, 'Threshold Crumey (C_th)',
                      f"{self.hasil['c_th_tel']:.4e}",
                      f"{opt_tel['c_th_tel']:.4e}" if has_opt_tel and 'c_th_tel' in opt_tel else na)

        row = wcr(row, 'Visib. Margin (D_m)',
                  f"{self.hasil['delta_m_tel']:.4f}" if 'delta_m_tel' in self.hasil else na,
                  f"{opt_tel['delta_m_tel']:.4f}" if has_opt_tel else na)

        row = wcr(row, 'Telescope Gain (mag)',
                  f"{self.hasil['telescope_gain']:+.4f}" if 'telescope_gain' in self.hasil else na,
                  f"{opt_tel['telescope_gain']:+.4f}" if has_opt_tel and 'telescope_gain' in opt_tel else na)

        # Status Teleskop
        is_sunset_tel = self.hasil.get('delta_m_tel', -1) > 0
        is_opt_tel_vis = has_opt_tel and opt_tel.get('delta_m_tel', -1) > 0
        row = wcr(row, 'Status',
                  "TERLIHAT" if is_sunset_tel else "TIDAK TERLIHAT",
                  "TERLIHAT" if is_opt_tel_vis else ("TIDAK TERLIHAT" if has_opt_tel else na),
                  bold=True, is_status=True,
                  status_pos_sunset=is_sunset_tel, status_pos_opt=is_opt_tel_vis)

        # Durasi dan timesteps (hanya mode optimal)
        if is_optimal:
            dur_tel = self.hasil.get('visibility_duration_tel', 0)
            row = wcr(row, 'Durasi Visibilitas (menit)', na, str(dur_tel))
            # Window visibilitas kontinu terpanjang
            ws_tel = self.hasil.get('best_window_start_tel')
            we_tel = self.hasil.get('best_window_end_tel')
            if ws_tel and we_tel:
                window_str_tel = f"{ws_tel.strftime('%H:%M:%S')} - {we_tel.strftime('%H:%M:%S')}"
            else:
                window_str_tel = '-'
            row = wcr(row, 'Window Visibilitas', na, window_str_tel)
            row = wcr(row, 'Total Timesteps', na, str(self.hasil.get('total_timesteps', 0)))

        # Interpretasi keseluruhan
        overall_tel = is_sunset_tel or is_opt_tel_vis
        if overall_tel:
            row = wdr(row, 'Interpretasi', '✓ Hilal TERDETEKSI dengan teleskop', bold=True)
        else:
            row = wdr(row, 'Interpretasi', '✗ Hilal SULIT terdeteksi dengan teleskop', bold=True)

        ws.freeze_panes = 'A2'

    def _write_timestep_sheet(self, ws, styles):
        """Menulis sheet Timestep Data ke worksheet."""
        ws.sheet_properties.tabColor = '548235'

        timestep_headers = [
            'No', 'Waktu Lokal', 'Moon Alt (°)', 'Sun Alt (°)',
            'Elongasi (°)', 'RH (%)', 'T (°C)',
            'Sky Brightness (nL)', 'Luminansi Hilal (nL)', 'k_V',
            'Δm Naked Eye', 'Margin Teleskop', 'Telescope Gain (mag)',
            'Kontras NE', 'C_obj Teleskop', 'C_th Teleskop'
        ]

        for col_idx, header in enumerate(timestep_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.alignment = styles['align_center']
            cell.border = styles['thin_border']

        all_results = self.hasil.get('all_timestep_results', [])
        if all_results:
            for i, r in enumerate(all_results, 1):
                row_data = [
                    i,
                    r['waktu_local'].strftime('%H:%M:%S') if r.get('waktu_local') else '',
                    round(r.get('moon_alt', 0), 4),
                    round(r.get('sun_alt', 0), 4),
                    round(r.get('elongation', 0), 4) if r.get('valid') else '',
                    round(r.get('rh', 0), 2) if r.get('rh') is not None else '',
                    round(r.get('temperature', 0), 2) if r.get('temperature') is not None else '',
                    f"{r.get('sky_brightness_nl', 0):.4e}" if r.get('valid') else '',
                    f"{r.get('luminansi_hilal_nl', 0):.4e}" if r.get('valid') else '',
                    round(r.get('k_v', 0), 4) if r.get('valid') else '',
                    round(r.get('delta_m_ne', 0), 4) if r.get('valid') else '',
                    round(r.get('delta_m_tel', 0), 4) if r.get('valid') else '',
                    round(r.get('telescope_gain', 0), 4) if r.get('valid') else '',
                    f"{r.get('rasio_kontras_ne', 0):.4e}" if r.get('valid') else '',
                    f"{r.get('rasio_kontras_tel', 0):.4e}" if r.get('valid') else '',
                    f"{r.get('c_th_tel', 0):.4e}" if r.get('valid') and 'c_th_tel' in r else ''
                ]

                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=i + 1, column=col_idx, value=val)
                    cell.font = styles['data_font']
                    cell.alignment = styles['align_center']
                    cell.border = styles['thin_border']

                    if r.get('valid') and r.get('delta_m_tel', -1) > 0:
                        cell.fill = styles['green_fill']
                    elif r.get('valid') and r.get('delta_m_ne', -1) > 0:
                        cell.fill = styles['light_green_fill']
        else:
            ws.cell(row=2, column=1, value='Tidak ada data timestep (mode sunset)').font = styles['data_font']

        # Auto-fit column widths
        for col_idx in range(1, len(timestep_headers) + 1):
            max_len = len(str(timestep_headers[col_idx - 1]))
            for row_idx in range(2, min(len(all_results) + 2, 100)):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 25)

        ws.freeze_panes = 'A2'

    def _write_info_sheet(self, ws, styles):
        """Menulis sheet Info Program ke worksheet."""
        ws.sheet_properties.tabColor = 'BF8F00'
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50

        info_data = [
            ('Program', 'Perhitungan Visibilitas Hilal'),
            ('Model Sky Brightness', 'Schaefer (1993)'),
            ('Model Luminansi Hilal', 'Kastner (1976)'),
            ('Model Teleskop', 'Schaefer (1990)'),
            ('Sumber Data Atmosfer', self.SUMBER_ATMOSFER_LABEL.get(self.sumber_atmosfer, self.sumber_atmosfer)),
            ('Mode Perhitungan', self.hasil.get('mode', 'sunset')),
            ('Tanggal Eksekusi', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('Catatan Schaefer', 'Moonlight excluded (Bulan = objek pengamatan)'),
        ]

        for col_idx, header in enumerate(['Parameter', 'Nilai'], 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.alignment = styles['align_center']
            cell.border = styles['thin_border']

        for i, (param, nilai) in enumerate(info_data, 2):
            cell_p = ws.cell(row=i, column=1, value=param)
            cell_p.font = styles['data_font']
            cell_p.border = styles['thin_border']
            cell_v = ws.cell(row=i, column=2, value=nilai)
            cell_v.font = styles['data_font']
            cell_v.border = styles['thin_border']

        ws.freeze_panes = 'A2'

    def simpan_ke_excel(self, filepath: str) -> str:
        """
        Menyimpan seluruh hasil perhitungan ke file Excel (.xlsx) dengan format rapi.

        Excel terdiri dari 3 sheet:
        1. Ringkasan   - Seluruh data summary hasil perhitungan
        2. Timestep Data - Data per-timestep (jika mode optimal)
        3. Info Program  - Metadata program

        Parameters:
        -----------
        filepath : str
            Path lengkap file Excel yang akan disimpan

        Returns:
        --------
        str
            Path file yang berhasil disimpan
        """
        wb = Workbook()
        styles = self._excel_styles()

        # Sheet 1: Ringkasan
        ws1 = wb.active
        ws1.title = "Ringkasan"
        self._write_ringkasan_sheet(ws1, styles)

        # Sheet 2: Timestep Data
        ws2 = wb.create_sheet(title="Timestep Data")
        self._write_timestep_sheet(ws2, styles)

        # Sheet 3: Info Program
        ws3 = wb.create_sheet(title="Info Program")
        self._write_info_sheet(ws3, styles)

        # Save
        output_dir = os.path.dirname(filepath)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)

        wb.save(filepath)
        print(f"\n  ✓ Hasil disimpan ke: {filepath}")

        return filepath


def tentukan_timezone_indonesia(longitude: float) -> str:
    """
    Menentukan timezone Indonesia berdasarkan bujur (longitude).
    
    Pembagian zona waktu Indonesia:
    - WIB  (UTC+7): bujur < 115° (Sumatera, Jawa, Kalimantan Barat & Tengah)
    - WITA (UTC+8): 115° <= bujur < 135° (Kalimantan Timur & Selatan, Sulawesi, Bali, Nusa Tenggara)
    - WIT  (UTC+9): bujur >= 135° (Maluku, Papua)
    
    Parameters:
    -----------
    longitude : float
        Bujur lokasi dalam derajat
        
    Returns:
    --------
    str
        Timezone IANA string
    """
    if longitude < 115.0:
        return "Asia/Jakarta"    # WIB (UTC+7)
    elif longitude < 135.0:
        return "Asia/Makassar"   # WITA (UTC+8)
    else:
        return "Asia/Jayapura"   # WIT (UTC+9)


def _get_input(prompt_text, default_val, type_func=float):
    """Helper untuk input dengan default value."""
    val = input(prompt_text).strip()
    if not val:
        return default_val
    try:
        return type_func(val)
    except ValueError:
        print(f"    [!] Input '{val}' tidak valid. Menggunakan default {default_val}.")
        return default_val


def _input_bulan_tahun_hijri():
    """Input bulan dan tahun Hijriah dari user. Return (bulan, tahun) atau None."""
    nama_bulan = {
        1: "Muharram", 2: "Shafar", 3: "Rabiul Awal", 4: "Rabiul Akhir",
        5: "Jumadil Awal", 6: "Jumadil Akhir", 7: "Rajab", 8: "Sya'ban",
        9: "Ramadhan", 10: "Syawwal", 11: "Dzulqa'dah", 12: "Dzulhijjah"
    }

    print("\n--- LANGKAH 2: INPUT BULAN DAN TAHUN HIJRIAH ---")
    print("  Daftar bulan Hijriah:")
    for num, nama in nama_bulan.items():
        print(f"    {num:2d}. {nama}")

    try:
        bulan_hijri = int(input("\n  Masukkan nomor bulan Hijriah (1-12): "))
        if bulan_hijri < 1 or bulan_hijri > 12:
            print("[!] Bulan tidak valid. Program dihentikan.")
            return None

        tahun_hijri = int(input("  Masukkan tahun Hijriah (contoh: 1446): "))
        if tahun_hijri < 1:
            print("[!] Tahun tidak valid. Program dihentikan.")
            return None

        print(f"\n  ✓ Bulan Hijriah: {nama_bulan[bulan_hijri]} {tahun_hijri} H")
        return bulan_hijri, tahun_hijri

    except ValueError:
        print("[!] Input harus berupa angka. Program dihentikan.")
        return None


def _input_mode_dan_offset():
    """Input mode perhitungan dan offset hari. Return (mode, delta_day)."""
    print("\n--- LANGKAH 3: PILIH MODE PERHITUNGAN ---")
    print("  1. Sunset  - Hitung visibilitas pada saat sunset saja")
    print("  2. Optimal - Loop dari sunset untuk mencari delta_m maksimum")

    try:
        mode_pilihan = int(input("\n  Pilih mode (1/2): "))
        mode = "sunset" if mode_pilihan == 1 else "optimal"
        print(f"  ✓ Mode: {mode}")
    except ValueError:
        mode = "optimal"
        print(f"  ✓ Mode default: {mode}")

    print("\n--- LANGKAH 3.5: PILIH WAKTU PENGAMATAN ---")
    print("  0. Sesuai Hisab (H)       - Default")
    print("  1. H + 1 Hari             - Besoknya")
    print("  2. H + 2 Hari             - Lusanya")
    print(" -1. H - 1 Hari             - Kemarin")
    print("  Catatan: ERA5 (1940-sekarang), MERRA-2 (1981-sekarang), BMKG (3 hari ke depan)")

    try:
        offset_pilihan = input("\n  Pilih waktu (-2/-1/0/1/2) [enter=0]: ").strip()
        delta_day = int(offset_pilihan) if offset_pilihan else 0
        print(f"  ✓ Offset waktu: H + {delta_day} hari")
    except ValueError:
        delta_day = 0
        print(f"  ✓ Offset default: H + 0 hari")

    return mode, delta_day


def _input_sumber_atmosfer(adm4_code: str = ''):
    """Input sumber data atmosfer. Return (sumber, manual_rh, manual_t, manual_p)."""
    print("\n--- LANGKAH 3.6: SUMBER DATA ATMOSFER ---")
    print("  Pilih sumber data atmosfer (RH, T, P):")
    print("  1. ERA5 Reanalysis  (Open-Meteo API)  — data historis 1940-sekarang")
    print("  2. MERRA-2          (NASA POWER API)   — data historis 1981-sekarang")
    if adm4_code:
        print(f"  3. BMKG Forecast    (API Prakiraan)    — prakiraan 3 hari [kode: {adm4_code}]")
    else:
        print("  3. BMKG Forecast    (API Prakiraan)    — prakiraan 3 hari (butuh kode wilayah)")
    print("  4. Input Manual     (tanpa API)        — masukkan RH, T, P secara manual")

    manual_rh, manual_t, manual_p = 80.0, 25.0, 1013.25

    try:
        pilihan = input("\n  Pilih sumber (1/2/3/4) [enter=1]: ").strip() or "1"

        if pilihan == "2":
            sumber = 'merra2'
            print(f"  ✓ Sumber atmosfer: MERRA-2 (NASA POWER API)")
        elif pilihan == "3":
            sumber = 'bmkg'
            if not adm4_code:
                adm4_input = input("  Masukkan kode wilayah BMKG (contoh: 33.74.10.1003): ").strip()
                if adm4_input:
                    adm4_code = adm4_input
                else:
                    print("  [!] Kode wilayah kosong. Fallback ke ERA5.")
                    sumber = 'era5'
            if sumber == 'bmkg':
                print(f"  ✓ Sumber atmosfer: BMKG Prakiraan Cuaca [kode: {adm4_code}]")
                print(f"  ⚠ BMKG tidak menyediakan tekanan udara — tekanan diestimasi dari elevasi.")
        elif pilihan == "4":
            sumber = 'manual'
            print("  Masukkan data atmosfer secara manual:")
            try:
                rh_input = input("    RH (%) [default=80.0]: ").strip()
                manual_rh = float(rh_input) if rh_input else 80.0
                t_input = input("    Suhu (°C) [default=25.0]: ").strip()
                manual_t = float(t_input) if t_input else 25.0
                p_input = input("    Tekanan (mbar) [default=1013.25]: ").strip()
                manual_p = float(p_input) if p_input else 1013.25
            except ValueError:
                print("  [!] Input tidak valid. Menggunakan nilai default.")
                manual_rh, manual_t, manual_p = 80.0, 25.0, 1013.25
            print(f"  ✓ Sumber atmosfer: Input Manual")
            print(f"    RH={manual_rh:.2f}%, T={manual_t:.2f}°C, P={manual_p:.2f} mbar")
        else:
            sumber = 'era5'
            print(f"  ✓ Sumber atmosfer: ERA5 Reanalysis (Open-Meteo API)")
    except EOFError:
        sumber = 'era5'
        print(f"  ✓ Sumber default: ERA5 Reanalysis")

    return sumber, adm4_code, manual_rh, manual_t, manual_p


def _input_koreksi_bias(bias_t: float, bias_rh: float, sumber_atmosfer: str = 'era5'):
    """Input koreksi bias reanalisis. Return (bias_t, bias_rh)."""
    # Koreksi bias hanya relevan untuk sumber API (bukan manual)
    if sumber_atmosfer == 'manual':
        print("\n--- LANGKAH 3.7: KOREKSI BIAS ---")
        print("  ✓ Koreksi bias dilewati (sumber: Input Manual)")
        return 0.0, 0.0

    label = sumber_atmosfer.upper().replace('MERRA2', 'MERRA-2')
    print(f"\n--- LANGKAH 3.7: KOREKSI BIAS {label} ---")
    print(f"  Koreksi bias digunakan untuk menyesuaikan data {label} dengan data observasi.")
    print(f"  Definisi: bias = {label} - Observasi")
    print(f"  Contoh: Jika {label} 30°C dan Observasi 28°C, maka bias_t = +2.0")
    print()

    if bias_t != 0.0 or bias_rh != 0.0:
        print(f"  [INFO] Lokasi ini memiliki data bias bawaan:")
        print(f"           bias_t = {bias_t:+.1f}°C")
        print(f"           bias_rh = {bias_rh:+.1f}%")
        print()

    print("  Pilih opsi koreksi bias:")
    print("  1. Gunakan data bawaan lokasi (jika ada)")
    print("  2. Tanpa koreksi (bias_t = 0, bias_rh = 0)")
    print("  3. Input manual nilai bias")

    try:
        bias_pilihan = input("\n  Pilih opsi (1/2/3) [enter=1]: ").strip() or "1"

        if bias_pilihan == "2":
            bias_t, bias_rh = 0.0, 0.0
            print(f"  ✓ Menggunakan data {label} tanpa koreksi")
        elif bias_pilihan == "3":
            try:
                bias_t_input = input("  Masukkan bias suhu (°C) [contoh: +1.5 atau -0.5, enter=0]: ").strip()
                bias_t = float(bias_t_input) if bias_t_input else 0.0
                bias_rh_input = input("  Masukkan bias RH (%) [contoh: +5.0 atau -3.0, enter=0]: ").strip()
                bias_rh = float(bias_rh_input) if bias_rh_input else 0.0
                print(f"  ✓ Koreksi bias: T={bias_t:+.1f}°C, RH={bias_rh:+.1f}%")
            except ValueError:
                print("  [!] Input tidak valid. Menggunakan default tanpa koreksi.")
                bias_t, bias_rh = 0.0, 0.0
        else:
            if bias_t == 0.0 and bias_rh == 0.0:
                print(f"  ✓ Lokasi tidak memiliki data bias. Menggunakan tanpa koreksi.")
            else:
                print(f"  ✓ Menggunakan data bawaan: T={bias_t:+.1f}°C, RH={bias_rh:+.1f}%")
    except ValueError:
        if bias_t == 0.0 and bias_rh == 0.0:
            print(f"  ✓ Lokasi tidak memiliki data bias. Menggunakan tanpa koreksi.")
        else:
            print(f"  ✓ Menggunakan data bawaan: T={bias_t:+.1f}°C, RH={bias_rh:+.1f}%")

    return bias_t, bias_rh


def _input_parameter_teleskop():
    """Input parameter teleskop dan field factor. Return dict parameter."""
    defaults = {
        'aperture': 100.0, 'magnification': 50.0, 'central_obstruction': 0.0,
        'transmission': 0.95, 'n_surfaces': 6, 'observer_age': 22.0, 'seeing': 3.0,
        'field_factor': 2.4, 'F_naked': 2.5
    }

    print("\n--- LANGKAH 4: PARAMETER TELESKOP & FIELD FACTOR ---")
    print("  Gunakan parameter default?")
    print("  (aperture=100mm, magnification=50x, obstruction=0mm,")
    print("   transmission=0.95, n_surfaces=6, age=22.0, seeing=3.0,")
    print("   field_factor_teleskop=2.4, field_factor_naked_eye=2.5)")

    try:
        default_tel = input("  Gunakan default? (Y/n): ").strip().lower()
        if default_tel == 'n':
            print("\n  Masukkan parameter (kosongkan untuk default):")
            defaults['aperture'] = _get_input("  Aperture (mm) [default: 100.0]: ", 100.0)
            defaults['magnification'] = _get_input("  Magnifikasi (x) [default: 50.0]: ", 50.0)
            defaults['central_obstruction'] = _get_input("  Obstruksi sentral (mm) [default: 0.0]: ", 0.0)
            defaults['transmission'] = _get_input("  Transmisi per permukaan/lensa (0-1) [default: 0.95]: ", 0.95)
            defaults['n_surfaces'] = _get_input("  Jumlah permukaan optik [default: 6]: ", 6, int)
            defaults['observer_age'] = _get_input("  Usia pengamat (tahun) [default: 22.0]: ", 22.0)
            defaults['seeing'] = _get_input("  Seeing atmosfir (arcsec) [default: 3.0]: ", 3.0)
            defaults['field_factor'] = _get_input("  Field factor teleskop (Crumey F) [default: 2.4]: ", 2.4)
            defaults['F_naked'] = _get_input("  Field factor naked eye (Crumey F) [default: 2.5]: ", 2.5)
            print(f"  ✓ Parameter kustom digunakan.")
        else:
            print(f"  ✓ Parameter default digunakan.")
    except Exception as e:
        print(f"  [!] Terjadi kesalahan: {e}")
        print(f"  ✓ Parameter default digunakan.")

    return defaults


# ═══════════════════════════════════════════════════════════════════════════════
# MULTI-LOKASI: Fungsi pendukung untuk batch processing di mode interaktif
# ═══════════════════════════════════════════════════════════════════════════════

def _parse_range_input(input_str: str, max_val: int) -> list:
    """
    Parse input range string menjadi list of integers.
    Contoh: "1,3,5-10,15" → [1, 3, 5, 6, 7, 8, 9, 10, 15]

    Parameters:
    -----------
    input_str : str
        String input dari user (contoh: "1,3,5-10,15")
    max_val : int
        Nilai maksimum yang valid (jumlah lokasi)

    Returns:
    --------
    list[int]
        List nomor lokasi (1-indexed) yang valid dan unik, terurut
    """
    result = set()
    parts = input_str.replace(' ', '').split(',')
    for part in parts:
        if not part:
            continue
        if '-' in part:
            try:
                start, end = part.split('-', 1)
                start, end = int(start), int(end)
                for i in range(start, end + 1):
                    if 1 <= i <= max_val:
                        result.add(i)
            except ValueError:
                continue
        else:
            try:
                val = int(part)
                if 1 <= val <= max_val:
                    result.add(val)
            except ValueError:
                continue
    return sorted(result)


def _input_mode_lokasi() -> str:
    """
    Input mode lokasi: tunggal atau multi.

    Returns:
    --------
    str
        'single' atau 'multi'
    """
    print("\n--- LANGKAH 0: MODE LOKASI ---")
    print("  1. Lokasi Tunggal   — hitung visibilitas untuk satu lokasi")
    print("  2. Multi-Lokasi     — hitung visibilitas untuk banyak lokasi sekaligus")

    try:
        pilihan = input("\n  Pilih mode (1/2) [enter=1]: ").strip() or "1"
        if pilihan == "2":
            print("  ✓ Mode: Multi-Lokasi")
            return 'multi'
        else:
            print("  ✓ Mode: Lokasi Tunggal")
            return 'single'
    except (ValueError, EOFError):
        print("  ✓ Mode default: Lokasi Tunggal")
        return 'single'


def _input_multi_lokasi() -> list:
    """
    Pilih beberapa lokasi dari daftar_lokasi untuk batch processing.

    Returns:
    --------
    list[dict] or None
        List dictionary lokasi yang dipilih, atau None jika dibatalkan
    """
    from daftar_lokasi import print_daftar_lokasi, get_list_lokasi

    print("\n--- LANGKAH 1: PILIH LOKASI PENGAMATAN (MULTI) ---")
    list_lokasi = print_daftar_lokasi()
    n = len(list_lokasi)

    print(f"\n  Pilih lokasi yang akan dihitung:")
    print(f"  1. Semua lokasi ({n} lokasi)")
    print(f"  2. Pilih beberapa (contoh: 1,3,5-10,15)")

    try:
        pilihan = input("\n  Pilih opsi (1/2) [enter=1]: ").strip() or "1"

        if pilihan == "2":
            range_str = input(f"  Masukkan nomor lokasi (1-{n}): ").strip()
            if not range_str:
                print("  [!] Input kosong. Menggunakan semua lokasi.")
                indices = list(range(1, n + 1))
            else:
                indices = _parse_range_input(range_str, n)
                if not indices:
                    print("  [!] Tidak ada nomor valid. Menggunakan semua lokasi.")
                    indices = list(range(1, n + 1))
        else:
            indices = list(range(1, n + 1))

        selected = [list_lokasi[i - 1] for i in indices]

        print(f"\n  ✓ {len(selected)} lokasi dipilih:")
        for i, lok in enumerate(selected, 1):
            print(f"    {i:2d}. {lok['nama']}  ({lok['lat']:.4f}°, {lok['lon']:.4f}°)")

        return selected

    except (ValueError, EOFError):
        print("  [!] Input error. Menggunakan semua lokasi.")
        return list_lokasi


def _run_multi_lokasi(lokasi_list: list, shared_params: dict) -> list:
    """
    Jalankan perhitungan visibilitas hilal untuk banyak lokasi.

    Parameters:
    -----------
    lokasi_list : list[dict]
        Daftar lokasi (dari daftar_lokasi)
    shared_params : dict
        Parameter bersama: bulan_hijri, tahun_hijri, mode, delta_day,
        sumber_atmosfer, tel_params, F_naked, dll.

    Returns:
    --------
    list[dict]
        List hasil per lokasi, masing-masing berisi 'lokasi', 'hasil', 'calculator', 'success'
    """
    import time

    results = []
    total = len(lokasi_list)

    print("\n" + "█" * 70)
    print("  BATCH PROCESSING: Visibilitas Hilal Multi-Lokasi")
    print(f"  {total} lokasi × Bulan {shared_params['bulan_hijri']}/{shared_params['tahun_hijri']} H")
    print(f"  Mode: {shared_params['mode']}  |  Atmosfer: {shared_params['sumber_atmosfer']}")
    print("█" * 70)

    t_batch_start = time.time()

    for i, lokasi in enumerate(lokasi_list, 1):
        nama = lokasi['nama']
        lat = lokasi['lat']
        lon = lokasi['lon']
        elv = lokasi['elevasi']
        tz = tentukan_timezone_indonesia(lon)

        print(f"\n{'═' * 70}")
        print(f"  [{i:2d}/{total}] {nama}")
        print(f"         Lat={lat:.4f}°  Lon={lon:.4f}°  Elv={elv:.0f}m  TZ={tz}")
        print(f"{'═' * 70}")

        t_start = time.time()

        try:
            calculator = HilalVisibilityCalculator(
                nama_tempat=nama,
                lintang=lat,
                bujur=lon,
                elevasi=elv,
                timezone_str=tz,
                bulan_hijri=shared_params['bulan_hijri'],
                tahun_hijri=shared_params['tahun_hijri'],
                delta_day_offset=shared_params['delta_day'],
                bias_t=lokasi.get('bias_t', 0.0),
                bias_rh=lokasi.get('bias_rh', 0.0),
                sumber_atmosfer=shared_params['sumber_atmosfer'],
                adm4_code=lokasi.get('adm4_code', ''),
                manual_rh=shared_params.get('manual_rh', 80.0),
                manual_t=shared_params.get('manual_t', 25.0),
                manual_p=shared_params.get('manual_p', 1013.25),
            )

            hasil = calculator.jalankan_perhitungan_lengkap(
                use_telescope=True,
                mode=shared_params['mode'],
                min_moon_alt=2.0,
                F_naked=shared_params['F_naked'],
                **shared_params['tel_params'],
            )

            elapsed = time.time() - t_start

            # Ringkasan singkat
            dm_ne = hasil.get('delta_m_ne', -99)
            dm_tel = hasil.get('delta_m_tel', -99)
            if shared_params['mode'] == 'optimal':
                dm_ne = hasil.get('optimal_delta_m_ne', dm_ne)
                dm_tel = hasil.get('optimal_delta_m_tel', dm_tel)

            status_ne = "TERLIHAT" if dm_ne >= 0 else "TIDAK"
            status_tel = "TERDETEKSI" if dm_tel > 0 else "TIDAK"

            print(f"\n  RINGKASAN: Δm_NE={dm_ne:+.3f} ({status_ne})  "
                  f"Δm_Tel={dm_tel:+.3f} ({status_tel})  "
                  f"⏱ {elapsed:.1f}s")

            results.append({
                'lokasi': lokasi,
                'hasil': hasil,
                'calculator': calculator,
                'success': True,
                'elapsed': elapsed,
            })

        except Exception as e:
            elapsed = time.time() - t_start
            print(f"\n  ✗ ERROR: {e}")
            results.append({
                'lokasi': lokasi,
                'hasil': {},
                'calculator': None,
                'success': False,
                'elapsed': elapsed,
                'error': str(e),
            })

    total_time = time.time() - t_batch_start
    n_success = sum(1 for r in results if r['success'])
    print(f"\n{'█' * 70}")
    print(f"  BATCH SELESAI: {n_success}/{total} berhasil dalam {total_time:.0f} detik")
    print(f"{'█' * 70}")

    # Tabel ringkasan akhir
    _print_tabel_ringkasan_multi(results, shared_params)

    return results


def _print_tabel_ringkasan_multi(results: list, shared_params: dict):
    """Cetak tabel ringkasan hasil multi-lokasi ke terminal."""
    mode = shared_params['mode']
    is_opt = (mode == 'optimal')

    print(f"\n{'═' * 110}")
    print(f"  TABEL RINGKASAN — Bulan {shared_params['bulan_hijri']}/{shared_params['tahun_hijri']} H"
          f"  |  Mode: {mode}  |  Atmosfer: {shared_params['sumber_atmosfer'].upper()}")
    print(f"{'═' * 110}")

    header = (f"{'No':>3} | {'Lokasi':<35} | {'Lat':>9} | {'Lon':>9} | "
              f"{'Moon Alt':>8} | {'Elong':>7} | {'Δm NE':>8} | {'Δm Tel':>8} | "
              f"{'NE':>6} | {'Tel':>6}")
    print(header)
    print("-" * 110)

    for i, r in enumerate(results, 1):
        lok = r['lokasi']
        if not r['success']:
            print(f"{i:3d} | {lok['nama']:<35} | {lok['lat']:9.4f} | {lok['lon']:9.4f} | "
                  f"{'ERROR':>8} | {'':>7} | {'':>8} | {'':>8} | {'':>6} | {'':>6}")
            continue

        h = r['hasil']
        moon_alt = h.get('moon_alt', 0)
        elong = h.get('elongation', 0)

        if is_opt:
            dm_ne = h.get('optimal_delta_m_ne', h.get('delta_m_ne', -99))
            dm_tel = h.get('optimal_delta_m_tel', h.get('delta_m_tel', -99))
        else:
            dm_ne = h.get('delta_m_ne', -99)
            dm_tel = h.get('delta_m_tel', -99)

        st_ne = "Y" if dm_ne >= 0 else "N"
        st_tel = "Y" if dm_tel > 0 else "N"

        print(f"{i:3d} | {lok['nama']:<35} | {lok['lat']:9.4f} | {lok['lon']:9.4f} | "
              f"{moon_alt:8.3f} | {elong:7.3f} | {dm_ne:+8.3f} | {dm_tel:+8.3f} | "
              f"{st_ne:>6} | {st_tel:>6}")

    print(f"{'═' * 110}")

    # Hitung statistik
    success_results = [r for r in results if r['success']]
    if success_results:
        n_ne = 0
        n_tel = 0
        for r in success_results:
            h = r['hasil']
            if is_opt:
                dm_ne_val = h.get('optimal_delta_m_ne', h.get('delta_m_ne', -1))
                dm_tel_val = h.get('optimal_delta_m_tel', h.get('delta_m_tel', -1))
            else:
                dm_ne_val = h.get('delta_m_ne', -1)
                dm_tel_val = h.get('delta_m_tel', -1)
            if dm_ne_val >= 0:
                n_ne += 1
            if dm_tel_val > 0:
                n_tel += 1
        print(f"  Naked Eye terlihat: {n_ne}/{len(success_results)}  |  "
              f"Teleskop terdeteksi: {n_tel}/{len(success_results)}")


def _simpan_excel_multi(results: list, shared_params: dict, filepath: str) -> str:
    """
    Simpan hasil multi-lokasi ke file Excel gabungan.

    Sheet 1: "Ringkasan Multi-Lokasi" — tabel semua lokasi
    Sheet 2: "Info & Parameter" — konfigurasi yang digunakan

    Parameters:
    -----------
    results : list[dict]
        Hasil dari _run_multi_lokasi()
    shared_params : dict
        Parameter bersama
    filepath : str
        Path file Excel

    Returns:
    --------
    str
        Path file yang disimpan
    """
    wb = Workbook()
    mode = shared_params['mode']
    is_opt = (mode == 'optimal')

    # --- Style definitions ---
    header_font = Font(name='Segoe UI', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    data_font = Font(name='Segoe UI', size=10)
    data_font_bold = Font(name='Segoe UI', size=10, bold=True)
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='BDD7EE'),
        right=Side(style='thin', color='BDD7EE'),
        top=Side(style='thin', color='BDD7EE'),
        bottom=Side(style='thin', color='BDD7EE')
    )
    green_fill = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')
    red_fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

    # ═══════════════════════════════════════════════════════════════
    # Sheet 1: Ringkasan Multi-Lokasi
    # ═══════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Ringkasan Multi-Lokasi"
    ws.sheet_properties.tabColor = '1F4E79'

    headers = [
        'No', 'Lokasi', 'Lat (°)', 'Lon (°)', 'Elv (m)',
        'Sunset', 'Moon Alt (°)', 'Elongasi (°)', 'Phase Angle (°)',
        'RH (%)', 'T (°C)', 'k_V',
        'Sky Bright. (nL)', 'Lumin. Hilal (nL)',
        'Δm NE (sunset)', 'Δm Tel (sunset)',
    ]
    if is_opt:
        headers += [
            'Δm NE (optimal)', 'Δm Tel (optimal)',
            'Waktu Opt NE', 'Waktu Opt Tel',
            'Durasi NE (min)', 'Durasi Tel (min)',
            'Tel Gain (mag)',
        ]
    headers += ['Status NE', 'Status Tel']

    # Write headers
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    # Write data rows
    for i, r in enumerate(results, 1):
        row_idx = i + 1
        lok = r['lokasi']

        if not r['success']:
            row_data = [i, lok['nama'], lok['lat'], lok['lon'], lok['elevasi']]
            row_data += ['ERROR'] + [''] * (len(headers) - 6) + ['ERROR']
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.font = data_font
                cell.alignment = align_center if col != 2 else align_left
                cell.border = thin_border
                if val == 'ERROR':
                    cell.fill = red_fill
            continue

        h = r['hasil']

        # Tentukan delta_m final
        if is_opt:
            dm_ne_final = h.get('optimal_delta_m_ne', h.get('delta_m_ne', -99))
            dm_tel_final = h.get('optimal_delta_m_tel', h.get('delta_m_tel', -99))
        else:
            dm_ne_final = h.get('delta_m_ne', -99)
            dm_tel_final = h.get('delta_m_tel', -99)

        st_ne = "TERLIHAT" if dm_ne_final >= 0 else "TIDAK TERLIHAT"
        st_tel = "TERDETEKSI" if dm_tel_final > 0 else "TIDAK TERDETEKSI"

        sunset_str = ''
        sunset_val = h.get('sunset_local')
        if sunset_val and hasattr(sunset_val, 'strftime'):
            sunset_str = sunset_val.strftime('%H:%M:%S')

        row_data = [
            i, lok['nama'], lok['lat'], lok['lon'], lok['elevasi'],
            sunset_str,
            round(h.get('moon_alt', 0), 4),
            round(h.get('elongation', 0), 4),
            round(h.get('phase_angle', 0), 4),
            round(h.get('rh', 0), 2),
            round(h.get('temperature', 0), 2),
            round(h.get('k_v', 0), 4),
            f"{h.get('sky_brightness_nl', 0):.4e}",
            f"{h.get('luminansi_hilal_nl', 0):.4e}",
            round(h.get('delta_m_ne', -99), 4),
            round(h.get('delta_m_tel', -99), 4),
        ]

        if is_opt:
            opt_time_ne = h.get('optimal_time_ne')
            opt_time_tel = h.get('optimal_time_tel')
            row_data += [
                round(h.get('optimal_delta_m_ne', -99), 4),
                round(h.get('optimal_delta_m_tel', -99), 4),
                opt_time_ne.strftime('%H:%M:%S') if opt_time_ne and hasattr(opt_time_ne, 'strftime') else '',
                opt_time_tel.strftime('%H:%M:%S') if opt_time_tel and hasattr(opt_time_tel, 'strftime') else '',
                h.get('visibility_duration_ne', 0),
                h.get('visibility_duration_tel', 0),
                round(h.get('optimal_telescope_gain', h.get('telescope_gain', 0)), 4),
            ]

        row_data += [st_ne, st_tel]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = data_font
            cell.alignment = align_center if col != 2 else align_left
            cell.border = thin_border

        # Warnai kolom status
        col_st_ne = len(headers) - 1
        col_st_tel = len(headers)
        ws.cell(row=row_idx, column=col_st_ne).fill = green_fill if dm_ne_final >= 0 else red_fill
        ws.cell(row=row_idx, column=col_st_ne).font = data_font_bold
        ws.cell(row=row_idx, column=col_st_tel).fill = green_fill if dm_tel_final > 0 else red_fill
        ws.cell(row=row_idx, column=col_st_tel).font = data_font_bold

    # Auto-fit column widths
    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for row_idx in range(2, len(results) + 2):
            cell_val = ws.cell(row=row_idx, column=col).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 30)
    # Kolom lokasi lebih lebar
    ws.column_dimensions['B'].width = 38

    ws.freeze_panes = 'A2'

    # ═══════════════════════════════════════════════════════════════
    # Sheet 2: Info & Parameter
    # ═══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet(title="Info & Parameter")
    ws2.sheet_properties.tabColor = 'BF8F00'
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 50

    tel = shared_params['tel_params']
    sumber_label = HilalVisibilityCalculator.SUMBER_ATMOSFER_LABEL.get(
        shared_params['sumber_atmosfer'], shared_params['sumber_atmosfer'].upper())

    info_data = [
        ('Program', 'Perhitungan Visibilitas Hilal — Multi-Lokasi'),
        ('Bulan Hijriah', f"{shared_params['bulan_hijri']}"),
        ('Tahun Hijriah', f"{shared_params['tahun_hijri']}"),
        ('Mode Perhitungan', mode),
        ('Offset Hari', f"H + {shared_params['delta_day']} hari"),
        ('Sumber Data Atmosfer', sumber_label),
        ('Jumlah Lokasi', f"{len(results)}"),
        ('Berhasil', f"{sum(1 for r in results if r['success'])}"),
        ('', ''),
        ('--- Parameter Teleskop ---', ''),
        ('Aperture (mm)', f"{tel.get('aperture', 100.0)}"),
        ('Magnification (x)', f"{tel.get('magnification', 50.0)}"),
        ('Central Obstruction (mm)', f"{tel.get('central_obstruction', 0.0)}"),
        ('Transmission', f"{tel.get('transmission', 0.95)}"),
        ('N Surfaces', f"{tel.get('n_surfaces', 6)}"),
        ('Observer Age', f"{tel.get('observer_age', 22.0)}"),
        ('Seeing (arcsec)', f"{tel.get('seeing', 3.0)}"),
        ('Field Factor Teleskop (F)', f"{tel.get('field_factor', 2.4)}"),
        ('Field Factor Naked Eye (F)', f"{shared_params['F_naked']}"),
        ('', ''),
        ('Tanggal Eksekusi', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
    ]

    for col, hdr in enumerate(['Parameter', 'Nilai'], 1):
        cell = ws2.cell(row=1, column=col, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    for i, (param, nilai) in enumerate(info_data, 2):
        cell_p = ws2.cell(row=i, column=1, value=param)
        cell_p.font = data_font_bold if param.startswith('---') else data_font
        cell_p.border = thin_border
        cell_v = ws2.cell(row=i, column=2, value=nilai)
        cell_v.font = data_font
        cell_v.border = thin_border

    ws2.freeze_panes = 'A2'

    # Save
    output_dir = os.path.dirname(filepath)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    wb.save(filepath)
    print(f"\n  ✓ Hasil multi-lokasi disimpan ke: {filepath}")
    return filepath


def _plot_multi_lokasi(results: list, shared_params: dict,
                       save_path: Optional[str] = None) -> bool:
    """
    Plot perbandingan delta_m (visibility margin) antar lokasi.

    Horizontal bar chart: lokasi di sumbu Y, delta_m di sumbu X.
    Warna hijau jika terlihat (delta_m ≥ 0), merah jika tidak.

    Parameters:
    -----------
    results : list[dict]
        Hasil dari _run_multi_lokasi()
    shared_params : dict
        Parameter bersama
    save_path : str or None
        Path untuk menyimpan gambar. None = hanya tampilkan.

    Returns:
    -------
    bool
        True jika berhasil, False jika gagal.
    """
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        import matplotlib.ticker as ticker
    except ImportError:
        print("  [!] matplotlib belum terinstall. Jalankan: pip install matplotlib")
        return False

    success_results = [r for r in results if r['success']]
    if not success_results:
        print("  [!] Tidak ada hasil valid untuk diplot.")
        return False

    mode = shared_params['mode']
    is_opt = (mode == 'optimal')

    # Kumpulkan data
    nama_list = []
    dm_ne_list = []
    dm_tel_list = []

    for r in reversed(success_results):  # reversed agar urutan atas→bawah sesuai input
        lok = r['lokasi']
        h = r['hasil']
        # Singkat nama jika terlalu panjang
        nama = lok['nama']
        if len(nama) > 30:
            nama = nama[:28] + '…'
        nama_list.append(nama)

        if is_opt:
            dm_ne_list.append(h.get('optimal_delta_m_ne', h.get('delta_m_ne', -99)))
            dm_tel_list.append(h.get('optimal_delta_m_tel', h.get('delta_m_tel', -99)))
        else:
            dm_ne_list.append(h.get('delta_m_ne', -99))
            dm_tel_list.append(h.get('delta_m_tel', -99))

    n = len(nama_list)
    fig_height = max(6, n * 0.5 + 2)
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, fig_height), sharey=True)

    y_pos = range(n)
    bar_height = 0.6

    # --- Panel kiri: Naked Eye ---
    colors_ne = ['#4CAF50' if v >= 0 else '#F44336' for v in dm_ne_list]
    ax1.barh(y_pos, dm_ne_list, height=bar_height, color=colors_ne, edgecolor='white', linewidth=0.5)
    ax1.axvline(x=0, color='#333333', linewidth=1.2, linestyle='-')
    ax1.set_yticks(y_pos)
    ax1.set_yticklabels(nama_list, fontsize=9)
    ax1.set_xlabel('Visibility Margin Δm (mag)', fontsize=11, fontweight='bold')
    ax1.set_title('Naked Eye', fontsize=13, fontweight='bold', pad=10)
    ax1.grid(axis='x', alpha=0.3, linestyle='--')
    # Anotasi nilai
    for j, v in enumerate(dm_ne_list):
        ax1.text(v + (0.3 if v >= 0 else -0.3), j, f'{v:+.2f}',
                 va='center', ha='left' if v >= 0 else 'right', fontsize=8)

    # --- Panel kanan: Teleskop ---
    colors_tel = ['#4CAF50' if v > 0 else '#F44336' for v in dm_tel_list]
    ax2.barh(y_pos, dm_tel_list, height=bar_height, color=colors_tel, edgecolor='white', linewidth=0.5)
    ax2.axvline(x=0, color='#333333', linewidth=1.2, linestyle='-')
    ax2.set_xlabel('Visibility Margin Δm (mag)', fontsize=11, fontweight='bold')
    ax2.set_title('Teleskop', fontsize=13, fontweight='bold', pad=10)
    ax2.grid(axis='x', alpha=0.3, linestyle='--')
    # Anotasi nilai
    for j, v in enumerate(dm_tel_list):
        ax2.text(v + (0.3 if v > 0 else -0.3), j, f'{v:+.2f}',
                 va='center', ha='left' if v > 0 else 'right', fontsize=8)

    # Judul utama
    bln = shared_params['bulan_hijri']
    thn = shared_params['tahun_hijri']
    sumber = shared_params['sumber_atmosfer'].upper()
    mode_label = "Optimal" if is_opt else "Sunset"
    fig.suptitle(
        f'Perbandingan Visibilitas Hilal — Bulan {bln}/{thn} H\n'
        f'Mode: {mode_label}  |  Atmosfer: {sumber}  |  {n} Lokasi',
        fontsize=14, fontweight='bold', y=1.02
    )

    plt.tight_layout()

    if save_path:
        os.makedirs(os.path.dirname(save_path) or '.', exist_ok=True)
        fig.savefig(save_path, dpi=150, bbox_inches='tight')
        print(f"  [✓] Grafik perbandingan disimpan: {save_path}")

    plt.close(fig)
    return True


def main():
    """Fungsi utama dengan mode interaktif untuk pemilihan lokasi dan parameter."""
    from daftar_lokasi import pilih_lokasi_interaktif

    print("\n" + "=" * 70)
    print("  PROGRAM PERHITUNGAN VISIBILITAS HILAL")
    print("  Model: Schaefer (Sky Brightness) + Kastner (Luminansi Hilal)")
    print("=" * 70)

    # LANGKAH 0: Pilih Mode Lokasi (Tunggal / Multi)
    mode_lokasi = _input_mode_lokasi()

    if mode_lokasi == 'multi':
        return _main_multi()
    else:
        return _main_single()


def _main_single():
    """Alur interaktif untuk perhitungan lokasi tunggal (alur asli)."""
    from daftar_lokasi import pilih_lokasi_interaktif

    # LANGKAH 1: Pilih Lokasi
    print("\n--- LANGKAH 1: PILIH LOKASI PENGAMATAN ---")
    lokasi = pilih_lokasi_interaktif()
    if lokasi is None:
        print("[!] Lokasi tidak valid. Program dihentikan.")
        return None

    nama_tempat = lokasi["nama"]
    lintang = lokasi["lat"]
    bujur = lokasi["lon"]
    elevasi = lokasi["elevasi"]
    adm4_code = lokasi.get("adm4_code", "")
    bias_t = lokasi.get("bias_t", 0.0)
    bias_rh = lokasi.get("bias_rh", 0.0)
    timezone_str = tentukan_timezone_indonesia(bujur)

    print(f"\n  Lokasi    : {nama_tempat}")
    print(f"  Koordinat : {lintang}°, {bujur}°")
    print(f"  Elevasi   : {elevasi} m")
    print(f"  Timezone  : {timezone_str}")
    if adm4_code:
        print(f"  Kode BMKG : {adm4_code}")
    if bias_t != 0.0 or bias_rh != 0.0:
        print(f"  Bias Data : T={bias_t:+.1f}°C, RH={bias_rh:+.1f}%")
    else:
        print(f"  Bias Data : Tidak ada data bias (tanpa koreksi)")

    # LANGKAH 2: Input Bulan dan Tahun Hijriah
    result = _input_bulan_tahun_hijri()
    if result is None:
        return None
    bulan_hijri, tahun_hijri = result

    # LANGKAH 3 & 3.5: Mode dan Offset
    mode, delta_day = _input_mode_dan_offset()

    # LANGKAH 3.6: Sumber Data Atmosfer
    sumber_atmosfer, adm4_code, manual_rh, manual_t, manual_p = _input_sumber_atmosfer(adm4_code)

    # LANGKAH 3.7: Koreksi Bias
    bias_t, bias_rh = _input_koreksi_bias(bias_t, bias_rh, sumber_atmosfer)

    # LANGKAH 4: Parameter Teleskop & Field Factor
    tel_params = _input_parameter_teleskop()
    F_naked = tel_params.pop('F_naked')  # pisahkan, bukan parameter teleskop

    # JALANKAN PERHITUNGAN
    print("\n" + "=" * 70)
    print("  MEMULAI PERHITUNGAN...")
    print("=" * 70)

    calculator = HilalVisibilityCalculator(
        nama_tempat=nama_tempat,
        lintang=lintang,
        bujur=bujur,
        elevasi=elevasi,
        timezone_str=timezone_str,
        bulan_hijri=bulan_hijri,
        tahun_hijri=tahun_hijri,
        delta_day_offset=delta_day,
        bias_t=bias_t,
        bias_rh=bias_rh,
        sumber_atmosfer=sumber_atmosfer,
        adm4_code=adm4_code,
        manual_rh=manual_rh,
        manual_t=manual_t,
        manual_p=manual_p
    )

    hasil = calculator.jalankan_perhitungan_lengkap(
        use_telescope=True,
        mode=mode,
        min_moon_alt=2.0,
        F_naked=F_naked,
        **tel_params
    )

    status_ne = "TERLIHAT" if hasil.get('delta_m_ne', -1) >= 0 else "SULIT TERLIHAT"
    status_tel = "TERDETEKSI" if hasil.get('delta_m_tel', -1) > 0 else "SULIT TERDETEKSI"
    print(f"\nPROGRAM SELESAI   : Hilal {status_ne} (naked eye), {status_tel} (teleskop)")

    # LANGKAH 5: Simpan ke Excel
    print("\n--- LANGKAH 5: SIMPAN HASIL KE EXCEL ---")
    try:
        simpan_excel = input("  Simpan hasil ke file Excel? (Y/n): ").strip().lower()
    except EOFError:
        simpan_excel = 'n'

    if simpan_excel != 'n':
        nama_file_safe = nama_tempat.replace(' ', '_').replace('/', '-').replace('\\', '-')
        sumber_tag = sumber_atmosfer.upper().replace('MERRA2', 'MERRA2')
        nama_file = f"Hilal_{nama_file_safe}_{bulan_hijri}_{tahun_hijri}_{sumber_tag}.xlsx"
        script_dir = os.path.dirname(os.path.abspath(__file__))
        output_dir = os.path.join(script_dir, 'output')
        filepath = os.path.join(output_dir, nama_file)
        calculator.simpan_ke_excel(filepath)
        print(f"  File: {filepath}")
    else:
        print("  Hasil tidak disimpan ke Excel.")

    # LANGKAH 6: Grafik Visualisasi Visibility Margin
    if mode.lower() == 'optimal' and hasil.get('all_timestep_results'):
        print("\n--- LANGKAH 6: GRAFIK VISUALISASI ---")
        try:
            simpan_grafik = input("  Simpan grafik visibilitas margin ke file gambar? (Y/n): ").strip().lower()
        except EOFError:
            simpan_grafik = 'n'

        if simpan_grafik != 'n':
            nama_file_safe = nama_tempat.replace(' ', '_').replace('/', '-').replace('\\', '-')
            sumber_tag = sumber_atmosfer.upper()
            nama_grafik = f"Grafik_Hilal_{nama_file_safe}_{bulan_hijri}_{tahun_hijri}_{sumber_tag}.png"
            script_dir = os.path.dirname(os.path.abspath(__file__))
            output_dir = os.path.join(script_dir, 'output')
            grafik_path = os.path.join(output_dir, nama_grafik)
            calculator.plot_visibility_margin(save_path=grafik_path)
        else:
            print("  Grafik tidak disimpan.")

    return hasil


def _main_multi():
    """Alur interaktif untuk perhitungan multi-lokasi (batch)."""

    # LANGKAH 1B: Pilih Lokasi-Lokasi
    lokasi_list = _input_multi_lokasi()
    if not lokasi_list:
        print("[!] Tidak ada lokasi yang dipilih. Program dihentikan.")
        return None

    # LANGKAH 2: Input Bulan dan Tahun Hijriah (shared)
    result = _input_bulan_tahun_hijri()
    if result is None:
        return None
    bulan_hijri, tahun_hijri = result

    # LANGKAH 3 & 3.5: Mode dan Offset (shared)
    mode, delta_day = _input_mode_dan_offset()

    # LANGKAH 3.6: Sumber Data Atmosfer (shared)
    # Untuk multi-lokasi, adm4_code diambil per lokasi, jadi pass empty string
    sumber_atmosfer, _, manual_rh, manual_t, manual_p = _input_sumber_atmosfer('')

    # LANGKAH 3.7: Koreksi Bias
    # Untuk multi-lokasi, gunakan bias bawaan masing-masing lokasi
    if sumber_atmosfer != 'manual':
        print(f"\n--- LANGKAH 3.7: KOREKSI BIAS ---")
        print(f"  ✓ Mode multi-lokasi: menggunakan bias bawaan per lokasi")
        print(f"    (setiap lokasi memiliki bias_t dan bias_rh dari database)")

    # LANGKAH 4: Parameter Teleskop & Field Factor (shared)
    tel_params = _input_parameter_teleskop()
    F_naked = tel_params.pop('F_naked')

    # Susun shared_params
    shared_params = {
        'bulan_hijri': bulan_hijri,
        'tahun_hijri': tahun_hijri,
        'mode': mode,
        'delta_day': delta_day,
        'sumber_atmosfer': sumber_atmosfer,
        'manual_rh': manual_rh,
        'manual_t': manual_t,
        'manual_p': manual_p,
        'F_naked': F_naked,
        'tel_params': tel_params,
    }

    # JALANKAN BATCH
    results = _run_multi_lokasi(lokasi_list, shared_params)

    # LANGKAH 5B: Simpan ke Excel
    print("\n--- LANGKAH 5: SIMPAN HASIL KE EXCEL ---")
    try:
        simpan_excel = input("  Simpan hasil multi-lokasi ke file Excel? (Y/n): ").strip().lower()
    except EOFError:
        simpan_excel = 'n'

    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(script_dir, 'output')

    if simpan_excel != 'n':
        sumber_tag = sumber_atmosfer.upper()
        nama_file = f"Hilal_Multi_{bulan_hijri}_{tahun_hijri}_{sumber_tag}_{len(results)}lokasi.xlsx"
        filepath = os.path.join(output_dir, nama_file)
        _simpan_excel_multi(results, shared_params, filepath)
        print(f"  File: {filepath}")
    else:
        print("  Hasil tidak disimpan ke Excel.")

    # LANGKAH 6B: Grafik Perbandingan Multi-Lokasi
    print("\n--- LANGKAH 6: GRAFIK PERBANDINGAN MULTI-LOKASI ---")
    try:
        simpan_grafik = input("  Simpan grafik perbandingan visibilitas antar lokasi? (Y/n): ").strip().lower()
    except EOFError:
        simpan_grafik = 'n'

    if simpan_grafik != 'n':
        sumber_tag = sumber_atmosfer.upper()
        nama_grafik = f"Grafik_Multi_{bulan_hijri}_{tahun_hijri}_{sumber_tag}_{len(results)}lokasi.png"
        grafik_path = os.path.join(output_dir, nama_grafik)
        _plot_multi_lokasi(results, shared_params, save_path=grafik_path)
    else:
        print("  Grafik tidak disimpan.")

    return results


if __name__ == "__main__":
    hasil = main()

