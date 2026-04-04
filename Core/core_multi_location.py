#!/usr/bin/env python3
"""
══════════════════════════════════════════════════════════════════════
VALIDASI MODEL CRUMEY (2014) vs DATA OBSERVASI HILAL INDONESIA
══════════════════════════════════════════════════════════════════════

Menjalankan model visibilitas hilal pada seluruh data observasi rukyatul
hilal dari lokasi-lokasi BMKG Indonesia, lalu membandingkan hasil prediksi
model (naked eye & teleskop) dengan data observasi aktual.

Fitur:
  1. Batch processing observasi di banyak lokasi sekaligus
  2. Perhitungan visibilitas naked eye dan teleskop
  3. Perbandingan prediksi model vs observasi (Y/N)
  4. Output Excel dengan format rapi

CARA PAKAI:
  Letakkan script ini di direktori Core/ bersama modul-modul lainnya,
  lalu jalankan:
      python batch_validation_crumey.py

PRASYARAT:
  - Koneksi internet (untuk ERA5 API)
  - Semua modul Core/ sudah terinstall dan berfungsi
  - File de440s.bsp ada di direktori data-hisab/

Author: Pipeline validasi untuk skripsi hilal
Referensi: Crumey, A. (2014), MNRAS 442, 2600-2619
══════════════════════════════════════════════════════════════════════
"""

import csv
import os
import sys
import time
import traceback
from typing import List

# Pastikan Core/ di PATH
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from core_crescent_visibility import (
    HilalVisibilityCalculator,
    tentukan_timezone_indonesia,
)


# ═══════════════════════════════════════════════════════════════════
# KONFIGURASI GLOBAL (nilai default)
# ═══════════════════════════════════════════════════════════════════

# --- Field factor referensi ---
F_NAKED_REF = 1.8       # naked-eye reference
FIELD_FACTOR_REF = 1.8   # telescope reference

# --- Parameter teleskop default BMKG ---
TEL_PARAMS = dict(
    aperture=100.0,             # mm (refraktor BMKG tipikal)
    magnification=50.0,         # pembesaran
    transmission=0.95,          # transmisi per permukaan
    n_surfaces=6,               # jumlah permukaan optik
    central_obstruction=0.0,    # refraktor → 0
    observer_age=22.0,          # usia pengamat tipikal
    field_factor=FIELD_FACTOR_REF,
)

# --- Mode perhitungan ---
CALC_MODE = "optimal"       # "sunset" atau "optimal"
SUMBER_ATMOSFER = "era5"    # "era5", "merra2", "manual"

# --- Interval loop optimal ---
INTERVAL_MENIT = 1
MIN_MOON_ALT = 2.0
START_DELAY_MENIT = 1


def _input_float(prompt: str, default: float) -> float:
    """Minta input float dari user, return default jika kosong."""
    try:
        val = input(f"  {prompt} [{default}]: ").strip()
        return float(val) if val else default
    except ValueError:
        print(f"    [!] Input tidak valid, menggunakan default: {default}")
        return default


def _input_int(prompt: str, default: int) -> int:
    """Minta input int dari user, return default jika kosong."""
    try:
        val = input(f"  {prompt} [{default}]: ").strip()
        return int(val) if val else default
    except ValueError:
        print(f"    [!] Input tidak valid, menggunakan default: {default}")
        return default


def _input_konfigurasi_interaktif():
    """Konfigurasi global secara interaktif. Tekan Enter untuk pakai default."""
    global F_NAKED_REF, FIELD_FACTOR_REF, TEL_PARAMS
    global CALC_MODE, SUMBER_ATMOSFER
    global INTERVAL_MENIT, MIN_MOON_ALT, START_DELAY_MENIT

    print("\n" + "═" * 70)
    print("  KONFIGURASI GLOBAL (tekan Enter untuk pakai nilai default)")
    print("═" * 70)

    # --- Mode perhitungan ---
    print("\n  ── Mode Perhitungan ──")
    print(f"  Pilihan: 1=optimal, 2=sunset")
    mode_input = input(f"  Mode perhitungan [{CALC_MODE}]: ").strip().lower()
    if mode_input == "2" or mode_input == "sunset":
        CALC_MODE = "sunset"
    elif mode_input == "1" or mode_input == "optimal":
        CALC_MODE = "optimal"
    elif mode_input:
        print(f"    [!] Input tidak valid, menggunakan default: {CALC_MODE}")
    print(f"    → Mode: {CALC_MODE}")

    # --- Sumber atmosfer ---
    print("\n  ── Sumber Data Atmosfer ──")
    print(f"  Pilihan: 1=era5, 2=merra2, 3=manual")
    atm_input = input(f"  Sumber atmosfer [{SUMBER_ATMOSFER}]: ").strip().lower()
    if atm_input in ("1", "era5"):
        SUMBER_ATMOSFER = "era5"
    elif atm_input in ("2", "merra2"):
        SUMBER_ATMOSFER = "merra2"
    elif atm_input in ("3", "manual"):
        SUMBER_ATMOSFER = "manual"
    elif atm_input:
        print(f"    [!] Input tidak valid, menggunakan default: {SUMBER_ATMOSFER}")
    print(f"    → Sumber: {SUMBER_ATMOSFER}")

    # --- Field factor ---
    print("\n  ── Field Factor ──")
    F_NAKED_REF = _input_float("F naked eye", F_NAKED_REF)
    FIELD_FACTOR_REF = _input_float("F teleskop", FIELD_FACTOR_REF)

    # --- Parameter teleskop ---
    print("\n  ── Parameter Teleskop ──")
    TEL_PARAMS['aperture'] = _input_float("Aperture (mm)", TEL_PARAMS['aperture'])
    TEL_PARAMS['magnification'] = _input_float("Magnifikasi", TEL_PARAMS['magnification'])
    TEL_PARAMS['transmission'] = _input_float("Transmisi/permukaan", TEL_PARAMS['transmission'])
    TEL_PARAMS['n_surfaces'] = _input_int("Jumlah permukaan optik", int(TEL_PARAMS['n_surfaces']))
    TEL_PARAMS['central_obstruction'] = _input_float("Central obstruction", TEL_PARAMS['central_obstruction'])
    TEL_PARAMS['observer_age'] = _input_float("Usia pengamat (tahun)", TEL_PARAMS['observer_age'])
    TEL_PARAMS['field_factor'] = FIELD_FACTOR_REF

    # --- Interval loop optimal ---
    print("\n  ── Interval Loop Optimal ──")
    INTERVAL_MENIT = _input_int("Interval (menit)", INTERVAL_MENIT)
    MIN_MOON_ALT = _input_float("Min altitude bulan (°)", MIN_MOON_ALT)
    START_DELAY_MENIT = _input_int("Start delay (menit)", START_DELAY_MENIT)

    # --- Ringkasan ---
    print("\n" + "─" * 70)
    print("  RINGKASAN KONFIGURASI:")
    print(f"    Mode          : {CALC_MODE}")
    print(f"    Atmosfer      : {SUMBER_ATMOSFER}")
    print(f"    F naked eye   : {F_NAKED_REF}")
    print(f"    F teleskop    : {FIELD_FACTOR_REF}")
    print(f"    Aperture      : {TEL_PARAMS['aperture']} mm")
    print(f"    Magnifikasi   : {TEL_PARAMS['magnification']}")
    print(f"    Transmisi     : {TEL_PARAMS['transmission']}")
    print(f"    N surfaces    : {TEL_PARAMS['n_surfaces']}")
    print(f"    Obstruction   : {TEL_PARAMS['central_obstruction']}")
    print(f"    Usia pengamat : {TEL_PARAMS['observer_age']} tahun")
    print(f"    Interval      : {INTERVAL_MENIT} menit")
    print(f"    Min moon alt  : {MIN_MOON_ALT}°")
    print(f"    Start delay   : {START_DELAY_MENIT} menit")
    print("─" * 70)


# ═══════════════════════════════════════════════════════════════════
# DATA OBSERVASI
# ═══════════════════════════════════════════════════════════════════
# Format per entry:
#   no, tanggal (YYYY-MM-DD), nama lokasi,
#   lat, lon, elv, adm4_code,
#   bulan_hijri, tahun_hijri,
#   bias_t, bias_rh,
#   observed (True=Y, False=N)

OBSERVATIONS = [
    # ── Muharram 1444 (29 Juli 2022) ─────────────────────────────
    ( 1, "2022-07-29", "Rooftop Observatorium UIN WS",
      -6.99167, 110.34806, 89.0, "33.74.15.1010",
       1, 1444,  -1, 2,  False),
    ( 2, "2022-07-29", "Tower Hilal Sulamu_BMKG Kupang",
      -10.14333, 123.62667, 10.42, "53.01.07.2001",
       1, 1444,  -1, -2,  True),
    ( 3, "2022-07-29", "Hotel mina tanjung",
      -8.346986, 116.149, 6.0, "52.08.01.2001",
       1, 1444,  -1, 2,  False),
    ( 4, "2022-07-29", "Pantai Lhoknga_chiek kuta_BMKG Aceh Besar",
       5.46667, 95.24233, 11.65, "11.06.02.2001",
       1, 1444,  -1, 4,  False),
    ( 5, "2022-07-29", "POB Condrodipo",
      -7.16967, 112.61733, 61.0, "35.25.14.2002",
       1, 1444,   0, -4,  False),
    ( 6, "2022-07-29", "Tower Hilal Marana_BMKG Palu",
      -0.57861, 119.79070, 17.49, "72.03.10.2007",
       1, 1444,  -4, 10,  False),
    ( 7, "2022-07-29", "Tower Hilal Cikelet",
      -7.59367, 107.62350, 7.0, "32.05.02.2005",
       1, 1444,  -1, 2,  False),
    ( 8, "2022-07-29", "Lapangan Tembak Desa Kebutuhjurang_BMKG Banjarnegara",
      -7.480125, 109.677931, 496.0, "33.04.20.2005",
       1, 1444,   0, 0,  False),
    ( 9, "2022-07-29", "POB Syekh Bela-Belu",
      -7.73983, 110.35017, 45.0, "34.02.04.2005",
       1, 1444,   0, 0,  False),
    (10, "2022-07-29", "Tower Hilal Ternate",
      -0.79983, 127.29483, 33.61, "82.71.01.1005",
       1, 1444,  -1, 2,  False),

    # ── Ramadhan 1444 (22 Maret 2023) ────────────────────────────
    (11, "2023-03-22", "Rooftop Observatorium UIN WS",
      -6.99167, 110.34806, 89.0, "33.74.15.1010",
       9, 1444,  -1, 2,  False),
    (12, "2023-03-22", "POB Pedalen Kebumen_BMKG Banjarnegara",
      -7.731322, 109.390878, 9.0, "33.05.01.2001",
       9, 1444,   0, 0,  False),
    (13, "2023-03-22", "Pantai Loang Baloq-MATARAM",
      -8.60408, 116.07440, 5.0, "52.71.04.1002",
       9, 1444,  -1, 2,  True),
    (14, "2023-03-22", "Tower Hilal Marana_BMKG Palu",
      -0.57861, 119.79070, 17.49, "72.03.10.2007",
       9, 1444,  -4, 10,  True),
    (15, "2023-03-22", "Pantai Lhoknga_chiek kuta_BMKG Aceh Besar",
       5.46667, 95.24233, 11.65, "11.06.02.2001",
       9, 1444,  -1, 4,  True),
    (16, "2023-03-22", "Tower Hilal Meras_MTC Manado_BMKG Manado",
       1.48033, 124.83367, 15.14, "71.71.06.1005",
       9, 1444,  -1, 4,  False),
    (17, "2023-03-22", "POB Cibeas Pel. Ratu",
      -7.07400, 106.53133, 114.0, "32.02.02.2003",
       9, 1444,   0, -2,  False),
    (18, "2023-03-22", "POB Syekh Bela-Belu",
      -7.73983, 110.35017, 45.0, "34.02.04.2005",
       9, 1444,   0, 0,  False),
    (19, "2023-03-22", "Tower Hilal Cikelet",
      -7.59367, 107.62350, 7.0, "32.05.02.2005",
       9, 1444,  -1, 2,  False),
    (20, "2023-03-22", "Gedung BMKG NTT-Kupang",
      -10.15278, 123.60833, 40.0, "53.71.06.1007",
       9, 1444,  -1, -2,  False),
 
    # ── Syawal 1445 (9 April 2024) ───────────────────────────────
    (21, "2024-04-09", "Rooftop Observatorium UIN WS",
      -6.99167, 110.34806, 89.0, "33.74.15.1010",
      10, 1445,  -1, 2,  False),
    (22, "2024-04-09", "POB Pedalen Kebumen_BMKG Banjarnegara",
      -7.731322, 109.390878, 9.0, "33.05.01.2001",
      10, 1445,   0, 0,  False),
    (23, "2024-04-09", "Tower Hilal Meras_MTC Manado_BMKG Manado",
       1.48033, 124.83367, 15.14, "71.71.06.1005",
      10, 1445,  -1, 4,  True),
    (24, "2024-04-09", "Pantai Loang Baloq-MATARAM",
      -8.60408, 116.07440, 5.0, "52.71.04.1002",
      10, 1445,  -1, 2,  False),
    (25, "2024-04-09", "Pantai Lhoknga_chiek kuta_BMKG Aceh Besar",
       5.46667, 95.24233, 11.65, "11.06.02.2001",
      10, 1445,  -1, 4,  False),
    (26, "2024-04-09", "Gedung BMKG NTT-Kupang",
      -10.15278, 123.60833, 40.0, "53.71.06.1007",
      10, 1445,  -1, -2,  False),
    (27, "2024-04-09", "Tower Hilal Ternate",
      -0.79983, 127.29483, 33.61, "82.71.01.1005",
      10, 1445,  -1, 2,  False),
    (28, "2024-04-09", "Pantai Ngliyep Malang_BMKG Malang",
      -8.35000, 112.43333, 10.0, "35.07.15.2001",
      10, 1445,   0, 0,  False),
    (29, "2024-04-09", "Tower Hilal Marana_BMKG Palu",
      -0.57861, 119.79070, 17.49, "72.03.10.2007",
      10, 1445,  -4, 10,  False),
    (30, "2024-04-09", "kantor stageof lampung utara_BMKG Lampung Utara",
       4.836136, 104.87005, 33.0, "18.03.07.1007",
      10, 1445,  -1, 6,  False),
    (31, "2024-04-09", "POB Syekh Bela-Belu",
      -7.73983, 110.35017, 45.0, "34.02.04.2005",
      10, 1445,   0, 0,  False), 
]


# ═══════════════════════════════════════════════════════════════════
# HELPER
# ═══════════════════════════════════════════════════════════════════

N_OBS = len(OBSERVATIONS)


def parse_obs(entry: tuple) -> dict:
    """Parse satu entry tuple menjadi dictionary."""
    return {
        'no': entry[0],
        'tanggal': entry[1],
        'nama': entry[2],
        'lat': entry[3],
        'lon': entry[4],
        'elv': entry[5],
        'adm4': entry[6],
        'bulan_hijri': entry[7],
        'tahun_hijri': entry[8],
        'bias_t': entry[9],
        'bias_rh': entry[10],
        'observed': entry[11],  # True = Y, False = N
    }


# ═══════════════════════════════════════════════════════════════════
# BATCH PROCESSING
# ═══════════════════════════════════════════════════════════════════

def run_single_observation(obs: dict, verbose: bool = True) -> dict:
    """Jalankan model untuk satu observasi.

    Returns
    -------
    dict : berisi semua hasil + metadata observasi
    """
    no = obs['no']
    nama = obs['nama']
    tanggal = obs['tanggal']
    tz = tentukan_timezone_indonesia(obs['lon'])

    if verbose:
        print(f"\n{'═' * 70}")
        print(f"[{no:2d}/{N_OBS}] {nama}")
        print(f"        Tanggal: {tanggal}  |  Hijri: {obs['bulan_hijri']}/{obs['tahun_hijri']}")
        print(f"        Lat={obs['lat']:.4f}  Lon={obs['lon']:.4f}  Elv={obs['elv']:.0f}m")
        print(f"        Observasi: {'TERLIHAT (Y)' if obs['observed'] else 'TIDAK TERLIHAT (N)'}")
        print(f"{'═' * 70}")

    try:
        calc = HilalVisibilityCalculator(
            nama_tempat=nama,
            lintang=obs['lat'],
            bujur=obs['lon'],
            elevasi=obs['elv'],
            timezone_str=tz,
            bulan_hijri=obs['bulan_hijri'],
            tahun_hijri=obs['tahun_hijri'],
            delta_day_offset=0,
            bias_t=obs['bias_t'],
            bias_rh=obs['bias_rh'],
            sumber_atmosfer=SUMBER_ATMOSFER,
            adm4_code=obs['adm4'],
        )

        hasil = calc.jalankan_perhitungan_lengkap(
            use_telescope=True,
            mode=CALC_MODE,
            F_naked=F_NAKED_REF,
            interval_menit=INTERVAL_MENIT,
            min_moon_alt=MIN_MOON_ALT,
            start_delay_menit=START_DELAY_MENIT,
            **TEL_PARAMS,
        )

        # Tanggal pengamatan yang dihitung model
        tgl_model = hasil.get('tanggal_pengamatan')
        if tgl_model:
            tgl_model_str = tgl_model.strftime('%Y-%m-%d')
        else:
            tgl_model_str = "?"

        # Verifikasi tanggal
        tgl_cocok = (tgl_model_str == tanggal)
        if not tgl_cocok and verbose:
            print(f"\n  ⚠ TANGGAL TIDAK COCOK: model={tgl_model_str}, "
                  f"observasi={tanggal}")

        # Ambil nilai datetime untuk parsing ke time
        sunset_local_dt = hasil.get('sunset_local')
        optimal_time_ne_dt = hasil.get('optimal_time_ne') if CALC_MODE == "optimal" else None
        optimal_time_tel_dt = hasil.get('optimal_time_tel') if CALC_MODE == "optimal" else None

        # Kumpulkan hasil
        result = {
            # Metadata observasi
            'no': no,
            'nama': nama,
            'tanggal_obs': tanggal,
            'tanggal_model': tgl_model_str,
            'tanggal_cocok': tgl_cocok,
            'bulan_hijri': obs['bulan_hijri'],
            'tahun_hijri': obs['tahun_hijri'],
            'lat': obs['lat'],
            'lon': obs['lon'],
            'elv': obs['elv'],
            'observed': obs['observed'],
            'success': True,

            # Hasil saat sunset
            'sunset_local': sunset_local_dt,
            'moon_alt_sunset': hasil.get('moon_alt', 0),
            'sun_alt_sunset': hasil.get('sun_alt', 0),
            'elongation': hasil.get('elongation', 0),
            'moon_width': hasil.get('moon_width', 0),
            'phase_angle': hasil.get('phase_angle', 0),
            'sky_brightness_nl': hasil.get('sky_brightness_nl', 0),
            'luminansi_hilal_nl': hasil.get('luminansi_hilal_nl', 0),
            'k_v': hasil.get('k_v', 0),
            'rh': hasil.get('rh', 0),
            'temperature': hasil.get('temperature', 0),
            'delta_m_ne_sunset': hasil.get('delta_m_ne', -99.0),
            'delta_m_tel_sunset': hasil.get('delta_m_tel', -99.0),
            'telescope_gain_sunset': hasil.get('telescope_gain', 0),
        }

        # Hasil optimal (jika mode optimal)
        if CALC_MODE == "optimal":
            # Ambil data astronomis dari result dict optimal
            opt_ne = hasil.get('optimal_result_ne') or {}
            opt_tel = hasil.get('optimal_result_tel') or {}

            result.update({
                'delta_m_ne_opt': hasil.get('optimal_delta_m_ne', -99.0),
                'delta_m_tel_opt': hasil.get('optimal_delta_m_tel', -99.0),
                'optimal_time_ne': optimal_time_ne_dt,
                'optimal_time_tel': optimal_time_tel_dt,
                'optimal_moon_alt_ne': hasil.get('optimal_moon_alt_ne', 0),
                'optimal_moon_alt_tel': hasil.get('optimal_moon_alt_tel', 0),
                'optimal_sun_alt_tel': hasil.get('optimal_sun_alt_tel', 0),
                'telescope_gain_opt': hasil.get('optimal_telescope_gain', 0),
                'vis_duration_ne': hasil.get('visibility_duration_ne', 0),
                'vis_duration_tel': hasil.get('visibility_duration_tel', 0),
                # Data astronomis optimal NE
                'opt_ne_elongation': opt_ne.get('elongation', 0),
                'opt_ne_sky_brightness_nl': opt_ne.get('sky_brightness_nl', 0),
                'opt_ne_luminansi_hilal_nl': opt_ne.get('luminansi_hilal_nl', 0),
                'opt_ne_k_v': opt_ne.get('k_v', 0),
                'opt_ne_rh': opt_ne.get('rh', 0),
                'opt_ne_temperature': opt_ne.get('temperature', 0),
                # Data astronomis optimal Teleskop
                'opt_tel_elongation': opt_tel.get('elongation', 0),
                'opt_tel_sky_brightness_nl': opt_tel.get('sky_brightness_nl', 0),
                'opt_tel_luminansi_hilal_nl': opt_tel.get('luminansi_hilal_nl', 0),
                'opt_tel_k_v': opt_tel.get('k_v', 0),
                'opt_tel_rh': opt_tel.get('rh', 0),
                'opt_tel_temperature': opt_tel.get('temperature', 0),
            })
        else:
            result.update({
                'delta_m_ne_opt': hasil.get('delta_m_ne', -99.0),
                'delta_m_tel_opt': hasil.get('delta_m_tel', -99.0),
                'optimal_time_ne': None,
                'optimal_time_tel': None,
                'optimal_moon_alt_ne': 0,
                'optimal_moon_alt_tel': 0,
                'optimal_sun_alt_tel': 0,
                'telescope_gain_opt': 0,
                'vis_duration_ne': 0,
                'vis_duration_tel': 0,
                'opt_ne_elongation': 0, 'opt_ne_sky_brightness_nl': 0,
                'opt_ne_luminansi_hilal_nl': 0, 'opt_ne_k_v': 0,
                'opt_ne_rh': 0, 'opt_ne_temperature': 0,
                'opt_tel_elongation': 0, 'opt_tel_sky_brightness_nl': 0,
                'opt_tel_luminansi_hilal_nl': 0, 'opt_tel_k_v': 0,
                'opt_tel_rh': 0, 'opt_tel_temperature': 0,
            })

        # Ringkasan cepat
        dm_ne = result['delta_m_ne_opt']
        dm_tel = result['delta_m_tel_opt']
        pred_ne = "Y" if dm_ne > 0 else "N"
        pred_tel = "Y" if dm_tel > 0 else "N"
        obs_str = "Y" if obs['observed'] else "N"
        match_tel = "✓" if (dm_tel > 0) == obs['observed'] else "✗"

        if verbose:
            print(f"\n  RINGKASAN:")
            print(f"    Naked Eye : Δm={dm_ne:+.3f}  Pred={pred_ne}")
            print(f"    Teleskop  : Δm={dm_tel:+.3f}  Pred={pred_tel}  Obs={obs_str}  {match_tel}")

        return result

    except Exception as e:
        print(f"\n  ✗ ERROR pada obs #{no}: {e}")
        traceback.print_exc()
        return {
            'no': no,
            'nama': nama,
            'tanggal_obs': tanggal,
            'tanggal_model': '?',
            'tanggal_cocok': False,
            'bulan_hijri': obs['bulan_hijri'],
            'tahun_hijri': obs['tahun_hijri'],
            'lat': obs['lat'],
            'lon': obs['lon'],
            'elv': obs['elv'],
            'observed': obs['observed'],
            'success': False,
            'sunset_local': None,
            'optimal_time_ne': None,
            'optimal_time_tel': None,
            'delta_m_ne_sunset': -99.0,
            'delta_m_tel_sunset': -99.0,
            'delta_m_ne_opt': -99.0,
            'delta_m_tel_opt': -99.0,
            'error': str(e),
        }


def run_batch(bias_mode: str = '1', manual_bias_t: float = 0.0, manual_bias_rh: float = 0.0) -> List[dict]:
    """Jalankan model untuk semua observasi."""
    print("\n" + "█" * 70)
    print("  BATCH VALIDATION: Model Crumey (2014) vs Observasi Hilal")
    print(f"  {N_OBS} data observasi × 4 event rukyat (2022–2024)")
    print(f"  Mode: {CALC_MODE}  |  Atmosfer: {SUMBER_ATMOSFER}")
    print(f"  F_naked_ref={F_NAKED_REF}  |  F_tel_ref={FIELD_FACTOR_REF}")
    if bias_mode == '2':
        print("  Koreksi Bias: Tanpa koreksi (bias = 0)")
    elif bias_mode == '3':
        print(f"  Koreksi Bias: Manual Seragam (T={manual_bias_t:+.1f}°C, RH={manual_bias_rh:+.1f}%)")
    else:
        print("  Koreksi Bias: Data Bawaan Lokasi")
    print("█" * 70)

    results = []
    t_start = time.time()

    for seq_no, entry in enumerate(OBSERVATIONS, start=1):
        obs = parse_obs(entry)
        obs['no'] = seq_no
        
        # Terapkan opsi bias
        if bias_mode == '2':
            obs['bias_t'] = 0.0
            obs['bias_rh'] = 0.0
        elif bias_mode == '3':
            obs['bias_t'] = manual_bias_t
            obs['bias_rh'] = manual_bias_rh

        t_obs_start = time.time()
        result = run_single_observation(obs)
        elapsed = time.time() - t_obs_start
        result['elapsed_sec'] = elapsed
        results.append(result)
        print(f"  ⏱ Waktu: {elapsed:.1f} detik")

    total_time = time.time() - t_start
    n_success = sum(1 for r in results if r.get('success', False))
    print(f"\n{'═' * 70}")
    print(f"BATCH SELESAI: {n_success}/{N_OBS} berhasil dalam {total_time:.0f} detik")
    print(f"{'═' * 70}")

    return results


# ═══════════════════════════════════════════════════════════════════
# TABEL RINGKASAN
# ═══════════════════════════════════════════════════════════════════

def print_results_table(results: List[dict]):
    """Cetak tabel ringkasan hasil."""
    print(f"\n{'═' * 120}")
    print(f"TABEL HASIL OBSERVASI (F_naked={F_NAKED_REF:.1f}, F_tel={FIELD_FACTOR_REF:.1f})")
    print(f"{'═' * 120}")
    hdr = (f"{'No':>3} {'Tanggal':>10} {'Lokasi':<35} "
           f"{'Obs':>3} {'Moon°':>6} {'Elong°':>6} {'Lebar':>6} "
           f"{'Δm_NE':>8} {'Pr_NE':>5} "
           f"{'Δm_Tel':>8} {'Pr_Tel':>6} {'Match':>5}")
    print(hdr)
    print("─" * 120)

    for r in results:
        if not r.get('success'):
            print(f"{r['no']:>3} {r['tanggal_obs']:>10} {r['nama']:<35} "
                  f"{'?' :>3} {'?':>6} {'?':>6} {'?':>6} {'ERROR':>8} {'?':>5} "
                  f"{'ERROR':>8} {'?':>6} {'?':>5}")
            continue

        obs_str = "Y" if r['observed'] else "N"
        dm_ne = r.get('delta_m_ne_opt', -99)
        dm_tel = r.get('delta_m_tel_opt', -99)
        pred_ne = "Y" if dm_ne > 0 else "N"
        pred_tel = "Y" if dm_tel > 0 else "N"
        # Cocok jika prediksi teleskop sesuai observasi
        match = "✓" if pred_tel == obs_str else "✗"
        moon_alt = r.get('moon_alt_sunset', 0)
        elong = r.get('elongation', 0)
        moon_width_arcmin = r.get('moon_width', 0) * 60.0

        print(f"{r['no']:>3} {r['tanggal_obs']:>10} {r['nama']:<35} "
              f"{obs_str:>3} {moon_alt:>6.2f} {elong:>6.2f} {moon_width_arcmin:>6.2f} "
              f"{dm_ne:>+8.3f} {pred_ne:>5} "
              f"{dm_tel:>+8.3f} {pred_tel:>6} {match:>5}")

    print("─" * 120)

    # Ringkasan kecocokan (observasi = data teleskop)
    valid = [r for r in results if r.get('success', False)]
    if valid:
        n_match_tel = sum(1 for r in valid if (r.get('delta_m_tel_opt', -99) > 0) == r['observed'])
        print(f"\n  Kecocokan Teleskop vs Observasi : {n_match_tel}/{len(valid)} ({n_match_tel/len(valid):.1%})")


# ═══════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════════

def save_to_excel(results: List[dict], filepath: str, bias_mode_str: str = "Data Bawaan Lokasi"):
    """Simpan semua hasil ke file Excel dengan format rapi 2-baris header."""
    from datetime import time as dt_time
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.styles.colors import Color
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Styles ──
    thin_border = Border(
        left=Side('thin'), right=Side('thin'),
        top=Side('thin'), bottom=Side('thin'))
    thin_border_no_top = Border(
        left=Side('thin'), right=Side('thin'),
        bottom=Side('thin'))

    hdr_font = Font(name='Times New Roman', bold=True, size=11,
                    color=Color(theme=1))
    data_font = Font(name='Times New Roman', size=10)
    center = Alignment(horizontal='center', vertical='center')
    center_no_v = Alignment(horizontal='center')
    left_align = Alignment(horizontal='left', vertical='center')

    # Header fills (warna grup)
    meta_fill = PatternFill('solid', fgColor='0070C0')
    sunset_fill = PatternFill('solid',
                              fgColor=Color(theme=9, tint=-0.249977111117893))
    opt_ne_fill = PatternFill('solid',
                              fgColor=Color(theme=7, tint=-0.249977111117893))
    tel_fill = PatternFill('solid',
                           fgColor=Color(theme=6, tint=-0.249977111117893))

    # Data fills (warna muda untuk kolom Δm & prediksi)
    sunset_data_fill = PatternFill('solid',
                                   fgColor=Color(theme=9, tint=0.3999755851924192))
    opt_ne_data_fill = PatternFill('solid',
                                   fgColor=Color(theme=7, tint=0.3999755851924192))
    tel_data_fill = PatternFill('solid',
                                fgColor=Color(theme=6, tint=0.3999755851924192))

    green_fill = PatternFill('solid', fgColor='C6EFCE')
    red_fill = PatternFill('solid', fgColor='FFC7CE')

    # Ringkasan styles (tetap Arial)
    ringkasan_hdr_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    ringkasan_hdr_fill = PatternFill('solid', fgColor='1F4E79')
    ringkasan_data_font = Font(name='Arial', size=10)
    ringkasan_border = Border(
        left=Side('thin', 'B0B0B0'), right=Side('thin', 'B0B0B0'),
        top=Side('thin', 'B0B0B0'), bottom=Side('thin', 'B0B0B0'))

    # ═══ Sheet 1: Hasil Observasi ═══
    ws = wb.active
    ws.title = "Hasil Observasi"

    # ── Row 1: Header grup (merged) + metadata kolom (merged 2 baris) ──
    meta_headers = ['No', 'Tanggal', 'Lokasi', 'Lat', 'Lon', 'Elv',
                    'Bulan Hijri', 'Time Zone']
    for ci, h in enumerate(meta_headers, 1):
        col_letter = get_column_letter(ci)
        ws.merge_cells(f'{col_letter}1:{col_letter}2')
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font
        c.fill = meta_fill
        c.alignment = center
        c.border = thin_border

    # Grup: Sunset (I1:W1)
    ws.merge_cells('I1:W1')
    c = ws.cell(row=1, column=9,
                value='DATA VISIBILITAS HILAL NAKED EYE DAN TELESKOP SAAT SUNSET')
    c.font = hdr_font; c.fill = sunset_fill
    c.alignment = center_no_v; c.border = thin_border

    # Grup: Optimal NE (X1:AG1)
    ws.merge_cells('X1:AG1')
    c = ws.cell(row=1, column=24,
                value='DATA VISIBILITAS HILAL NAKED EYE WAKTU OPTIMAL ATAU BEST TIME')
    c.font = hdr_font; c.fill = opt_ne_fill
    c.alignment = center_no_v; c.border = thin_border

    # Grup: Teleskop (AH1:AV1)
    ws.merge_cells('AH1:AV1')
    c = ws.cell(row=1, column=34,
                value='DATA VISIBILITAS HILAL BERBANTUAN TELESKOP BEST TIME')
    c.font = hdr_font; c.fill = tel_fill
    c.alignment = center_no_v; c.border = thin_border

    # ── Row 2: Sub-header kolom per grup ──
    # Sunset (I-W)
    sunset_hdrs = {
        9: 'Sunset Lokal', 10: 'Moon Alt (\u00b0)', 11: 'Sun Alt (\u00b0)',
        12: 'Elongasi (\u00b0)',
        13: 'Lebar Sabit (arcmin)', 14: 'Phase Angle (\u00b0)',
        15: 'Sky Bright (nL)', 16: 'Lum Hilal (nL)',
        17: 'k_V', 18: 'RH (%)', 19: 'T (\u00b0C)',
    }
    for ci, h in sunset_hdrs.items():
        c = ws.cell(row=2, column=ci, value=h)
        c.font = hdr_font; c.fill = sunset_fill
        c.alignment = center; c.border = thin_border

    # Δm NE sunset (T2:U2)
    ws.merge_cells('T2:U2')
    c = ws.cell(row=2, column=20, value='\u0394m NE')
    c.font = hdr_font; c.fill = sunset_fill
    c.alignment = center; c.border = thin_border

    # Δm Tel sunset (V2:W2)
    ws.merge_cells('V2:W2')
    c = ws.cell(row=2, column=22, value='\u0394m Tel ')
    c.font = hdr_font; c.fill = sunset_fill
    c.alignment = center; c.border = thin_border

    # Optimal NE (X-AG)
    opt_ne_hdrs = {
        24: 'Best Time NE', 25: 'Moon Alt (\u00b0)', 26: 'Elongasi (\u00b0)',
        27: 'Sky Bright (nL)', 28: 'Lum Hilal (nL)',
        29: 'k_V', 30: 'RH (%)', 31: 'T (\u00b0C)',
    }
    for ci, h in opt_ne_hdrs.items():
        c = ws.cell(row=2, column=ci, value=h)
        c.font = hdr_font; c.fill = opt_ne_fill
        c.alignment = center; c.border = thin_border

    # Δm NE optimal (AF2:AG2)
    ws.merge_cells('AF2:AG2')
    c = ws.cell(row=2, column=32, value='\u0394m NE')
    c.font = hdr_font; c.fill = opt_ne_fill
    c.alignment = center; c.border = thin_border

    # Teleskop (AH-AV)
    tel_hdrs = {
        34: 'Best Time Tel', 35: 'Moon Alt (\u00b0)', 36: 'Sun Alt (\u00b0)',
        37: 'Elongasi (\u00b0)',
        38: 'Sky Bright (nL)', 39: 'Lum Hilal (nL)',
        40: 'k_V', 41: 'RH (%)', 42: 'T (\u00b0C)',
        43: 'Tel Gain Opt', 44: 'Leg time (mnt)',
        47: 'Observasi Tel', 48: 'Correct',
    }
    for ci, h in tel_hdrs.items():
        c = ws.cell(row=2, column=ci, value=h)
        c.font = hdr_font; c.fill = tel_fill
        c.alignment = center; c.border = thin_border

    # Δm Tel teleskop (AS2:AT2)
    ws.merge_cells('AS2:AT2')
    c = ws.cell(row=2, column=45, value='\u0394m Tel')
    c.font = hdr_font; c.fill = tel_fill
    c.alignment = center; c.border = thin_border

    # ── Helper functions ──
    def _parse_time(time_val):
        """Parse waktu (datetime, string, atau datetime.time) ke datetime.time."""
        if not time_val:
            return None
        try:
            # Jika sudah datetime.time, return langsung
            if isinstance(time_val, dt_time):
                return time_val
            # Jika datetime object, ambil bagian time
            if hasattr(time_val, 'hour') and hasattr(time_val, 'minute') and hasattr(time_val, 'second'):
                return dt_time(time_val.hour, time_val.minute, time_val.second)
            # Jika string, parse seperti sebelumnya
            s = str(time_val).strip()
            if ' ' in s:
                time_part = s.split(' ')[1].split('+')[0].split('-')[0]
            else:
                time_part = s
            parts = time_part.split(':')
            return dt_time(int(parts[0]), int(parts[1]),
                           round(float(parts[2])) if len(parts) > 2 else 0)
        except Exception:
            return None

    def _tz_offset(lon):
        """UTC offset zona waktu Indonesia dari bujur."""
        if lon < 115:
            return 7   # WIB
        elif lon < 135:
            return 8   # WITA
        return 9       # WIT

    # ── Data rows (mulai baris 3) ──
    # Kolom yang diberi warna Δm data
    SUNSET_DM_COLS = {20, 21, 22, 23}
    OPT_NE_DM_COLS = {32, 33}
    TEL_DM_COLS = {45, 46, 47}

    for i, r in enumerate(results, 3):
        obs_str = "Y" if r['observed'] else "N"
        dm_ne_sun = r.get('delta_m_ne_sunset', -99)
        dm_tel_sun = r.get('delta_m_tel_sunset', -99)
        dm_ne_opt = r.get('delta_m_ne_opt', -99)
        dm_tel_opt = r.get('delta_m_tel_opt', -99)
        pred_ne_sun = "Y" if dm_ne_sun > 0 else "N"
        pred_tel_sun = "Y" if dm_tel_sun > 0 else "N"
        pred_ne_opt = "Y" if dm_ne_opt > 0 else "N"
        pred_tel_opt = "Y" if dm_tel_opt > 0 else "N"
        cocok = "\u2713" if pred_tel_opt == obs_str else "\u2717"

        row_data = {
            # Metadata (A-H)
            1: r['no'],
            2: r['tanggal_obs'],
            3: r['nama'],
            4: r.get('lat', 0),
            5: r.get('lon', 0),
            6: r.get('elv', 0),
            7: f"{r.get('bulan_hijri', 0)}/{r.get('tahun_hijri', 0)}",
            8: _tz_offset(r.get('lon', 0)),
            # Sunset (I-W)
            9: _parse_time(r.get('sunset_local', '')),
            10: round(r.get('moon_alt_sunset', 0), 4),
            11: round(r.get('sun_alt_sunset', 0), 4),
            12: round(r.get('elongation', 0), 4),
            13: round(r.get('moon_width', 0) * 60.0, 4),
            14: round(r.get('phase_angle', 0), 4),
            15: f"{r.get('sky_brightness_nl', 0):.4e}",
            16: f"{r.get('luminansi_hilal_nl', 0):.4e}",
            17: round(r.get('k_v', 0), 4),
            18: round(r.get('rh', 0), 2),
            19: round(r.get('temperature', 0), 2),
            20: round(dm_ne_sun, 4),
            21: pred_ne_sun,
            22: round(dm_tel_sun, 4),
            23: pred_tel_sun,
            # Optimal NE (X-AG)
            24: _parse_time(r.get('optimal_time_ne', '')),
            25: round(r.get('optimal_moon_alt_ne', 0), 4),
            26: round(r.get('opt_ne_elongation', 0), 4),
            27: f"{r.get('opt_ne_sky_brightness_nl', 0):.4e}",
            28: f"{r.get('opt_ne_luminansi_hilal_nl', 0):.4e}",
            29: round(r.get('opt_ne_k_v', 0), 4),
            30: round(r.get('opt_ne_rh', 0), 2),
            31: round(r.get('opt_ne_temperature', 0), 2),
            32: round(dm_ne_opt, 4),
            33: pred_ne_opt,
            # Teleskop (AH-AV)
            34: _parse_time(r.get('optimal_time_tel', '')),
            35: round(r.get('optimal_moon_alt_tel', 0), 4),
            36: round(r.get('optimal_sun_alt_tel', 0), 4),
            37: round(r.get('opt_tel_elongation', 0), 4),
            38: f"{r.get('opt_tel_sky_brightness_nl', 0):.4e}",
            39: f"{r.get('opt_tel_luminansi_hilal_nl', 0):.4e}",
            40: round(r.get('opt_tel_k_v', 0), 4),
            41: round(r.get('opt_tel_rh', 0), 2),
            42: round(r.get('opt_tel_temperature', 0), 2),
            43: round(r.get('telescope_gain_opt', 0), 4),
            44: int(round(r.get('vis_duration_tel', 0))),
            45: round(dm_tel_opt, 4),
            46: pred_tel_opt,
            47: obs_str,
            48: cocok,
        }

        for ci, val in row_data.items():
            c = ws.cell(row=i, column=ci, value=val)
            c.font = data_font
            c.alignment = left_align if ci == 3 else center
            c.border = thin_border_no_top

            # Format waktu
            if ci in (9, 24, 34) and isinstance(val, dt_time):
                c.number_format = 'h:mm:ss'

            # Warna Δm data
            if ci in SUNSET_DM_COLS:
                c.fill = sunset_data_fill
            elif ci in OPT_NE_DM_COLS:
                c.fill = opt_ne_data_fill
            elif ci in TEL_DM_COLS:
                c.fill = tel_data_fill

            # Kolom Correct
            if ci == 48:
                c.fill = green_fill if cocok == "\u2713" else red_fill

    # ── Column widths ──
    col_widths = {
        'A': 4, 'B': 12, 'C': 30, 'D': 11, 'E': 12, 'F': 7,
        'G': 13, 'H': 7.57,
        'I': 14.29, 'J': 12.71, 'K': 12.71, 'L': 14, 'M': 17,
        'N': 17, 'O': 13, 'P': 16, 'Q': 8, 'R': 13, 'S': 13,
        'T': 9.29, 'U': 4, 'V': 9.29, 'W': 4.43,
        'X': 14.57, 'Y': 12.86, 'Z': 11.86, 'AA': 15.71, 'AB': 15,
        'AC': 8.29, 'AD': 9.14, 'AE': 7.57, 'AF': 9, 'AG': 3.86,
        'AH': 14.71, 'AI': 12.86, 'AJ': 12.71, 'AK': 11.86,
        'AL': 15.71, 'AM': 15, 'AN': 7.29, 'AO': 9, 'AP': 7.43,
        'AQ': 13.29, 'AR': 15.14, 'AS': 8.43, 'AT': 4.29,
        'AU': 14.43, 'AV': 8.29,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = 'A3'

    # ═══ Sheet 2: Ringkasan ═══
    ws2 = wb.create_sheet("Ringkasan")
    ws2.sheet_properties.tabColor = 'BF8F00'
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 25

    valid = [r for r in results if r.get('success', False)]
    n_match_tel = sum(1 for r in valid
                      if (r.get('delta_m_tel_opt', -99) > 0) == r['observed'])

    summary_data = [
        ("KONFIGURASI", ""),
        ("Mode Perhitungan", CALC_MODE),
        ("Sumber Atmosfer", SUMBER_ATMOSFER),
        ("Koreksi Bias", bias_mode_str),
        ("F Naked Eye", F_NAKED_REF),
        ("F Teleskop", FIELD_FACTOR_REF),
        ("Aperture Teleskop (mm)", TEL_PARAMS['aperture']),
        ("Magnifikasi", TEL_PARAMS['magnification']),
        ("Transmisi/permukaan", TEL_PARAMS['transmission']),
        ("Jumlah permukaan", TEL_PARAMS['n_surfaces']),
        ("", ""),
        ("HASIL VALIDASI", ""),
        ("Total Observasi", len(results)),
        ("Observasi Berhasil", len(valid)),
        ("Observasi Terlihat (Y)", sum(1 for r in results if r['observed'])),
        ("Observasi Tidak Terlihat (N)",
         sum(1 for r in results if not r['observed'])),
        ("", ""),
        ("KECOCOKAN PREDIKSI TELESKOP vs OBSERVASI", ""),
        ("Kecocokan Teleskop",
         f"{n_match_tel}/{len(valid)} ({n_match_tel/len(valid):.1%})"
         if valid else "N/A"),
    ]

    for i, (label, val) in enumerate(summary_data, 1):
        cl = ws2.cell(row=i, column=1, value=label)
        cv = ws2.cell(row=i, column=2, value=val)
        if label and not val and val != 0:
            cl.font = ringkasan_hdr_font
            cl.fill = ringkasan_hdr_fill
            cv.fill = ringkasan_hdr_fill
        else:
            cl.font = ringkasan_data_font
            cv.font = ringkasan_data_font
        cl.border = ringkasan_border
        cv.border = ringkasan_border

    # Save
    out_dir = os.path.dirname(filepath)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(filepath)
    print(f"\n  \u2713 Excel disimpan: {filepath}")


# ═══════════════════════════════════════════════════════════════════
# CSV OUTPUT
# ═══════════════════════════════════════════════════════════════════

def save_to_csv(results: List[dict], filepath: str):
    """Simpan hasil ke file CSV dengan format flat yang mudah dibaca program.

    Kolom: No, Tanggal, Lokasi, Lat, Lon, Elv, Bulan_Hijri,
           sun_alt_sunset, Moon_Alt_Sunset, Elongasi_Sunset,
           W_arcmin, Phase_Angle, Sky_Bright_Sunset, Lum_Hilal_Sunset,
           kV_Sunset, RH_Sunset, T_Sunset, Dm_NE_Sunset, Dm_Tel_Sunset,
           Best_Time_Tel, sun_alt_BT, Moon_Alt_BT, Elongasi_BT,
           Sky_Bright_BT, Lum_Hilal_BT, kV_BT, RH_BT, T_BT,
           Tel_Gain, Leg_Time_min, Dm_Tel_BT, Prediksi, Observasi
    """
    headers = [
        'No', 'Tanggal', 'Lokasi', 'Lat', 'Lon', 'Elv', 'Bulan_Hijri',
        'sun_alt_sunset', 'Moon_Alt_Sunset', 'Elongasi_Sunset',
        'W_arcmin', 'Phase_Angle', 'Sky_Bright_Sunset', 'Lum_Hilal_Sunset',
        'kV_Sunset', 'RH_Sunset', 'T_Sunset', 'Dm_NE_Sunset', 'Dm_Tel_Sunset',
        'Best_Time_Tel', 'sun_alt_BT', 'Moon_Alt_BT', 'Elongasi_BT',
        'Sky_Bright_BT', 'Lum_Hilal_BT', 'kV_BT', 'RH_BT', 'T_BT',
        'Tel_Gain', 'Leg_Time_min', 'Dm_Tel_BT', 'Prediksi', 'Observasi',
    ]

    out_dir = os.path.dirname(filepath)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)

    with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(headers)

        for r in results:
            # Parse Best_Time_Tel ke string HH:MM:SS
            bt_val = r.get('optimal_time_tel')
            bt_str = ''
            if bt_val:
                try:
                    if hasattr(bt_val, 'hour'):
                        bt_str = f"{bt_val.hour:02d}:{bt_val.minute:02d}:{bt_val.second:02d}"
                    else:
                        bt_str = str(bt_val)
                except Exception:
                    bt_str = ''

            dm_tel_bt = r.get('delta_m_tel_opt', -99)
            prediksi = "Y" if dm_tel_bt > 0 else "N"
            observasi = "Y" if r.get('observed', False) else "N"

            # Helper: format angka ke string (format standar internasional)
            def _f4(val):
                return f"{val:.4f}" if r.get('success') else ''
            def _f2(val):
                return f"{val:.2f}" if r.get('success') else ''
            def _full(val):
                return repr(val) if r.get('success') else ''

            row = [
                str(r.get('no', '')),
                r.get('tanggal_obs', ''),
                r.get('nama', ''),
                _f4(r.get('lat', 0)),
                _f4(r.get('lon', 0)),
                _f2(r.get('elv', 0)),
                f"{r.get('bulan_hijri', 0)}/{r.get('tahun_hijri', 0)}",
                _f4(r.get('sun_alt_sunset', 0)),
                _f4(r.get('moon_alt_sunset', 0)),
                _f4(r.get('elongation', 0)),
                _f4(r.get('moon_width', 0) * 60.0),
                _f4(r.get('phase_angle', 0)),
                _full(r.get('sky_brightness_nl', 0)),
                _full(r.get('luminansi_hilal_nl', 0)),
                _f4(r.get('k_v', 0)),
                _f2(r.get('rh', 0)),
                _f2(r.get('temperature', 0)),
                _f4(r.get('delta_m_ne_sunset', -99)),
                _f4(r.get('delta_m_tel_sunset', -99)),
                bt_str,
                _f4(r.get('optimal_sun_alt_tel', 0)),
                _f4(r.get('optimal_moon_alt_tel', 0)),
                _f4(r.get('opt_tel_elongation', 0)),
                _full(r.get('opt_tel_sky_brightness_nl', 0)),
                _full(r.get('opt_tel_luminansi_hilal_nl', 0)),
                _f4(r.get('opt_tel_k_v', 0)),
                _f2(r.get('opt_tel_rh', 0)),
                _f2(r.get('opt_tel_temperature', 0)),
                _f4(r.get('telescope_gain_opt', 0)),
                str(int(round(r.get('vis_duration_tel', 0)))) if r.get('success') else '',
                _f4(dm_tel_bt),
                prediksi if r.get('success') else '',
                observasi,
            ]

            writer.writerow(row)

    print(f"  \u2713 CSV disimpan : {filepath}")


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════

def _input_koreksi_bias_batch() -> tuple:
    print("\n--- KOREKSI BIAS ---")
    print("  Pilih opsi koreksi bias untuk semua lokasi:")
    print("  1. Gunakan data bawaan lokasi (jika ada)")
    print("  2. Tanpa koreksi (bias_t = 0, bias_rh = 0)")
    print("  3. Input manual nilai bias seragam untuk semua lokasi")
    try:
        pilihan = input("\n  Pilih opsi (1/2/3) [enter=1]: ").strip() or "1"
        if pilihan == "2":
            print("  ✓ Menggunakan data tanpa koreksi")
            return "2", 0.0, 0.0, "Tanpa koreksi (bias = 0)"
        elif pilihan == "3":
            t_str = input("  Masukkan bias suhu (°C) [enter=0]: ").strip()
            bias_t = float(t_str) if t_str else 0.0
            rh_str = input("  Masukkan bias RH (%) [enter=0]: ").strip()
            bias_rh = float(rh_str) if rh_str else 0.0
            print(f"  ✓ Koreksi bias seragam: T={bias_t:+.1f}°C, RH={bias_rh:+.1f}%")
            return "3", bias_t, bias_rh, f"Manual Seragam (T={bias_t:+.1f}°C, RH={bias_rh:+.1f}%)"
        else:
            print("  ✓ Menggunakan data bawaan lokasi")
            return "1", 0.0, 0.0, "Data Bawaan Lokasi"
    except EOFError:
        return "1", 0.0, 0.0, "Data Bawaan Lokasi"
    except ValueError:
        print("  [!] Input tidak valid, menggunakan bawaan lokasi.")
        return "1", 0.0, 0.0, "Data Bawaan Lokasi"

def main():
    """Entry point utama."""
    # 0. Konfigurasi interaktif
    _input_konfigurasi_interaktif()

    # 1. Prompt Bias
    bias_mode, manual_bias_t, manual_bias_rh, bias_mode_str = _input_koreksi_bias_batch()

    # 2. Batch run
    results = run_batch(bias_mode, manual_bias_t, manual_bias_rh)

    # 3. Tampilkan tabel
    print_results_table(results)

    # 4. Simpan ke Excel
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output_dir = os.path.join(script_dir, 'output')

    bias_tag = "NoBias" if bias_mode == '2' else ("ManualBias" if bias_mode == '3' else "BiasBawaan")
    excel_path = os.path.join(output_dir,
        f"Validasi_Crumey_{SUMBER_ATMOSFER}_{CALC_MODE}_{bias_tag}.xlsx")
    save_to_excel(results, excel_path, bias_mode_str)

    # 5. Simpan ke CSV
    csv_path = os.path.join(output_dir,
        f"Validasi_Crumey_{SUMBER_ATMOSFER}_{CALC_MODE}_{bias_tag}.csv")
    save_to_csv(results, csv_path)

    # 6. Ringkasan akhir
    print(f"\n{'█' * 70}")
    print("  VALIDASI SELESAI")
    print(f"{'█' * 70}")
    print(f"  Excel : {excel_path}")
    print(f"  CSV   : {csv_path}")
    print(f"{'█' * 70}\n")

    return results


if __name__ == "__main__":
    results = main()