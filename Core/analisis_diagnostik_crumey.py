#!/usr/bin/env python3
"""
══════════════════════════════════════════════════════════════════════
ANALISIS DIAGNOSTIK MODEL CRUMEY (2014)
Tahap 4.5: Analisis Sensitivitas dan Ketidakpastian
══════════════════════════════════════════════════════════════════════

Script ini membaca output dari batch_validation_crumey.py, kemudian
melakukan analisis diagnostik LOKAL (tanpa API call) untuk menentukan
penyebab utama error prediksi.

TAHAP YANG DICAKUP (sesuai arsitektur_bab4_final.md):
  4.5.1  Sensitivitas OAT terhadap RH + Kategorisasi FN (A/B/C)
  4.5.2  Dekomposisi jalur error (Kastner vs Schaefer)
  4.5.3  Error bar Δm dari ketidakpastian ERA5
  4.5.4  Perbandingan ERA5 vs MERRA-2

INPUT:
  - data_hilal_clean.csv    (ERA5, dari batch run)
  - data_merra2_clean.csv   (MERRA-2, dari batch run)

OUTPUT: Analisis_Diagnostik_Crumey.xlsx + plot

CARA PAKAI:
  1. Letakkan di direktori Core/
  2. Pastikan file CSV input ada di Core/output/
  3. Jalankan: python analisis_diagnostik_crumey.py
══════════════════════════════════════════════════════════════════════
"""

import math
import os
import sys
import warnings
warnings.filterwarnings('ignore')

# Fix Windows console encoding
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import pandas as pd
import numpy as np
from typing import List, Dict, Tuple, Optional

# Pastikan Core/ di PATH
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from visual_limit_schaefer import hitung_sky_brightness
from visual_limit_kastner import hitung_luminansi_kastner
from full_rumus_crumey import (
    nL_to_cdm2, cdm2_to_nL, arcmin2_to_sr,
    contrast_threshold, crescent_area_arcmin2,
)


# ═══════════════════════════════════════════════════════════════════
# KONFIGURASI
# ═══════════════════════════════════════════════════════════════════

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(SCRIPT_DIR, 'output')

INPUT_ERA5 = os.path.join(OUTPUT_DIR, 'data_hilal_clean.csv')
INPUT_MERRA2 = os.path.join(OUTPUT_DIR, 'data_merra2_clean.csv')

# RH variations (percentage points, added to baseline)
RH_DELTAS = [-30, -25, -20, -15, -10, -5, 0, +5, +10]

# Telescope parameters (harus sama dengan batch run)
TEL_APERTURE = 100.0     # mm
TEL_MAG = 50.0
TEL_TRANS = 0.95
TEL_NSURFACES = 6
TEL_OBSTRUCTION = 0.0
TEL_AGE = 30.0

# Field factors — F = 1.5 (optimal, sudah terkalibrasi; arsitektur §4.2)
F_NAKED = 1.5
F_TEL = 1.5

# Moon semidiameter (default, ~0.26°)
MOON_SD_DEG = 0.26


# ═══════════════════════════════════════════════════════════════════
# DATA LOADING
# ═══════════════════════════════════════════════════════════════════

def load_observation_data(filepath: str) -> pd.DataFrame:
    """Baca CSV ERA5 dan kembalikan DataFrame dengan kolom standar.

    CSV output dari batch run (data_hilal_clean.csv) memakai kolom flat:
      No, Tanggal, Event, Lokasi, Lat, Lon, Elv, Bulan_Hijri,
      Moon_Alt_Sunset, Elongasi_Sunset, W_arcmin, Phase_Angle,
      Sky_Bright_Sunset, ..., Best_Time_Tel, Moon_Alt_BT, Elongasi_BT,
      Sky_Bright_BT, Lum_Hilal_BT, kV_BT, RH_BT, T_BT, ...,
      Dm_Tel_BT, Prediksi, Observasi
    """
    df_raw = pd.read_csv(filepath)

    df = pd.DataFrame()
    df['No'] = df_raw['No']
    df['Tanggal'] = df_raw['Tanggal']
    df['Lokasi'] = df_raw['Lokasi']
    df['Lat'] = df_raw['Lat']
    df['Lon'] = df_raw['Lon']
    df['Elv'] = df_raw['Elv']

    df['Phase Angle (°)'] = df_raw['Phase_Angle']
    df['Lebar Sabit (arcmin)'] = df_raw['W_arcmin']

    # Best Time section (dipakai untuk analisis)
    df['Waktu Optimal Tel'] = df_raw['Best_Time_Tel']
    df['Moon Alt (°)'] = df_raw['Moon_Alt_BT']
    df['Elongasi (°)'] = df_raw['Elongasi_BT']
    df['Sky Bright (nL)'] = df_raw['Sky_Bright_BT']
    df['Lum Hilal (nL)'] = df_raw['Lum_Hilal_BT']
    df['k_V'] = df_raw['kV_BT']
    df['RH (%)'] = df_raw['RH_BT']
    df['T (°C)'] = df_raw['T_BT']
    df['Leg_Time_min'] = df_raw['Leg_Time_min']
    df['Δm Tel Opt'] = df_raw['Dm_Tel_BT']
    df['Obs (Y/N)'] = df_raw['Observasi']
    df['Cocok?'] = (df_raw['Prediksi'] == df_raw['Observasi']).apply(
        lambda x: '✓' if x else '✗')

    return df


def load_merra2_data(filepath: str) -> pd.DataFrame:
    """Baca CSV MERRA-2 dan kembalikan DataFrame dengan kolom standar.

    CSV MERRA-2 (data_merra2_clean.csv) memakai kolom:
      No, Tanggal, Event, Lokasi, Moon_Alt_BT_M2, Elongasi_BT_M2,
      Sky_Bright_BT_M2, Lum_Hilal_BT_M2, kV_BT_M2, RH_BT_M2,
      T_BT_M2, Dm_Tel_BT_M2, Pred_M2, Observasi
    """
    df_raw = pd.read_csv(filepath)

    df = pd.DataFrame()
    df['No'] = df_raw['No']
    df['Tanggal'] = df_raw['Tanggal']
    df['Lokasi'] = df_raw['Lokasi']

    df['Moon Alt (°)'] = df_raw['Moon_Alt_BT_M2']
    df['Elongasi (°)'] = df_raw['Elongasi_BT_M2']
    df['Sky Bright (nL)'] = df_raw['Sky_Bright_BT_M2']
    df['Lum Hilal (nL)'] = df_raw['Lum_Hilal_BT_M2']
    df['k_V'] = df_raw['kV_BT_M2']
    df['RH (%)'] = df_raw['RH_BT_M2']
    df['T (°C)'] = df_raw['T_BT_M2']
    df['Δm Tel Opt'] = df_raw['Dm_Tel_BT_M2']
    df['Obs (Y/N)'] = df_raw['Observasi']
    df['Cocok?'] = (df_raw['Pred_M2'] == df_raw['Observasi']).apply(
        lambda x: '✓' if x else '✗')

    return df


# ═══════════════════════════════════════════════════════════════════
# FUNGSI UTILITAS
# ═══════════════════════════════════════════════════════════════════

def estimate_sun_alt_at_optimal(leg_time_min: float) -> float:
    """Estimasi sun altitude di waktu optimal.
    Matahari turun ~0.25°/menit setelah sunset di tropis.
    leg_time_min: menit setelah sunset (dari kolom Leg_Time_min di CSV).
    """
    delta_min = max(leg_time_min if isinstance(leg_time_min, (int, float)) else 7.0, 0)
    # Sun descent rate: ~0.25°/menit di tropis saat sunset
    return -0.833 - 0.25 * delta_min


def estimate_azimuth_diff(elongation: float, sun_alt: float, moon_alt: float) -> float:
    """Hitung selisih azimuth Sun-Moon dari elongasi via spherical trig."""
    rd = math.pi / 180.0
    cos_elong = math.cos(elongation * rd)
    sin_sun = math.sin(sun_alt * rd)
    sin_moon = math.sin(moon_alt * rd)
    cos_sun = math.cos(sun_alt * rd)
    cos_moon = math.cos(moon_alt * rd)

    denom = cos_sun * cos_moon
    if abs(denom) < 1e-10:
        return elongation  # fallback

    cos_daz = (cos_elong - sin_sun * sin_moon) / denom
    cos_daz = max(-1.0, min(1.0, cos_daz))
    return math.degrees(math.acos(cos_daz))


def compute_telescope_Ba(B_sky_nL: float) -> float:
    """Hitung apparent background melalui teleskop (Crumey Eq. 66)."""
    d_exit = TEL_APERTURE / TEL_MAG  # exit pupil [mm]
    De = 7.0 * math.exp(-0.5 * (TEL_AGE / 100.0) ** 2)  # pupil mata [mm]
    Ft = 1.0 / (TEL_TRANS ** TEL_NSURFACES * (1.0 - (TEL_OBSTRUCTION / TEL_APERTURE) ** 2
                 if TEL_OBSTRUCTION > 0 else 1.0))

    delta_min = min(d_exit, De)
    Ba_factor = (delta_min / De) ** 2 / Ft
    return B_sky_nL * Ba_factor


def compute_delta_m(L_nL: float, B_nL: float, phase_angle: float,
                    mode: str = 'naked_eye') -> Dict[str, float]:
    """Hitung Δm (naked eye atau teleskop) dari luminansi dan sky brightness."""
    B_cd = nL_to_cdm2(B_nL)
    L_cd = nL_to_cdm2(L_nL)

    if B_cd <= 0 or L_cd <= 0:
        return {'delta_m': -99.0, 'C_obj': 0, 'C_th': 0, 'k_ratio': 0}

    C_obj = (L_cd - B_cd) / B_cd
    A_arcmin2 = crescent_area_arcmin2(phase_angle, MOON_SD_DEG)
    A_sr = arcmin2_to_sr(A_arcmin2)
    if A_sr <= 0:
        A_sr = 1e-12

    if mode == 'naked_eye':
        C_th = contrast_threshold(A_sr, B_cd, F=F_NAKED, mode='auto')
    else:  # telescope
        B_tel_nL = compute_telescope_Ba(B_nL)
        B_tel_cd = nL_to_cdm2(B_tel_nL)
        if B_tel_cd <= 0:
            B_tel_cd = 1e-10
        A_eff = A_sr * (TEL_MAG ** 2)
        phi = math.sqrt(2) * 1.0 * F_TEL
        C_th = contrast_threshold(A_eff, B_tel_cd, F=phi, mode='auto')

    if C_obj > 0 and C_th > 0:
        delta_m = 2.5 * math.log10(C_obj / C_th)
    elif C_th > 0:
        L_over_B = max(1.0 + C_obj, 1e-15)
        delta_m = 2.5 * (math.log10(L_over_B) - math.log10(C_th))
    else:
        delta_m = -99.0

    return {'delta_m': delta_m, 'C_obj': C_obj, 'C_th': C_th}


def compute_full_chain(row: pd.Series, rh_override: float) -> Dict[str, float]:
    """Jalankan rantai fisika lengkap untuk satu observasi dengan RH tertentu.

    Returns dict dengan: k_V, B_sky_nL, L_hilal_nL, dm_ne, dm_tel, dll.
    """
    # Parameter astronomi (tetap)
    moon_alt = row['Moon Alt (°)']
    phase_angle = row['Phase Angle (°)']
    elongation = row['Elongasi (°)']
    temperature = row['T (°C)']
    lat = row['Lat']
    elv = row['Elv']

    # Parse tanggal
    tgl = pd.to_datetime(row['Tanggal'])
    month = tgl.month
    year = tgl.year

    # Estimasi sun_alt dan azimuth_diff
    sun_alt = estimate_sun_alt_at_optimal(
        row.get('Leg_Time_min', 7.0)
    )
    azisun = estimate_azimuth_diff(elongation, sun_alt, moon_alt)

    # Clamp RH
    rh_used = max(5.0, min(100.0, rh_override))

    # ── RANTAI 1: Sky Brightness (Schaefer) ──
    try:
        result_sky = hitung_sky_brightness(
            month=month, year=year,
            altsun=sun_alt, azisun=azisun,
            humidity=rh_used, temperature=temperature,
            latitude=lat, elevation=elv,
            alt_objek=max(moon_alt, 0.01),
        )
        B_sky_nL = float(result_sky.get('sky_brightness', 0))
        K_list = result_sky.get('K', [])
        k_v = float(K_list[2]) if len(K_list) > 2 else 0.3
    except Exception:
        B_sky_nL = 1e9
        k_v = 0.5

    if B_sky_nL <= 0:
        B_sky_nL = 1.0

    # ── RANTAI 2: Luminansi Hilal (Kastner) ──
    z = 90.0 - moon_alt
    try:
        L_hilal_nL = hitung_luminansi_kastner(
            alpha=phase_angle, r=MOON_SD_DEG, z=z, k=k_v
        )
    except Exception:
        L_hilal_nL = 0.0

    # ── RANTAI 3: Visibilitas (Crumey) ──
    result_ne = compute_delta_m(L_hilal_nL, B_sky_nL, phase_angle, 'naked_eye')
    result_tel = compute_delta_m(L_hilal_nL, B_sky_nL, phase_angle, 'telescope')

    return {
        'rh_used': rh_used,
        'k_v': k_v,
        'B_sky_nL': B_sky_nL,
        'L_hilal_nL': L_hilal_nL,
        'C_obj': result_ne['C_obj'],
        'C_th_ne': result_ne['C_th'],
        'C_th_tel': result_tel['C_th'],
        'dm_ne': result_ne['delta_m'],
        'dm_tel': result_tel['delta_m'],
        'sun_alt_est': sun_alt,
    }


def classify_obs_type(obs_yn: str, cocok: str) -> str:
    """Klasifikasi tipe observasi: TP, TN, FN, FP."""
    if obs_yn == 'Y' and cocok == '✓':
        return 'TP'
    elif obs_yn == 'Y' and cocok != '✓':
        return 'FN'
    elif obs_yn == 'N' and cocok == '✓':
        return 'TN'
    else:
        return 'FP'


def categorize_fn(butuh_delta_rh) -> str:
    """Kategorisasi FN sesuai arsitektur 4.5.1:
      A (near-miss)  : |ΔRH| ≤ 5 pp
      B (correctable): 5 < |ΔRH| ≤ 20 pp — dalam range bias ERA5
      C (structural) : |ΔRH| > 20 pp atau tidak ada RH kritis
    """
    if not isinstance(butuh_delta_rh, (int, float)):
        return 'C'
    abs_drh = abs(butuh_delta_rh)
    if abs_drh <= 5:
        return 'A'
    elif abs_drh <= 20:
        return 'B'
    else:
        return 'C'


# ═══════════════════════════════════════════════════════════════════
# 4.5.1 — SENSITIVITAS OAT TERHADAP RH
# ═══════════════════════════════════════════════════════════════════

def analyze_rh_sensitivity(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Variasikan RH dan ukur dampak pada Δm untuk semua observasi.

    Sesuai arsitektur 4.5.1:
    - Variasi ΔRH ∈ {-30,...,+10} pp
    - Plot Δm vs ΔRH per observasi (terutama FN dan FP)
    - Hitung RH kritis per FN (Δm = 0)
    - Kategorisasi FN: A (near-miss), B (correctable), C (structural)
    """
    print("\n" + "═" * 70)
    print("4.5.1  SENSITIVITAS OAT Δm TERHADAP VARIASI RH")
    print("═" * 70)

    all_rows = []
    critical_rows = []

    for idx, row in df.iterrows():
        obs_no = int(row['No'])
        rh_baseline = row['RH (%)']
        obs_yn = row['Obs (Y/N)']
        tipe = classify_obs_type(obs_yn, row['Cocok?'])

        print(f"\n  Obs #{obs_no:2d} {row['Lokasi'][:30]:30s} "
              f"({tipe}) RH_base={rh_baseline:.1f}%")

        dm_at_deltas = {}
        kv_at_deltas = {}

        for delta_rh in RH_DELTAS:
            rh_test = rh_baseline + delta_rh
            if rh_test < 5 or rh_test > 100:
                continue

            result = compute_full_chain(row, rh_test)

            dm_at_deltas[delta_rh] = result['dm_tel']
            kv_at_deltas[delta_rh] = result['k_v']

            entry = {
                'No': obs_no,
                'Lokasi': row['Lokasi'],
                'Obs': obs_yn,
                'Tipe': tipe,
                'RH_baseline': rh_baseline,
                'ΔRH': delta_rh,
                'RH_test': rh_test,
                'k_V': result['k_v'],
                'B_sky_nL': result['B_sky_nL'],
                'L_hilal_nL': result['L_hilal_nL'],
                'Δm_NE': result['dm_ne'],
                'Δm_Tel': result['dm_tel'],
                'C_obj': result['C_obj'],
            }
            all_rows.append(entry)

        # Cari critical RH (Δm_tel = 0) via interpolasi linear
        deltas_sorted = sorted(dm_at_deltas.keys())
        critical_rh = None
        for i in range(len(deltas_sorted) - 1):
            d1, d2 = deltas_sorted[i], deltas_sorted[i+1]
            dm1, dm2 = dm_at_deltas[d1], dm_at_deltas[d2]
            if dm1 <= -90 or dm2 <= -90:
                continue
            if (dm1 <= 0 <= dm2) or (dm2 <= 0 <= dm1):
                frac = -dm1 / (dm2 - dm1) if (dm2 - dm1) != 0 else 0
                critical_delta = d1 + frac * (d2 - d1)
                critical_rh = rh_baseline + critical_delta
                break

        # Δm improvement from -20% RH
        dm_base = dm_at_deltas.get(0, -99)
        dm_minus20 = dm_at_deltas.get(-20, dm_at_deltas.get(-25, -99))
        improvement = dm_minus20 - dm_base if dm_base > -90 and dm_minus20 > -90 else 0

        butuh_drh = (critical_rh - rh_baseline) if critical_rh else 'N/A'
        kategori = categorize_fn(butuh_drh) if tipe == 'FN' else '-'

        critical_rows.append({
            'No': obs_no,
            'Lokasi': row['Lokasi'],
            'Obs': obs_yn,
            'Tipe': tipe,
            'RH_baseline': rh_baseline,
            'k_V_baseline': kv_at_deltas.get(0, 0),
            'Δm_Tel_baseline': dm_base,
            'Δm_Tel_RH-20': dm_minus20,
            'Improvement_20': improvement,
            'RH_kritis': critical_rh if critical_rh else 'N/A',
            'Butuh_ΔRH': butuh_drh,
            'Kategori_FN': kategori,
        })

        # Print ringkasan
        status = f"RH_kritis={critical_rh:.1f}%" if critical_rh else "di luar range"
        kat_str = f"  [{kategori}]" if tipe == 'FN' else ""
        print(f"    Δm(base)={dm_base:+.2f}  Δm(RH-20)={dm_minus20:+.2f}  "
              f"gain={improvement:+.2f}  {status}{kat_str}")

    critical_df = pd.DataFrame(critical_rows)

    # ── Ringkasan Kategorisasi FN ──
    fn_rows = critical_df[critical_df['Tipe'] == 'FN']
    if len(fn_rows) > 0:
        print("\n  ┌─────────────────────────────────────────────────┐")
        print("  │         KATEGORISASI FALSE NEGATIVE (FN)        │")
        print("  ├─────────────────────────────────────────────────┤")
        for cat, desc in [('A', 'near-miss, |ΔRH| ≤ 5 pp'),
                          ('B', 'correctable, 5 < |ΔRH| ≤ 20 pp'),
                          ('C', 'structural, |ΔRH| > 20 pp')]:
            n = sum(fn_rows['Kategori_FN'] == cat)
            print(f"  │  Kategori {cat} ({desc}): {n}/{len(fn_rows)}")
        print("  └─────────────────────────────────────────────────┘")

        # Detail per FN
        print("\n  Detail FN:")
        for _, r in fn_rows.iterrows():
            drh_str = f"{r['Butuh_ΔRH']:+.1f} pp" if isinstance(r['Butuh_ΔRH'], (int, float)) else "N/A"
            print(f"    #{int(r['No']):2d} {r['Lokasi'][:25]:25s} "
                  f"ΔRH={drh_str:>10s}  → Kategori {r['Kategori_FN']}")

    return pd.DataFrame(all_rows), critical_df


# ═══════════════════════════════════════════════════════════════════
# 4.5.2 — DEKOMPOSISI JALUR ERROR
# ═══════════════════════════════════════════════════════════════════

def analyze_error_decomposition(df: pd.DataFrame) -> pd.DataFrame:
    """Dekomposisi: kontribusi jalur Kastner (luminansi) vs Schaefer (sky brightness).

    Sesuai arsitektur 4.5.2:
    - Dua jalur pengaruh kV:
      Kastner: kV → ekstingsi → luminansi hilal
      Schaefer: kV → sky brightness
    - Teknik OAT: ubah kV hanya di satu jalur
    - Hitung persentase kontribusi masing-masing
    """
    print("\n" + "═" * 70)
    print("4.5.2  DEKOMPOSISI JALUR ERROR (KASTNER vs SCHAEFER)")
    print("═" * 70)

    fn_fp_obs = df[
        ((df['Obs (Y/N)'] == 'Y') & (df['Cocok?'] != '✓')) |
        ((df['Obs (Y/N)'] == 'N') & (df['Cocok?'] != '✓'))
    ]
    if len(fn_fp_obs) == 0:
        print("  Tidak ada observasi FN/FP untuk didekomposisi.")
        return pd.DataFrame()

    decomp_rows = []

    for _, row in fn_fp_obs.iterrows():
        obs_no = int(row['No'])
        rh_base = row['RH (%)']
        tipe = classify_obs_type(row['Obs (Y/N)'], row['Cocok?'])

        print(f"\n  Obs #{obs_no}: {row['Lokasi']} ({tipe})")

        # Baseline
        res_base = compute_full_chain(row, rh_base)

        # Skenario 1: Kurangi RH 20pp → dampak penuh
        rh_low = max(rh_base - 20, 5.0)
        res_low = compute_full_chain(row, rh_low)

        # Skenario 2: Gunakan k_V rendah tapi B_sky dari baseline RH
        # (isolasi efek k_V pada luminansi saja — jalur Kastner)
        z = 90.0 - row['Moon Alt (°)']
        L_with_low_kv = hitung_luminansi_kastner(
            alpha=row['Phase Angle (°)'], r=MOON_SD_DEG, z=z, k=res_low['k_v']
        )
        res_only_L = compute_delta_m(
            L_with_low_kv, res_base['B_sky_nL'], row['Phase Angle (°)'], 'telescope'
        )

        # Skenario 3: Gunakan B_sky rendah tapi L_hilal dari baseline k_V
        # (isolasi efek k_V pada sky brightness — jalur Schaefer)
        res_only_B = compute_delta_m(
            res_base['L_hilal_nL'], res_low['B_sky_nL'], row['Phase Angle (°)'], 'telescope'
        )

        # Hitung kontribusi
        dm_full_change = res_low['dm_tel'] - res_base['dm_tel']
        dm_L_only = res_only_L['delta_m'] - res_base['dm_tel']
        dm_B_only = res_only_B['delta_m'] - res_base['dm_tel']
        dm_interaction = dm_full_change - dm_L_only - dm_B_only

        decomp_rows.append({
            'No': obs_no,
            'Lokasi': row['Lokasi'],
            'Tipe': tipe,
            'RH_base': rh_base,
            'RH_low': rh_low,
            'k_V_base': res_base['k_v'],
            'k_V_low': res_low['k_v'],
            'Δk_V': res_low['k_v'] - res_base['k_v'],
            'B_sky_base': res_base['B_sky_nL'],
            'B_sky_low': res_low['B_sky_nL'],
            'L_hilal_base': res_base['L_hilal_nL'],
            'L_hilal_low': res_low['L_hilal_nL'],
            'L_ratio': res_low['L_hilal_nL'] / res_base['L_hilal_nL']
                       if res_base['L_hilal_nL'] > 0 else 0,
            'Δm_base': res_base['dm_tel'],
            'Δm_full_change': dm_full_change,
            'Δm_dari_L (Kastner)': dm_L_only,
            'Δm_dari_B (Schaefer)': dm_B_only,
            'Δm_interaksi': dm_interaction,
            'Kontribusi_L (%)': abs(dm_L_only) / abs(dm_full_change) * 100
                                if abs(dm_full_change) > 0.001 else 0,
            'Kontribusi_B (%)': abs(dm_B_only) / abs(dm_full_change) * 100
                                if abs(dm_full_change) > 0.001 else 0,
        })

        print(f"    k_V: {res_base['k_v']:.4f} → {res_low['k_v']:.4f} (Δ={res_low['k_v']-res_base['k_v']:+.4f})")
        print(f"    Δm_tel: {res_base['dm_tel']:+.3f} → {res_low['dm_tel']:+.3f} (gain={dm_full_change:+.3f})")
        print(f"    Kontribusi: Kastner(L)={dm_L_only:+.3f} | Schaefer(B)={dm_B_only:+.3f} | interaksi={dm_interaction:+.3f}")

    decomp_df = pd.DataFrame(decomp_rows)

    # Ringkasan
    if len(decomp_df) > 0:
        mean_L = decomp_df['Kontribusi_L (%)'].mean()
        mean_B = decomp_df['Kontribusi_B (%)'].mean()
        print(f"\n  Rata-rata kontribusi: Kastner(L)={mean_L:.1f}%  Schaefer(B)={mean_B:.1f}%")
        dominant = "Kastner (luminansi)" if mean_L > mean_B else "Schaefer (sky brightness)"
        print(f"  → Jalur dominan: {dominant}")

    return decomp_df


# ═══════════════════════════════════════════════════════════════════
# 4.5.3 — ERROR BAR DARI KETIDAKPASTIAN ERA5
# ═══════════════════════════════════════════════════════════════════

def analyze_error_bars(sensitivity_df: pd.DataFrame, critical_df: pd.DataFrame) -> pd.DataFrame:
    """Estimasi error bar Δm berdasarkan ketidakpastian RH ERA5.

    Sesuai arsitektur 4.5.3:
    - σ_RH dari literatur validasi ERA5 tropis: ±5, ±10, ±15 pp
    - Per observasi: Δm_low, Δm_high → error bar
    - Per FN: apakah interval mencakup Δm = 0?
    - Fraksi FN konsisten per level σ_RH
    """
    print("\n" + "═" * 70)
    print("4.5.3  ERROR BAR Δm DARI KETIDAKPASTIAN ERA5")
    print("═" * 70)

    RH_UNCERTAINTY = [5, 10, 15]  # pp

    rows = []
    for _, cr in critical_df.iterrows():
        obs_no = cr['No']
        dm_base = cr['Δm_Tel_baseline']
        tipe = cr['Tipe']

        obs_sens = sensitivity_df[sensitivity_df['No'] == obs_no]

        for unc in RH_UNCERTAINTY:
            # Δm at RH ± uncertainty
            dm_low = obs_sens[obs_sens['ΔRH'] == -unc]
            dm_high = obs_sens[obs_sens['ΔRH'] == unc]

            dm_low_val = dm_low['Δm_Tel'].values[0] if len(dm_low) > 0 else dm_base
            dm_high_val = dm_high['Δm_Tel'].values[0] if len(dm_high) > 0 else dm_base

            error_bar = (dm_low_val - dm_high_val) / 2.0
            covers_zero = (min(dm_low_val, dm_high_val) <= 0 <= max(dm_low_val, dm_high_val))

            rows.append({
                'No': obs_no,
                'Lokasi': cr['Lokasi'],
                'Obs': cr['Obs'],
                'Tipe': tipe,
                'Δm_baseline': dm_base,
                'σ_RH (±pp)': unc,
                'Δm_low_RH': dm_low_val,
                'Δm_high_RH': dm_high_val,
                'Error_bar (±mag)': abs(error_bar),
                'Mencakup_Δm=0': 'Ya' if covers_zero else 'Tidak',
            })

    result = pd.DataFrame(rows)

    # Ringkasan per level σ_RH
    print("\n  Fraksi FN yang error bar mencakup Δm=0:")
    for unc in RH_UNCERTAINTY:
        sub = result[result['σ_RH (±pp)'] == unc]
        fn_sub = sub[sub['Tipe'] == 'FN']
        n_covers = sum(fn_sub['Mencakup_Δm=0'] == 'Ya')
        if len(fn_sub) > 0:
            frac = n_covers / len(fn_sub) * 100
            print(f"    ±{unc:2d} pp:  {n_covers}/{len(fn_sub)} FN ({frac:.0f}%) konsisten")
        else:
            print(f"    ±{unc:2d} pp:  tidak ada FN")

    # Argumen kunci
    fn_15 = result[(result['σ_RH (±pp)'] == 15) & (result['Tipe'] == 'FN')]
    if len(fn_15) > 0:
        n_covers_15 = sum(fn_15['Mencakup_Δm=0'] == 'Ya')
        if n_covers_15 / len(fn_15) > 0.5:
            print(f"\n  → Mayoritas FN ({n_covers_15}/{len(fn_15)}) konsisten dalam ±15 pp:")
            print("    masalah utama kemungkinan kualitas data input, bukan model fisika")

    return result


# ═══════════════════════════════════════════════════════════════════
# 4.5.4 — PERBANDINGAN ERA5 vs MERRA-2
# ═══════════════════════════════════════════════════════════════════

def analyze_era5_vs_merra2(df_era5: pd.DataFrame, df_merra2: pd.DataFrame) -> pd.DataFrame:
    """Bandingkan output model menggunakan ERA5 vs MERRA-2.

    Sesuai arsitektur 4.5.4:
    - Scatter plot: Δm(ERA5) vs Δm(MERRA-2)
    - Korelasi Pearson/Spearman
    - Selisih RH dan kV antara kedua sumber
    - Confusion matrix MERRA-2 pada F optimal yang sama
    """
    print("\n" + "═" * 70)
    print("4.5.4  PERBANDINGAN ERA5 vs MERRA-2")
    print("═" * 70)

    # Merge on No
    merged = pd.merge(
        df_era5[['No', 'Lokasi', 'Obs (Y/N)', 'Cocok?',
                 'RH (%)', 'k_V', 'T (°C)', 'Δm Tel Opt']],
        df_merra2[['No', 'RH (%)', 'k_V', 'T (°C)', 'Δm Tel Opt', 'Cocok?']],
        on='No', suffixes=('_ERA5', '_MERRA2')
    )

    # Hitung selisih
    merged['ΔRH (M2-E5)'] = merged['RH (%)_MERRA2'] - merged['RH (%)_ERA5']
    merged['Δk_V (M2-E5)'] = merged['k_V_MERRA2'] - merged['k_V_ERA5']
    merged['ΔΔm (M2-E5)'] = merged['Δm Tel Opt_MERRA2'] - merged['Δm Tel Opt_ERA5']

    print(f"\n  Observasi yang dibandingkan: {len(merged)}")

    # Statistik selisih
    print("\n  Selisih (MERRA-2 − ERA5):")
    for var, label in [('ΔRH (M2-E5)', 'RH (pp)'),
                       ('Δk_V (M2-E5)', 'k_V'),
                       ('ΔΔm (M2-E5)', 'Δm Tel')]:
        vals = merged[var].dropna()
        print(f"    {label:10s}: mean={vals.mean():+.3f}  std={vals.std():.3f}  "
              f"range=[{vals.min():+.3f}, {vals.max():+.3f}]")

    # Korelasi Δm
    from scipy import stats as sp_stats

    dm_era5 = merged['Δm Tel Opt_ERA5'].values
    dm_merra2 = merged['Δm Tel Opt_MERRA2'].values

    # Filter valid pairs
    valid = (dm_era5 > -90) & (dm_merra2 > -90)
    dm_e = dm_era5[valid]
    dm_m = dm_merra2[valid]

    if len(dm_e) >= 3:
        r_pearson, p_pearson = sp_stats.pearsonr(dm_e, dm_m)
        r_spearman, p_spearman = sp_stats.spearmanr(dm_e, dm_m)
        print(f"\n  Korelasi Δm_Tel:")
        print(f"    Pearson  r = {r_pearson:.4f}  (p = {p_pearson:.4e})")
        print(f"    Spearman ρ = {r_spearman:.4f}  (p = {p_spearman:.4e})")
    else:
        r_pearson = r_spearman = float('nan')
        print("\n  [!] Terlalu sedikit data valid untuk korelasi.")

    # Confusion matrix MERRA-2
    print("\n  Confusion Matrix MERRA-2 (F optimal yang sama):")
    obs_vals = merged['Obs (Y/N)'].values
    cocok_m2 = merged['Cocok?_MERRA2'].values

    tp_m2 = sum((obs_vals == 'Y') & (cocok_m2 == '✓'))
    fn_m2 = sum((obs_vals == 'Y') & (cocok_m2 != '✓'))
    tn_m2 = sum((obs_vals == 'N') & (cocok_m2 == '✓'))
    fp_m2 = sum((obs_vals == 'N') & (cocok_m2 != '✓'))

    # ERA5 confusion matrix untuk perbandingan
    cocok_e5 = merged['Cocok?_ERA5'].values
    tp_e5 = sum((obs_vals == 'Y') & (cocok_e5 == '✓'))
    fn_e5 = sum((obs_vals == 'Y') & (cocok_e5 != '✓'))
    tn_e5 = sum((obs_vals == 'N') & (cocok_e5 == '✓'))
    fp_e5 = sum((obs_vals == 'N') & (cocok_e5 != '✓'))

    n_total = len(merged)
    n_y = sum(obs_vals == 'Y')
    n_n = sum(obs_vals == 'N')

    def calc_metrics(tp, fn, tn, fp):
        n = tp + fn + tn + fp
        acc = (tp + tn) / n if n > 0 else 0
        sens = tp / (tp + fn) if (tp + fn) > 0 else 0
        spec = tn / (tn + fp) if (tn + fp) > 0 else 0
        ba = (sens + spec) / 2
        denom_mcc = math.sqrt((tp+fp)*(tp+fn)*(tn+fp)*(tn+fn))
        mcc = (tp*tn - fp*fn) / denom_mcc if denom_mcc > 0 else 0
        return acc, sens, spec, ba, mcc

    acc_e5, sens_e5, spec_e5, ba_e5, mcc_e5 = calc_metrics(tp_e5, fn_e5, tn_e5, fp_e5)
    acc_m2, sens_m2, spec_m2, ba_m2, mcc_m2 = calc_metrics(tp_m2, fn_m2, tn_m2, fp_m2)

    print(f"                  ERA5      MERRA-2")
    print(f"    TP          : {tp_e5:5d}      {tp_m2:5d}")
    print(f"    FN          : {fn_e5:5d}      {fn_m2:5d}")
    print(f"    TN          : {tn_e5:5d}      {tn_m2:5d}")
    print(f"    FP          : {fp_e5:5d}      {fp_m2:5d}")
    print(f"    Accuracy    : {acc_e5:.1%}     {acc_m2:.1%}")
    print(f"    Sensitivity : {sens_e5:.1%}     {sens_m2:.1%}")
    print(f"    Specificity : {spec_e5:.1%}     {spec_m2:.1%}")
    print(f"    Bal. Acc.   : {ba_e5:.1%}     {ba_m2:.1%}")
    print(f"    MCC         : {mcc_e5:+.3f}     {mcc_m2:+.3f}")

    # Interpretasi
    dm_diff_std = merged['ΔΔm (M2-E5)'].std()
    if dm_diff_std > 0.3:
        print(f"\n  → Hasil BERBEDA signifikan (σ_ΔΔm = {dm_diff_std:.3f}):")
        print("    data atmosfer merupakan bottleneck utama")
    else:
        print(f"\n  → Hasil CUKUP KONSISTEN (σ_ΔΔm = {dm_diff_std:.3f}):")
        print("    model robust terhadap sumber data atmosfer")

    # Prediksi diskordansi (obs yang berubah status)
    discordant = merged[cocok_e5 != cocok_m2]
    if len(discordant) > 0:
        print(f"\n  Observasi dengan prediksi berbeda ({len(discordant)}):")
        for _, d in discordant.iterrows():
            print(f"    #{int(d['No']):2d} {d['Lokasi'][:25]:25s} "
                  f"ERA5={'✓' if d['Cocok?_ERA5']=='✓' else '✗'} "
                  f"MERRA2={'✓' if d['Cocok?_MERRA2']=='✓' else '✗'} "
                  f"ΔRH={d['ΔRH (M2-E5)']:+.1f}pp")

    # Output dataframe
    output_df = merged[['No', 'Lokasi', 'Obs (Y/N)',
                        'RH (%)_ERA5', 'RH (%)_MERRA2', 'ΔRH (M2-E5)',
                        'k_V_ERA5', 'k_V_MERRA2', 'Δk_V (M2-E5)',
                        'Δm Tel Opt_ERA5', 'Δm Tel Opt_MERRA2', 'ΔΔm (M2-E5)',
                        'Cocok?_ERA5', 'Cocok?_MERRA2']].copy()

    # Tambah baris ringkasan korelasi
    summary_rows = pd.DataFrame([{
        'No': '', 'Lokasi': 'KORELASI',
        'Obs (Y/N)': '',
        'RH (%)_ERA5': '', 'RH (%)_MERRA2': '',
        'ΔRH (M2-E5)': merged['ΔRH (M2-E5)'].mean(),
        'k_V_ERA5': '', 'k_V_MERRA2': '',
        'Δk_V (M2-E5)': merged['Δk_V (M2-E5)'].mean(),
        'Δm Tel Opt_ERA5': f'r={r_pearson:.3f}',
        'Δm Tel Opt_MERRA2': f'ρ={r_spearman:.3f}',
        'ΔΔm (M2-E5)': merged['ΔΔm (M2-E5)'].mean(),
        'Cocok?_ERA5': f'{acc_e5:.1%}',
        'Cocok?_MERRA2': f'{acc_m2:.1%}',
    }])
    output_df = pd.concat([output_df, summary_rows], ignore_index=True)

    return output_df


# ═══════════════════════════════════════════════════════════════════
# OUTPUT: EXCEL
# ═══════════════════════════════════════════════════════════════════

def save_results(sensitivity_df, critical_df, decomp_df,
                 error_bar_df, merra2_df, filepath):
    """Simpan semua hasil ke Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    sty = {
        'hdr_font': Font(name='Arial', bold=True, size=10, color='FFFFFF'),
        'hdr_fill': PatternFill('solid', fgColor='1F4E79'),
        'data_font': Font(name='Arial', size=10),
        'center': Alignment(horizontal='center', vertical='center'),
        'left': Alignment(horizontal='left', vertical='center'),
        'border': Border(
            left=Side('thin', 'B0B0B0'), right=Side('thin', 'B0B0B0'),
            top=Side('thin', 'B0B0B0'), bottom=Side('thin', 'B0B0B0')),
        'fn_fill': PatternFill('solid', fgColor='FFC7CE'),
        'tn_fill': PatternFill('solid', fgColor='C6EFCE'),
        'highlight': PatternFill('solid', fgColor='FFEB9C'),
    }

    def write_df_to_sheet(ws, df, sheet_name):
        ws.title = sheet_name
        headers = list(df.columns)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = sty['hdr_font']
            c.fill = sty['hdr_fill']
            c.alignment = sty['center']
            c.border = sty['border']

        for ri, (_, row) in enumerate(df.iterrows(), 2):
            for ci, col in enumerate(headers, 1):
                val = row[col]
                if isinstance(val, float) and not math.isnan(val):
                    if abs(val) > 1e6:
                        val = f"{val:.4e}"
                    elif abs(val) < 0.001 and val != 0:
                        val = f"{val:.6f}"
                    else:
                        val = round(val, 4)
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = sty['data_font']
                c.alignment = sty['center']
                c.border = sty['border']

                # Highlight FN rows
                if 'Tipe' in headers:
                    tipe_ci = headers.index('Tipe') + 1
                    tipe_val = ws.cell(row=ri, column=tipe_ci).value
                    if tipe_val == 'FN':
                        c.fill = sty['fn_fill']
                    elif tipe_val == 'FP':
                        c.fill = sty['highlight']

        for ci in range(1, len(headers) + 1):
            max_len = max(len(str(headers[ci-1])),
                         max((len(str(ws.cell(row=r, column=ci).value or ''))
                              for r in range(2, min(len(df)+2, 50))), default=8))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 25)
        ws.freeze_panes = 'A2'

    # Sheet 1: Sensitivitas RH (pivot)
    pivot_rows = []
    for obs_no in sorted(sensitivity_df['No'].unique()):
        obs_data = sensitivity_df[sensitivity_df['No'] == obs_no]
        base_row = obs_data[obs_data['ΔRH'] == 0]
        if len(base_row) == 0:
            continue
        base = base_row.iloc[0]
        prow = {
            'No': obs_no,
            'Lokasi': base['Lokasi'],
            'Obs': base['Obs'],
            'Tipe': base['Tipe'],
            'RH_base': base['RH_baseline'],
            'k_V_base': base['k_V'],
        }
        for delta in RH_DELTAS:
            d = obs_data[obs_data['ΔRH'] == delta]
            if len(d) > 0:
                prow[f'Δm_Tel(ΔRH={delta:+d})'] = round(d.iloc[0]['Δm_Tel'], 3)
                prow[f'k_V(ΔRH={delta:+d})'] = round(d.iloc[0]['k_V'], 4)
        pivot_rows.append(prow)

    ws1 = wb.active
    write_df_to_sheet(ws1, pd.DataFrame(pivot_rows), "4.5.1 Sensitivitas RH")

    # Sheet 2: RH Kritis + Kategorisasi FN
    ws2 = wb.create_sheet()
    write_df_to_sheet(ws2, critical_df, "4.5.1 RH Kritis + Kat FN")

    # Sheet 3: Dekomposisi Error
    ws3 = wb.create_sheet()
    if len(decomp_df) > 0:
        write_df_to_sheet(ws3, decomp_df, "4.5.2 Dekomposisi Jalur")
    else:
        ws3.title = "4.5.2 Dekomposisi Jalur"
        ws3.cell(row=1, column=1, value="Tidak ada FN/FP untuk didekomposisi")

    # Sheet 4: Error Bar
    ws4 = wb.create_sheet()
    write_df_to_sheet(ws4, error_bar_df, "4.5.3 Error Bar ERA5")

    # Sheet 5: ERA5 vs MERRA-2
    ws5 = wb.create_sheet()
    if merra2_df is not None and len(merra2_df) > 0:
        write_df_to_sheet(ws5, merra2_df, "4.5.4 ERA5 vs MERRA-2")
    else:
        ws5.title = "4.5.4 ERA5 vs MERRA-2"
        ws5.cell(row=1, column=1, value="Data MERRA-2 tidak tersedia")

    # Sheet 6: Sensitivitas Detail (raw data)
    ws6 = wb.create_sheet()
    write_df_to_sheet(ws6, sensitivity_df, "Data Detail Sensitivitas")

    # Save
    os.makedirs(os.path.dirname(filepath) or '.', exist_ok=True)
    wb.save(filepath)
    print(f"\n  Saved: {filepath}")


# ═══════════════════════════════════════════════════════════════════
# OUTPUT: PLOT
# ═══════════════════════════════════════════════════════════════════

def plot_sensitivity(sensitivity_df: pd.DataFrame, critical_df: pd.DataFrame,
                     filepath: str):
    """Plot sensitivitas Δm vs ΔRH untuk observasi FN dan FP."""
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
    except ImportError:
        print("  [!] matplotlib tidak tersedia.")
        return

    # Plot FN dan FP (sesuai arsitektur: "terutama FN dan FP")
    misclass = sensitivity_df[sensitivity_df['Tipe'].isin(['FN', 'FP'])]
    if len(misclass) == 0:
        print("  [!] Tidak ada data FN/FP untuk diplot.")
        return

    fig, axes = plt.subplots(1, 2, figsize=(16, 7))

    # ── Plot 1: Δm_Tel vs ΔRH untuk setiap FN/FP ──
    ax1 = axes[0]
    obs_list = sorted(misclass['No'].unique())
    colors = plt.cm.tab10(range(len(obs_list)))

    for i, obs_no in enumerate(obs_list):
        obs_data = misclass[misclass['No'] == obs_no].sort_values('ΔRH')
        tipe = obs_data.iloc[0]['Tipe']
        label = f"#{obs_no} {obs_data.iloc[0]['Lokasi'][:20]} ({tipe})"
        marker = '*' if tipe == 'FN' else 's'
        ax1.plot(obs_data['ΔRH'], obs_data['Δm_Tel'],
                 marker=marker, markersize=6, color=colors[i],
                 linewidth=2, label=label)

    ax1.axhline(y=0, color='green', linewidth=2, linestyle='--',
                label='Threshold (Δm=0)', alpha=0.8)
    ax1.axvline(x=0, color='gray', linewidth=1, linestyle=':', alpha=0.5)

    # Zona kategorisasi FN
    ax1.axvspan(-5, 0, alpha=0.05, color='green', label='Zona A (near-miss)')
    ax1.axvspan(-20, -5, alpha=0.05, color='orange', label='Zona B (correctable)')

    ax1.set_xlabel('Perubahan RH (pp)', fontsize=12, fontweight='bold')
    ax1.set_ylabel('Δm Teleskop (mag)', fontsize=12, fontweight='bold')
    ax1.set_title('4.5.1  Sensitivitas Δm terhadap Perubahan RH\n'
                  '(Observasi Misclassified: FN/FP)',
                  fontsize=13, fontweight='bold')
    ax1.legend(fontsize=7, loc='best')
    ax1.grid(True, alpha=0.3)

    # ── Plot 2: k_V vs Δm untuk semua observasi ──
    ax2 = axes[1]
    all_base = sensitivity_df[sensitivity_df['ΔRH'] == 0]
    for tipe, color, marker, sz in [('TN', 'blue', 'o', 50),
                                     ('TP', 'green', 'D', 70),
                                     ('FN', 'red', '*', 100),
                                     ('FP', 'orange', 's', 70)]:
        sub = all_base[all_base['Tipe'] == tipe]
        if len(sub) > 0:
            ax2.scatter(sub['k_V'], sub['Δm_Tel'], c=color, s=sz, alpha=0.8,
                        label=f'{tipe} (n={len(sub)})', edgecolors='black',
                        linewidths=0.5, marker=marker, zorder=5 if tipe in ('FN','FP') else 3)

    ax2.axhline(y=0, color='green', linewidth=2, linestyle='--', alpha=0.8)
    ax2.set_xlabel('Koefisien Ekstingsi k_V', fontsize=12, fontweight='bold')
    ax2.set_ylabel('Δm Teleskop (mag)', fontsize=12, fontweight='bold')
    ax2.set_title('Korelasi k_V dengan Δm\n(Semua Observasi)',
                  fontsize=13, fontweight='bold')
    ax2.legend(fontsize=10)
    ax2.grid(True, alpha=0.3)

    plt.tight_layout()
    fig.savefig(filepath, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved: {filepath}")


def plot_decomposition(decomp_df: pd.DataFrame, filepath: str):
    """Plot dekomposisi kontribusi error (jalur Kastner vs Schaefer)."""
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
    except ImportError:
        return

    if len(decomp_df) == 0:
        return

    fig, ax = plt.subplots(figsize=(12, 6))

    labels = [f"#{int(r['No'])} {r['Lokasi'][:18]}" for _, r in decomp_df.iterrows()]
    x = range(len(labels))
    w = 0.25

    bars_L = [r['Δm_dari_L (Kastner)'] for _, r in decomp_df.iterrows()]
    bars_B = [r['Δm_dari_B (Schaefer)'] for _, r in decomp_df.iterrows()]
    bars_I = [r['Δm_interaksi'] for _, r in decomp_df.iterrows()]

    ax.bar([i - w for i in x], bars_L, w, label='Luminansi (Kastner)',
           color='#E74C3C', edgecolor='black', linewidth=0.5)
    ax.bar(x, bars_B, w, label='Sky Brightness (Schaefer)',
           color='#3498DB', edgecolor='black', linewidth=0.5)
    ax.bar([i + w for i in x], bars_I, w, label='Interaksi',
           color='#95A5A6', edgecolor='black', linewidth=0.5)

    ax.set_xlabel('Observasi Misclassified', fontsize=12, fontweight='bold')
    ax.set_ylabel('Kontribusi ke Δ(Δm) [mag]', fontsize=12, fontweight='bold')
    ax.set_title('4.5.2  Dekomposisi Perubahan Δm saat RH Dikurangi 20pp\n'
                 'Jalur Kastner (luminansi) vs Schaefer (sky brightness)',
                 fontsize=13, fontweight='bold')
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=30, ha='right', fontsize=9)
    ax.legend(fontsize=10)
    ax.grid(True, axis='y', alpha=0.3)
    ax.axhline(y=0, color='black', linewidth=0.5)

    plt.tight_layout()
    fig.savefig(filepath, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved: {filepath}")


def plot_era5_vs_merra2(df_era5: pd.DataFrame, df_merra2: pd.DataFrame, filepath: str):
    """Scatter plot Δm(ERA5) vs Δm(MERRA-2) sesuai arsitektur 4.5.4."""
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
    except ImportError:
        print("  [!] matplotlib tidak tersedia.")
        return

    merged = pd.merge(
        df_era5[['No', 'Lokasi', 'Obs (Y/N)', 'Δm Tel Opt']],
        df_merra2[['No', 'Δm Tel Opt']],
        on='No', suffixes=('_ERA5', '_MERRA2')
    )

    valid = (merged['Δm Tel Opt_ERA5'] > -90) & (merged['Δm Tel Opt_MERRA2'] > -90)
    merged = merged[valid]

    fig, axes = plt.subplots(1, 2, figsize=(16, 7))

    # ── Plot 1: Scatter Δm ──
    ax1 = axes[0]
    y_obs = merged[merged['Obs (Y/N)'] == 'Y']
    n_obs = merged[merged['Obs (Y/N)'] == 'N']

    ax1.scatter(n_obs['Δm Tel Opt_ERA5'], n_obs['Δm Tel Opt_MERRA2'],
                c='blue', s=50, alpha=0.7, label=f'N (n={len(n_obs)})',
                edgecolors='black', linewidths=0.5)
    ax1.scatter(y_obs['Δm Tel Opt_ERA5'], y_obs['Δm Tel Opt_MERRA2'],
                c='red', s=80, alpha=0.9, label=f'Y (n={len(y_obs)})',
                edgecolors='black', linewidths=0.5, marker='*', zorder=5)

    # Diagonal line (perfect agreement)
    all_dm = list(merged['Δm Tel Opt_ERA5']) + list(merged['Δm Tel Opt_MERRA2'])
    lims = [min(all_dm) - 0.2, max(all_dm) + 0.2]
    ax1.plot(lims, lims, 'k--', alpha=0.5, label='1:1 line')
    ax1.axhline(y=0, color='green', linewidth=1, linestyle=':', alpha=0.5)
    ax1.axvline(x=0, color='green', linewidth=1, linestyle=':', alpha=0.5)

    ax1.set_xlabel('Δm Teleskop (ERA5)', fontsize=12, fontweight='bold')
    ax1.set_ylabel('Δm Teleskop (MERRA-2)', fontsize=12, fontweight='bold')
    ax1.set_title('4.5.4  Δm(ERA5) vs Δm(MERRA-2)', fontsize=13, fontweight='bold')
    ax1.legend(fontsize=10)
    ax1.grid(True, alpha=0.3)
    ax1.set_aspect('equal')

    # ── Plot 2: Selisih RH ──
    ax2 = axes[1]
    merged_rh = pd.merge(
        df_era5[['No', 'Lokasi', 'Obs (Y/N)', 'RH (%)']],
        df_merra2[['No', 'RH (%)']],
        on='No', suffixes=('_ERA5', '_MERRA2')
    )
    merged_rh['ΔRH'] = merged_rh['RH (%)_MERRA2'] - merged_rh['RH (%)_ERA5']

    ax2.bar(range(len(merged_rh)), merged_rh['ΔRH'],
            color=['red' if v > 0 else 'blue' for v in merged_rh['ΔRH']],
            edgecolor='black', linewidth=0.5, alpha=0.7)
    ax2.set_xlabel('Observasi (No)', fontsize=12, fontweight='bold')
    ax2.set_ylabel('ΔRH = MERRA-2 − ERA5 (pp)', fontsize=12, fontweight='bold')
    ax2.set_title('Selisih RH antara MERRA-2 dan ERA5\nper Observasi',
                  fontsize=13, fontweight='bold')
    ax2.set_xticks(range(len(merged_rh)))
    ax2.set_xticklabels([f"#{int(n)}" for n in merged_rh['No']],
                        rotation=90, fontsize=8)
    ax2.axhline(y=0, color='black', linewidth=1)
    ax2.grid(True, axis='y', alpha=0.3)

    plt.tight_layout()
    fig.savefig(filepath, dpi=150, bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved: {filepath}")


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════

def main():
    print("\n" + "█" * 70)
    print("  ANALISIS DIAGNOSTIK MODEL CRUMEY (2014)")
    print("  Tahap 4.5: Analisis Sensitivitas dan Ketidakpastian")
    print("█" * 70)

    # ── Baca input ERA5 ──
    if not os.path.exists(INPUT_ERA5):
        print(f"\n  [!] File tidak ditemukan: {INPUT_ERA5}")
        print("  Jalankan batch_validation_crumey.py terlebih dahulu.")
        return

    print(f"\n  Membaca ERA5: {INPUT_ERA5}")
    df_era5 = load_observation_data(INPUT_ERA5)
    print(f"  {len(df_era5)} observasi dimuat.")

    # ── Baca input MERRA-2 (opsional) ──
    df_merra2 = None
    if os.path.exists(INPUT_MERRA2):
        print(f"  Membaca MERRA-2: {INPUT_MERRA2}")
        df_merra2 = load_merra2_data(INPUT_MERRA2)
        print(f"  {len(df_merra2)} observasi dimuat.")
    else:
        print(f"  [!] File MERRA-2 tidak ditemukan: {INPUT_MERRA2}")
        print("      Tahap 4.5.4 akan dilewati.")

    # ── 4.5.1: Sensitivitas OAT terhadap RH ──
    sensitivity_df, critical_df = analyze_rh_sensitivity(df_era5)

    # ── 4.5.2: Dekomposisi Jalur Error ──
    decomp_df = analyze_error_decomposition(df_era5)

    # ── 4.5.3: Error Bar dari Ketidakpastian ERA5 ──
    error_bar_df = analyze_error_bars(sensitivity_df, critical_df)

    # ── 4.5.4: Perbandingan ERA5 vs MERRA-2 ──
    merra2_df = None
    if df_merra2 is not None:
        merra2_df = analyze_era5_vs_merra2(df_era5, df_merra2)

    # ── Ringkasan Akhir ──
    print("\n" + "█" * 70)
    print("  RINGKASAN TEMUAN DIAGNOSTIK (4.5)")
    print("█" * 70)

    fn_critical = critical_df[critical_df['Tipe'] == 'FN']
    n_fn = len(fn_critical)
    fp_critical = critical_df[critical_df['Tipe'] == 'FP']
    n_fp = len(fp_critical)

    print(f"\n  False Negative (FN): {n_fn}")
    if n_fn > 0:
        for cat in ['A', 'B', 'C']:
            n_cat = sum(fn_critical['Kategori_FN'] == cat)
            print(f"    Kategori {cat}: {n_cat}/{n_fn}")
    print(f"  False Positive (FP): {n_fp}")

    if len(decomp_df) > 0:
        mean_L = decomp_df['Kontribusi_L (%)'].mean()
        mean_B = decomp_df['Kontribusi_B (%)'].mean()
        print(f"\n  Jalur error dominan (ΔRH=-20pp):")
        print(f"    Kastner (luminansi) : {mean_L:.1f}%")
        print(f"    Schaefer (sky bright): {mean_B:.1f}%")

    # ── Simpan ──
    excel_path = os.path.join(OUTPUT_DIR, "Analisis_Diagnostik_Crumey.xlsx")
    save_results(sensitivity_df, critical_df, decomp_df,
                 error_bar_df, merra2_df, excel_path)

    plot_path = os.path.join(OUTPUT_DIR, "Sensitivitas_RH_dan_kV.png")
    plot_sensitivity(sensitivity_df, critical_df, plot_path)

    decomp_plot_path = os.path.join(OUTPUT_DIR, "Dekomposisi_Jalur_Error.png")
    plot_decomposition(decomp_df, decomp_plot_path)

    if df_merra2 is not None:
        merra2_plot_path = os.path.join(OUTPUT_DIR, "ERA5_vs_MERRA2.png")
        plot_era5_vs_merra2(df_era5, df_merra2, merra2_plot_path)

    print(f"\n{'█' * 70}")
    print("  ANALISIS 4.5 SELESAI")
    print(f"{'█' * 70}")
    print(f"  Excel       : {excel_path}")
    print(f"  Plot 4.5.1  : {plot_path}")
    print(f"  Plot 4.5.2  : {decomp_plot_path}")
    if df_merra2 is not None:
        print(f"  Plot 4.5.4  : {os.path.join(OUTPUT_DIR, 'ERA5_vs_MERRA2.png')}")
    print(f"{'█' * 70}\n")


if __name__ == "__main__":
    main()
